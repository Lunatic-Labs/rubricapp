import openpyxl
from openpyxl import load_workbook
# from NamedAtomicLock import NamedAtomicLock
#
# myLock = NamedAtomicLock('./test1.xlsx')
#
# if myLock.acquire(timeout=15):
#
#     workbook = openpyxl.load_workbook("./test1.xlsx")
#     worksheet = workbook["Sheet1"]
#     i = 0
#     while i < 100000:
#         worksheet.append(["1", "2"])
#         i += 1
#     workbook.save("./test1.xlsx")
#     myLock.release()
from filelock import Timeout, FileLock

lock = FileLock("./test1.xlsx.lock", timeout=10)
with lock:
    workbook = openpyxl.load_workbook("./test1.xlsx")
    worksheet = workbook["Sheet1"]
    i = 0
    while i < 100000:
        worksheet.append(["1", "2"])
        i += 1
    workbook.save("./test1.xlsx")

