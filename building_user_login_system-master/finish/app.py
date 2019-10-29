from flask import Flask, render_template, redirect, url_for, request, send_file
from flask_bootstrap import Bootstrap
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, BooleanField
from flask_wtf.file import FileField, FileAllowed, FileRequired
from wtforms.validators import InputRequired, Email, Length, ValidationError
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash

from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
import os
import subprocess
# shutil used to delete whole directory(folder)
import shutil
import openpyxl
from openpyxl import load_workbook
import datetime
import json
import sys
from fpdf import FPDF, HTMLMixin

files_dir = None
if len(sys.argv) > 1:
    files_dir = sys.argv[1]
else:
    print(
        "Requires argument: path to put files and database (suggestion is `pwd` when already in directory containing app.py)")
    sys.exit(1)

app = Flask(__name__)
app.config['SECRET_KEY'] = 'Thisissupposedtobesecret!'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///{}/account.db'.format(files_dir)
bootstrap = Bootstrap(app)
db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Settings of Directory ======================================================================================================

# SET THE BASE DIRECTORY
os.chdir(files_dir)
base_directory = os.getcwd()


class HTML2PDF(FPDF, HTMLMixin):
    pass


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(15), unique=True, nullable=False)
    email = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(80), nullable=False)
    role = db.Column(db.String(20), nullable=True)
    University = db.Column(db.String(50), nullable=True)
    description = db.Column(db.String(255), nullable=True)


class Permission(UserMixin, db.Model):
    project_id = db.Column(db.String(50), primary_key=True)
    owner = db.Column(db.String(50), nullable=False)
    shareTo = db.Column(db.String(50), nullable=False)
    project = db.Column(db.String(30), nullable=False)
    status = db.Column(db.String(30), nullable=False)


class Project(UserMixin, db.Model):
    project_name = db.Column(db.String(30), primary_key=True)
    owner = db.Column(db.String(50), primary_key=True)
    project_status = db.Column(db.String(10), nullable=False)
    description = db.Column(db.String(255), nullable=True)


class Evaluation(UserMixin, db.Model):
    eva_name = db.Column(db.String(30), primary_key=True)
    project_name = db.Column(db.String(30), primary_key=True)
    project_owner = db.Column(db.String(50), primary_key=True)
    owner = db.Column(db.String(50), nullable=False)
    description = db.Column(db.String(255), nullable=True)
    last_edit = db.Column(db.String(15), nullable=True)


class Notification(UserMixin, db.Model):
    notification_id = db.Column(db.Integer, primary_key=True)
    from_user = db.Column(db.String(50), nullable=False)
    to_user = db.Column(db.String(50), nullable=False)
    message_type = db.Column(db.String(50), nullable=False)
    message_content = db.Column(db.String(255), nullable=True)
    status = db.Column(db.String(50), nullable=False)
    time = db.Column(db.String(50), nullable=False)
    appendix = db.Column(db.String(255), nullable=True)


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


class LoginForm(FlaskForm):
    username = StringField('username', validators=[InputRequired(), Length(min=4, max=15)])
    password = PasswordField('password', validators=[InputRequired(), Length(min=8, max=80)])
    remember = BooleanField('remember me')


class RegisterForm(FlaskForm):
    email = StringField('email', validators=[InputRequired(), Email(message='Invalid email'), Length(max=50)])
    username = StringField('username', validators=[InputRequired(), Length(min=4, max=15)])
    password = PasswordField('password', validators=[InputRequired(), Length(min=8, max=80)])
    # instructor = BooleanField('instructor')


@login_required
def project_name_validator(form, field):
    duplicate_project_name = Project.query.filter_by(project_name=field.data, owner=current_user.username).first()
    if duplicate_project_name is not None:
        raise ValidationError("The project name has been used before")


@login_required
def project_student_file_validator(form, field):
    path_to_current_user = "{}/{}".format(base_directory, current_user.username)
    path_to_student_file_stored = "{}/".format(path_to_current_user)
    student_file_filename = "student.xlsx"
    field.data.save(path_to_student_file_stored + student_file_filename)
    student_file_workbook = load_workbook(path_to_student_file_stored + student_file_filename)
    student_file_worksheet = student_file_workbook['Sheet1']
    find_group = True if 'group' in [x.value for x in list(student_file_worksheet.iter_rows())[0]] else False
    if find_group is False:
        # os.remove(path_to_student_file_stored)
        raise ValidationError("Can not find group information")
    # os.remove(path_to_student_file_stored+student_file_filename)


@login_required
def project_json_file_validator(form, field):
    path_to_current_user = "{}/{}".format(base_directory, current_user.username)
    path_to_json_file_stored = "{}/".format(path_to_current_user)
    json_file_filename = "TW.json"
    field.data.save(path_to_json_file_stored + json_file_filename)
    with open(path_to_json_file_stored + json_file_filename, 'r')as f:
        json_data = json.loads(f.read(), strict=False)

    if 'name' in json_data.keys() and 'category' in json_data.keys():
        for category in json_data['category']:
            if 'name' in category.keys() and 'section' in category.keys():
                category_name = (category['name'])
                for section in category['section']:
                    if 'name' in section.keys() and 'type' in section.keys() and 'values' in section.keys():
                        for value in section['values']:
                            if 'name' not in value.keys() or 'desc' not in value.keys():
                                raise ValidationError("lack of NAME or DESC in json file")

                    else:
                        raise ValidationError("lack of NAME or TYPE or VALUES in json file")
            else:
                raise ValidationError("lack of NAME or SECTIONS in json file")
    else:
        raise ValidationError("lack of NAME or CATEGORY in json file")
    # os.remove(path_to_json_file_stored+ json_file_filename)


class ProjectForm(FlaskForm):
    project_name = StringField('project name',
                               validators=[InputRequired(), Length(min=3, max=10), project_name_validator])
    project_description = StringField('description', validators=[InputRequired(), Length(min=0, max=255)])
    # group_file = FileField('group file',validators = [FileRequired(),FileAllowed(JSON, 'Json only')]
    # group_file = FileField('group file')
    student_file = FileField('students file', validators=[project_student_file_validator])
    grading_criteria = FileField('grading criteria', validators=[project_json_file_validator])


# class EmailForm(FlaskForm):
#     email = StringField('email', validators=[InputRequired(), Email(message='Invalid email'), Length(max=50)])
#     password = PasswordField('password', validators=[InputRequired(), Length(min=1, max=80)])

# class EvaluationForm(FlaskForm):
#     evaluation_name = StringField('evaluation name', validators=[InputRequired(), Length(min=3, max=10)])

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()

    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user:
            if check_password_hash(user.password, form.password.data):
                login_user(user, remember=form.remember.data)
                # instructor jump to instructor page, student jump to student page
                # if(user.instructor == "1"):
                return redirect(url_for('instructor_dashboard'))
                # else:
                #     return redirect(url_for('student_dashboard'))
            else:
                return render_template('login.html', msg="password not correct", form=form)
        else:
            return render_template('login.html', msg="user doesn't exist", form=form)
        # return '<h1>' + form.username.data + ' ' + form.password.data + '</h1>'

    return render_template('login.html', msg="", form=form)


@app.route('/signup', methods=['GET', 'POST'])
def signup():
    form = RegisterForm()
    if form.validate_on_submit():
        # check if the user and email has existed in the database
        check_username = User.query.filter_by(username=form.username.data).first()
        check_email = User.query.filter_by(email=form.email.data).first()
        if check_username:
            return render_template('signup.html', form=form, msg="Warning !!! The username has existed")
        elif check_email:
            return render_template('signup.html', form=form, msg="Warning !!! The email has been used")
        else:
            hashed_password = generate_password_hash(form.password.data, method='sha256')
            new_user = User(username=form.username.data, email=form.email.data, password=hashed_password)
            db.session.add(new_user)
            db.session.commit()

            # After making sure that the new user is created, the user's private folder can be created by using the user name

            path_to_user_folder = "{}/{}".format(base_directory, new_user.username)
            os.mkdir(path_to_user_folder)

            return redirect(url_for('login'))
            # return '<h1>' + form.username.data + ' ' + form.email.data + ' ' + form.password.data + '</h1>'

    return render_template('signup.html', form=form, msg="")


@app.route('/instructor_dashboard')
@login_required
def instructor_dashboard():
    # Load all projects to instructor_dashboard
    # Find all projects in User's private folder by using current user

    path_to_current_user = "{}/{}".format(base_directory, current_user.username)
    project_list = []
    for project in os.listdir(path_to_current_user):
        project_list.append(project)
    project_len = len(project_list)

    return render_template('instructor_dashboard.html', name=current_user.username, project_list=project_list,
                           project_len=project_len)


@app.route('/project_profile_jumptool', methods=["POST", "GET"])
@login_required
def project_profile_jumptool():
    # a jump tool before load project
    # display the information of projects (title and desc) and its recent evaluations

    path_to_current_user = "{}/{}".format(base_directory, current_user.username)
    # list of evaluation & list of groups relating to one project
    project_set_map = {}
    project_list = Permission.query.filter_by(owner=current_user.username, shareTo=current_user.username).all()
    project_information_map = {}
    for project in project_list:
        path_to_evaluation_file = "{}/{}/{}/evaluation.xlsx".format(base_directory, current_user.username,
                                                                    project.project)
        evaluation_workbook = openpyxl.load_workbook(path_to_evaluation_file)
        evaluation_worksheet = evaluation_workbook['eva']
        group_worksheet = evaluation_workbook['group']
        group_col = []
        for col_item in list(group_worksheet.iter_cols())[0]:
            if col_item.value != "groupid":
                group_col.append(col_item.value)
        set_of_eva = Evaluation.query.filter_by(project_name=project.project, project_owner=current_user.username).all()
        project_set_map[project.project] = (group_col, set_of_eva)
        project_information_map[project.project] = Project.query.filter_by(project_name=project.project,
                                                                           owner=project.owner).first()

    return render_template("project_profile_jumptool.html", project_list=project_list, project_set_map=project_set_map,
                           project_information_map=project_information_map)


@app.route('/project_profile/<string:project_id>', methods=["POST", "GET"])
@login_required
def project_profile(project_id):
    # show each grade in this project and divided into eva s
    project = Permission.query.filter_by(project_id=project_id).first()
    path_to_current_user = "{}/{}".format(base_directory, current_user.username)
    path_to_current_project = "{}/{}/{}".format(base_directory, current_user.username, project.project)

    list_of_shareTo_permission = [x for x in
                                  Permission.query.filter_by(project=project.project, owner=current_user.username).all()
                                  if x.shareTo != current_user.username]

    path_to_evaluation_file = "{}/{}/{}/evaluation.xlsx".format(base_directory, current_user.username, project.project)
    evaluation_workbook = openpyxl.load_workbook(path_to_evaluation_file)
    evaluation_worksheet = evaluation_workbook['eva']
    group_worksheet = evaluation_workbook['group']
    list_of_eva = select_by_col_name('eva_name', evaluation_worksheet)
    set_of_eva = set(list_of_eva)
    # get all groups and its owners
    dic_of_eva = {}

    for eva in set_of_eva:
        # update 9/13: simple profile
        dic_of_eva[eva] = []
        temp_eva = select_row_by_group_id("eva_name", eva, evaluation_worksheet)
        for eva_row in temp_eva:
            dic_of_eva[eva].append(eva_row)

    tags = [x.value for x in list(evaluation_worksheet.iter_rows())[0]]

    # group management
    management_groups = []
    rows_got_from_group_worksheet = list(group_worksheet.iter_rows())
    for row in rows_got_from_group_worksheet:
        management_groups.append([x.value for x in row])

    return render_template("project_profile.html", dic_of_eva=dic_of_eva,
                           list_of_shareTo_permission=list_of_shareTo_permission, management_groups=management_groups,
                           tags=tags, project=project, set_of_eva=list(set_of_eva))


@app.route('/project_profile_backed_up/<string:project_id>', methods=["POST", "GET"])
@login_required
def project_profile_backed_up(project_id):
    # prepare list of project and map of project like {project1: [eva1, eva2, eva3]}
    # project1:{eva1: [Ricky, Jason]}
    # project1: [Ricky, Json] (share to whom)
    project = Permission.query.filter_by(project_id=project_id).first()
    path_to_current_user = "{}/{}".format(base_directory, current_user.username)
    path_to_current_project = "{}/{}/{}".format(base_directory, current_user.username, project.project)

    list_of_shareTo_permission = [x for x in
                                  Permission.query.filter_by(project=project.project, owner=current_user.username).all()
                                  if x.shareTo != current_user.username]

    path_to_profile_json = "{}/{}/profile_json".format(path_to_current_user, project.project)
    if os.path.exists(path_to_profile_json):
        shutil.rmtree(path_to_profile_json)
    os.mkdir(path_to_profile_json)
    path_to_evaluation_file = "{}/{}/{}/evaluation.xlsx".format(base_directory, current_user.username, project.project)
    evaluation_workbook = openpyxl.load_workbook(path_to_evaluation_file)
    evaluation_worksheet = evaluation_workbook['eva']
    group_worksheet = evaluation_workbook['group']
    group_col = []
    for col_item in list(group_worksheet.iter_cols())[0]:
        if col_item.value != "groupid":
            group_col.append(col_item.value)
    list_of_eva = select_by_col_name('eva_name', evaluation_worksheet)
    set_of_eva = set(list_of_eva)
    # get all groups and its owners
    dic_of_eva = {}

    for eva in set_of_eva:
        # update 8/27: dump into json file
        dic_of_json = []
        start_index = 1
        # =================================================
        dic_of_eva[eva] = {}
        temp_eva = select_row_by_group_id("eva_name", eva, evaluation_worksheet)
        for group in group_col:
            dic_of_json.append({"id": start_index, "name": group, "parent_id": "0"})
            group_parent_id = start_index
            start_index += 1
            owners = set()
            for row in temp_eva:
                owners.add(row['owner'])
            for owner in owners:
                dic_of_json.append({"id": start_index, "name": owner, "parent_id": group_parent_id})
                owner_parent_id = start_index
                start_index += 1
                index_list_of_date = select_index_by_group_eva_owner(eva, group, owner, evaluation_worksheet)
                for date_index in index_list_of_date:
                    date = select_by_col_name("date", evaluation_worksheet)[date_index - 2]
                    dic_of_json.append({"id": start_index, "name": date, "parent_id": owner_parent_id})
                    start_index += 1
            dic_of_eva[eva][group] = owners

        path_to_this_json = "{}/{}_profile.json".format(path_to_profile_json, eva)
        with open(path_to_this_json, 'w') as f:  # writing JSON object
            json.dump(dic_of_json, f)

    return render_template("project_profile_backed_up.html", dic_of_eva=dic_of_eva,
                           list_of_shareTo_permission=list_of_shareTo_permission, project=project,
                           path_to_profile_json=path_to_profile_json, set_of_eva=list(set_of_eva))


@app.route('/management_group/<string:project_id>', methods=['GET', 'POST'])
@login_required
def managment_group(project_id):
    project = Permission.query.filter_by(project_id=project_id).first()
    path_to_evaluation_xlsx = "{}/{}/{}/evaluation.xlsx".format(base_directory, current_user.username, project.project)
    evaluation_workbook = openpyxl.load_workbook(path_to_evaluation_xlsx)
    group_worksheet = evaluation_workbook['group']
    for row_index in range(1, len(list(group_worksheet.iter_rows()))):
        for col_index in range(1, len(list(group_worksheet.iter_cols()))):
            student_email = request.form.get((list(group_worksheet.iter_cols())[0][row_index].value + str(col_index)),
                                             " ")

            # group management detector
            # if the given value is None, inserted should also be None
            if student_email == " " or student_email == "None":
                group_worksheet.cell(row_index + 1, col_index + 1).value = None
            else:
                group_worksheet.cell(row_index + 1, col_index + 1).value = student_email
    evaluation_workbook.save(path_to_evaluation_xlsx)
    return redirect(url_for("project_profile", project_id=project_id))


@app.route('/delete_eva/<string:project_id>/<string:evaluation>/<string:group>/<string:grader>/<string:datetime>',
           methods=['GET', 'POST'])
@login_required
def delete_eva(project_id, evaluation, group, grader, datetime):
    project = Permission.query.filter_by(project_id=project_id).first()
    path_to_evaluation_xlsx = "{}/{}/{}/evaluation.xlsx".format(base_directory, current_user.username, project.project)
    evaluation_workbook = openpyxl.load_workbook(path_to_evaluation_xlsx)
    evaluation_worksheet = evaluation_workbook['eva']
    group_worksheet = evaluation_workbook['group']
    allgroups = select_by_col_name('groupid', group_worksheet)
    students_worksheet = evaluation_workbook['students']

    index = int(select_index_by_group_eva_owner_date(evaluation, group, grader, datetime, evaluation_worksheet))
    evaluation_worksheet.delete_rows(index, 1)

    # check whether all group have at least one empty grade in this evaluation
    group_col_in_eva = set(select_by_col_name('group', evaluation_worksheet))
    empty_group = [x for x in allgroups if x not in group_col_in_eva]

    students = get_students_by_group(group_worksheet, students_worksheet)

    for empty in empty_group:
        students_name = []
        # couple is [email, student_name]
        for student_couple in students[str(group)]:
            students_name.append(student_couple[1])
        empty_row = new_row_generator(str(group), students_name, evaluation, evaluation_worksheet)
        evaluation_worksheet.append(empty_row)
    evaluation_workbook.save(path_to_evaluation_xlsx)
    return redirect(url_for("project_profile", project_id=project_id))


@app.route('/delete_project/<string:project_id>', methods=['GET', 'POST'])
@login_required
def delete_project(project_id):
    project = Permission.query.filter_by(project_id=project_id).first()
    permission_to_delete = Permission.query.filter_by(project=project.project).all()
    path_to_current_project = "{}/{}/{}".format(base_directory, current_user.username, project.project)
    if os.path.exists(path_to_current_project):
        shutil.rmtree(path_to_current_project)

        # after delete the folder, delete all the permissions that were send from the project
        for permission in permission_to_delete:
            db.session.delete(permission)
            db.session.commit()

        # delete the project in project table
        project_in_database = Project.query.filter_by(project_name=project.project, owner=project.owner).first()
        db.session.delete(project_in_database)
        db.session.commit()
        msg = "project has been deleted successfully"
    else:
        msg = "something went wrong with the project"

    return redirect(url_for("project_profile_jumptool", project_id=project_id))


@app.route('/update_permission/<string:project_id>', methods=["GET", "POST"])
@login_required
def update_permission(project_id):
    try:
        submit = request.form['submit']
        if submit == 'update':
            authority = request.form['authority']
            query = Permission.query.filter_by(project_id == project_id)
            query.status = authority
            db.session.commit()
            msg = "successfully updated authority"
        else:
            query = Permission.query.filter_by(project_id == project_id)
            db.session.delete(query)
            db.session.commit()
            msg = "successfully delete permission"
    except Exception as e:

        msg = "failure to update authority, {}".format(e)

    return redirect(url_for("project_profile"))


@app.route('/create_permission/<string:project_id>', methods=["GET", "POST"])
@login_required
def create_permission(project_id):
    username = request.form.get('username', " ")
    authority = request.form['authority']
    pending_authority = "pending|{}".format(authority)
    try:
        # create permission:
        project = Permission.query.filter_by(project_id=project_id).first()
        permission_projectid = "{}{}{}{}".format(current_user.username, username, project.project, authority)
        add_permission = Permission(project_id=permission_projectid, owner=current_user.username, shareTo=username,
                                    project=project.project, status=pending_authority)
        db.session.add(add_permission)
        db.session.commit()
        # create notification:
        time = datetime.today().strftime('%d/%m/%Y')
        message_content = "{} sends a project invitation to you".format(current_user.username)
        notification = Notification(from_user=current_user.username, to_user=username, message_type="permission",
                                    message_content=message_content, status="unread", time=time,
                                    appendix=permission_projectid)
        db.session.add(notification)
        db.session.commit()
        msg = "successfully created authority"
    except Exception as e:

        msg = "failure to create authority, {}".format(e)

    return redirect(url_for("project_profile"))


@app.route('/modify_group/<string:project>')
@login_required
def modify_group(project):
    return


@app.route('/modify_group_receiver/<string:project>')
@login_required
def modify_group_receiver(project):
    return


@app.route('/instructor_project', methods=["POST", "GET"])
@login_required
def instructor_project():
    # Load All project and shared project from database
    list_of_all_projects = Permission.query.filter_by(shareTo=current_user.username).all()
    list_of_personal_projects = Permission.query.filter_by(owner=current_user.username,
                                                           shareTo=current_user.username).all()
    # list_of_shared_project = [item for item in list_of_all_projects if item not in list_of_personal_projects]
    list_of_shared_project = []
    for project in list_of_all_projects:
        flag = True
        for personal_project in list_of_personal_projects:
            if project.project_id == personal_project.project_id:
                flag = False
        if flag:
            list_of_shared_project.append(project)

    list_of_personal_project_database = {}
    list_of_shared_project_database = {}
    # load the description of project
    for personal_project in list_of_personal_projects:
        project_in_project_db = Project.query.filter_by(project_name=personal_project.project,
                                                        owner=personal_project.owner).first()
        list_of_personal_project_database[project_in_project_db.project_name] = project_in_project_db
    for shared_project in list_of_shared_project:
        project_in_project_db = Project.query.filter_by(project_name=shared_project.project,
                                                        owner=shared_project.owner).first()
        list_of_shared_project_database[project_in_project_db.project_name] = project_in_project_db
    return render_template('instructor_project.html', personal_project_list=list_of_personal_projects,
                           shared_project_list=list_of_shared_project,
                           list_of_personal_project_database=list_of_personal_project_database,
                           list_of_shared_project_database=list_of_shared_project_database)


@app.route('/create_project', methods=["POST", "GET"])
@login_required
def create_project():
    # Request from file by WTF
    # Create a new project folder under 'path_to_current_user'
    # save files in new folder and build a evaluation doc depending on json file
    path_to_current_user = "{}/{}".format(base_directory, current_user.username)
    path_to_student_file = "{}/student.xlsx".format(path_to_current_user)
    path_to_json_file = "{}/TW.json".format(path_to_current_user)
    form = ProjectForm()

    if form.validate_on_submit():
        # create project folder
        path_to_current_user_project = "{}/{}/{}".format(base_directory, current_user.username, form.project_name.data)
        os.mkdir(path_to_current_user_project)

        path_to_student_file_stored = "{}/student.xlsx".format(path_to_current_user_project)
        shutil.move(path_to_student_file, path_to_student_file_stored)
        path_to_json_file_stored = "{}/TW.json".format(path_to_current_user_project)
        shutil.move(path_to_json_file, path_to_json_file_stored)
        # creating evaluation doc based on grading criteria json file

        # copy student sheet to evaluation doc
        student_file_workbook = openpyxl.load_workbook(path_to_student_file_stored)
        student_file_worksheet = student_file_workbook['Sheet1']

        # create group file depending on student file
        list_of_group = select_by_col_name('group', student_file_worksheet)
        set_of_group = set(list_of_group)
        # create a group workbook
        path_to_group_file = "{}/group.xlsx".format(path_to_current_user_project)
        group_workbook = openpyxl.Workbook()
        group_file_worksheet = group_workbook.create_sheet('Sheet1')
        meta_file_worksheet = group_workbook.create_sheet('Sheet2')
        # all student information map
        student_map_list = []
        for student_index in range(2, len(list(student_file_worksheet.iter_rows())) + 1):
            student_map_list.append(select_map_by_index(student_index, student_file_worksheet))
        # insert group columns
        group_file_worksheet.cell(1, 1).value = 'groupid'
        meta_file_worksheet.cell(1, 1).value = 'groupid'
        meta_file_worksheet.cell(1, 2).value = 'metaid'
        start_index = 2
        for group in set_of_group:
            group_file_worksheet.cell(start_index, 1).value = group
            student_emails = [x['Email'] for x in student_map_list if x['group'] == group]
            meta_file_worksheet.cell(start_index, 1).value = group
            meta_group = [x['meta'] for x in student_map_list if x['group'] == group][0]
            meta_file_worksheet.cell(start_index, 2).value = meta_group
            for insert_index in range(2, len(student_emails) + 2):
                group_file_worksheet.cell(start_index, insert_index).value = student_emails[insert_index - 2]
            start_index += 1
        group_workbook.save(path_to_group_file)

        path_to_evaluation = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_workbook = openpyxl.Workbook()
        evaluation_group = evaluation_workbook.create_sheet('group')
        evaluation_meta = evaluation_workbook.create_sheet('meta')
        evaluation_student = evaluation_workbook.create_sheet('students')
        copy_all_worksheet(evaluation_group, group_file_worksheet)
        copy_all_worksheet(evaluation_meta, meta_file_worksheet)
        copy_all_worksheet(evaluation_student, student_file_worksheet)
        # create EVA depending on the json file
        evaluation_eva = evaluation_workbook.create_sheet('eva')
        # open json file and load json
        with open(path_to_json_file_stored, 'r')as f:
            json_data = json.loads(f.read(), strict=False)
        # The group id, eva_name, date are defults
        tags_to_append = ['group_id', 'eva_name', 'owner', 'date', 'students']
        for category in json_data['category']:
            category_name = (category['name'])
            for section in category['section']:
                # instructors don't care about the text value, the text values will only be send to students.
                if section['type'] != 'text':
                    value_to_append = "{}|{}".format(category_name, section['name'])
                    tags_to_append.append(value_to_append)
        tags_to_append.append("comment")
        tags_to_append.append("last_updates")
        evaluation_eva.append(tags_to_append)

        evaluation_workbook.save(path_to_evaluation)

        # create permission to owener himself
        project_id = "{}{}{}{}".format(current_user.username, current_user.username, form.project_name.data, 'full')
        self_permission = Permission(project_id=project_id, owner=current_user.username, shareTo=current_user.username,
                                     project=form.project_name.data, status='full')
        db.session.add(self_permission)
        db.session.commit()

        # create the project in database
        project_id = "{}{}".format(current_user.username, form.project_name.data)
        project_to_add = Project(project_name=form.project_name.data, project_status='public',
                                 owner=current_user.username, description=form.project_description.data)
        db.session.add(project_to_add)
        db.session.commit()

        return redirect(url_for("instructor_project"))


    else:
        if os.path.exists(path_to_student_file):
            os.remove(path_to_student_file)
        if os.path.exists(path_to_json_file):
            os.remove(path_to_json_file)
        return render_template('create_project.html', form=form)


def copy_all_worksheet(copy_to, copy_from):
    for row in range(0, len(list(copy_from.iter_rows()))):
        for col in range(0, len(list(copy_from.iter_cols()))):
            copy_to.cell(row=row + 1, column=col + 1).value = copy_from.cell(row=row + 1, column=col + 1).value


@app.route('/load_project/<string:project_id>/<string:msg>', methods=["GET"])
@login_required
def load_project(project_id, msg):
    # get project by project_id
    project = Permission.query.filter_by(project_id=project_id).first()
    path_to_evaluation_xlsx = "{}/{}/{}/evaluation.xlsx".format(base_directory, project.owner, project.project)
    evaluation_workbook = openpyxl.load_workbook(path_to_evaluation_xlsx)
    evaluation_worksheet = evaluation_workbook['eva']
    meta_worksheet = evaluation_workbook['meta']
    set_of_eva = Evaluation.query.filter_by(project_name=project.project, project_owner=project.owner).all()
    set_of_meta = set(select_by_col_name('metaid', meta_worksheet))

    # get all owner
    # dic_of_eva = {}
    # for eva in set_of_eva:
    #     owners = set()
    #     temp_eva = select_row_by_group_id("eva_name", eva, evaluation_worksheet)
    #     for row in temp_eva:
    #         owners.add(row['owner'])
    #     dic_of_eva[eva] = owners
    return render_template("project.html", project=project, data_of_eva_set=set_of_eva, set_of_meta=set_of_meta,
                           msg=msg, useremail=current_user.email)


# @app.route('/create_new_evaluation/<string:project_name>')
# @login_required
# def create_new_evaluation(project_name):
#     return render_template('create_new_evaluation.html', project_name=project_name)

@app.route('/create_evaluation/<string:project_id>', methods=['GET', 'POST'])
@login_required
def create_evaluation(project_id):
    # get project by id
    project = Permission.query.filter_by(project_id=project_id).first()
    # load group columns and evaluation worksheet
    evaluation_name = request.form['evaluation_name']
    evaluation_name_find_in_db = Evaluation.query.filter_by(project_owner=current_user.username,
                                                            project_name=project.project,
                                                            eva_name=evaluation_name).first()
    if evaluation_name_find_in_db is None:
        evaluation_desc = request.form.get('evaluation_description', " ")
        path_to_load_project = "{}/{}/{}".format(base_directory, current_user.username, project.project)
        path_to_evaluation_file = "{}/evaluation.xlsx".format(path_to_load_project)
        eva_workbook = load_workbook(path_to_evaluation_file)
        group_worksheet = eva_workbook['group']
        eva_worksheet = eva_workbook['eva']
        meta_worksheet = eva_workbook['meta']
        students_worksheet = eva_workbook['students']

        group_col = []
        for col_item in list(group_worksheet.iter_cols())[0]:
            if col_item.value != "groupid":
                group_col.append(col_item.value)

        # get all students by students
        students = get_students_by_group(group_worksheet, students_worksheet)
        # create a empty row for each group in the new evaluation
        for group in group_col:
            students_name = []
            # couple is [email, student_name]
            for student_couple in students[str(group)]:
                students_name.append(student_couple[1])
            row_to_insert = new_row_generator(str(group), students_name, evaluation_name, eva_worksheet)
            eva_worksheet.append(row_to_insert)
        eva_workbook.save(path_to_evaluation_file)

        # create evaluation in database:
        evaluation_to_add = Evaluation(eva_name=evaluation_name, project_name=project.project,
                                       project_owner=project.owner, owner=current_user.username,
                                       description=evaluation_desc)
        db.session.add(evaluation_to_add)
        db.session.commit()
        msg = "New Evaluation has been created successfully"

        set_of_meta = set(select_by_col_name('metaid', meta_worksheet))
        return redirect(
            url_for('jump_to_evaluation_page', project_id=project.project_id, evaluation_name=evaluation_name, metaid=set_of_meta.pop(), group="***None***", msg=msg))

    else:
        print(evaluation_name_find_in_db)
        return redirect(url_for('load_project', project_id=project_id, msg="The evaluation_name has been used before"))


# @app.route('/evaluation_jump_tool/<string:project_id>/<string:evaluation_name>/<string:msg>', methods=["GET", "POST"])
# @login_required
# def evaluation_jump_tool(project_id, evaluation_name, msg):
#     # get project by project_id
#     project = Permission.query.filter_by(project_id=project_id).first()
#     # prepare the json data and group numbers before it jumps to evaluation page
#     path_to_load_project = "{}/{}/{}".format(base_directory, project.owner, project.project)
#     eva_workbook = load_workbook("{}/evaluation.xlsx".format(path_to_load_project))
#     eva_worksheet = eva_workbook['eva']
#     group_worksheet = eva_workbook['group']
#     meta_worksheet = eva_workbook['meta']
#     students_worksheet = eva_workbook['students']
#
#     # data of groups and last edit
#     group_col = []
#     group_last_edit = {}
#
#     for col_item in list(group_worksheet.iter_cols())[0]:
#         if col_item.value != "groupid":
#             group_col.append(col_item.value)
#             index_list = select_index_by_group_eva(evaluation_name, col_item.value, eva_worksheet)
#             #get the last edit
#             row = select_map_by_index(int(index_list[-1]), eva_worksheet)
#             if len(row['last_updates'].split('|')) > 1 or len(index_list) > 1:
#                 last_edit_owner = row['owner']
#                 last_edit_date = row['date']
#                 group_last_edit[col_item.value] = [last_edit_owner, last_edit_date]
#             else:
#                 group_last_edit[col_item.value] = ['none', 'none']
#
#
#
#
#     students = get_students_by_group(group_worksheet, students_worksheet)
#     return render_template("evaluation_jump_tool.html", project=project, evaluation_name=evaluation_name, group_col=group_col, group_last_edit=group_last_edit, students=students)
#
# @app.route('/jump_to_evaluation_page/<string:project_id>/<string:evaluation_name>/<string:group>/<string:msg>', methods=["GET", "POST"])
# @login_required
# def jump_to_evaluation_page(project_id, evaluation_name, group, msg):
#     #get project by project_id
#     project = Permission.query.filter_by(project_id=project_id).first()
#     #prepare the json data and group numbers before it jumps to evaluation page
#     path_to_load_project = "{}/{}/{}".format(base_directory, project.owner, project.project)
#     with open ("{}/TW.json".format(path_to_load_project), 'r')as f:
#         json_data = json.loads(f.read(), strict=False)
#     eva_workbook = load_workbook("{}/evaluation.xlsx".format(path_to_load_project))
#     group_worksheet = eva_workbook['group']
#     students_worksheet = eva_workbook['students']
#
#     #data of groups
#     group_col = []
#     for col_item in list(group_worksheet.iter_cols())[0]:
#         if col_item.value != "groupid":
#             group_col.append(col_item.value)
#
#     #check if evaluation exists in the worksheet
#     eva_worksheet = eva_workbook['eva']
#
#     #Transform ROWS in worksheet to DICTIONARY
#     new_row = {}
#     first_row = list(eva_worksheet.iter_rows())[0]
#     for tag in first_row:
#         new_row[tag.value] = ""
#
#     temp_eva = select_row_by_group_id("eva_name", evaluation_name, eva_worksheet)
#
#     #dictionary contains all data
#     eva_to_edit = {}
#     #list contains only owners
#     owner_list = []
#
#     # for group in group_col:
#     for row in temp_eva:
#         # if str(group) == str(row['group_id']) and str(row['owner'] == owner):
#         #     eva_to_edit[str(group)] = row
#         if str(group) == str(row['group_id']):
#             owner_per_row = str(row['owner'])
#             date = str(row['date'])
#             #tuple will be unique in this evaluation
#             tuple = (owner_per_row, date)
#             owner_list.append(tuple)
#             eva_to_edit[tuple] = row
#     students = get_students_by_group(group_worksheet, students_worksheet)
#
#     return render_template("evaluation_page.html",  project=project, json_data=json_data, group=group, group_col=group_col, msg=msg, evaluation_name=evaluation_name, edit_data=eva_to_edit, owner_list=owner_list, students=students, current_user=current_user.username)


@app.route(
    '/jump_to_evaluation_page/<string:project_id>/<string:evaluation_name>/<string:metaid>/<string:group>/<string:msg>',
    methods=["GET", "POST"])
@login_required
def jump_to_evaluation_page(project_id, evaluation_name, metaid, group, msg):
    # get project by project_id
    project = Permission.query.filter_by(project_id=project_id).first()
    # prepare the json data and group numbers before it jumps to evaluation page
    path_to_load_project = "{}/{}/{}".format(base_directory, project.owner, project.project)
    with open("{}/TW.json".format(path_to_load_project), 'r')as f:
        json_data = json.loads(f.read(), strict=False)
    eva_workbook = load_workbook("{}/evaluation.xlsx".format(path_to_load_project))
    group_worksheet = eva_workbook['group']
    meta_worksheet = eva_workbook['meta']
    students_worksheet = eva_workbook['students']

    # data of meta groups
    set_of_meta = set(select_by_col_name('metaid', meta_worksheet))
    meta_group_map_list = []
    for group_index in range(2, len(list(meta_worksheet.iter_rows())) + 1):
        meta_group_map_list.append(select_map_by_index(group_index, meta_worksheet))
    group_col = [x['groupid'] for x in meta_group_map_list if str(x['metaid']) == str(metaid)]
    # if only click on meta group, by default choose its first group
    if group == "***None***":
        group = group_col[0]
    # check if evaluation exists in the worksheet
    eva_worksheet = eva_workbook['eva']

    # Transform ROWS in worksheet to DICTIONARY
    new_row = {}
    first_row = list(eva_worksheet.iter_rows())[0]
    for tag in first_row:
        new_row[tag.value] = ""

    temp_eva = select_row_by_group_id("eva_name", evaluation_name, eva_worksheet)

    # dictionary contains all data
    eva_to_edit = {}
    # list contains only owners
    owner_list = []

    # for group in group_col:
    for row in temp_eva:
        # if str(group) == str(row['group_id']) and str(row['owner'] == owner):
        #     eva_to_edit[str(group)] = row
        if str(group) == str(row['group_id']):
            owner_per_row = str(row['owner'])
            date = str(row['date'])
            # tuple will be unique in this evaluation
            tuple = (owner_per_row, date)
            owner_list.append(tuple)
            eva_to_edit[tuple] = row
    students = get_students_by_group(group_worksheet, students_worksheet)

    return render_template("evaluation_page.html", project=project, json_data=json_data, group=group, metaid=metaid,
                           group_col=group_col, set_of_meta=set_of_meta, msg=msg, evaluation_name=evaluation_name,
                           edit_data=eva_to_edit, owner_list=owner_list, students=students,
                           current_user=current_user.username)


@app.route(
    '/evaluation_receiver/<project_id>/<string:evaluation_name>/<string:metaid>/<string:group>/<string:owner>/<string:past_date>',
    methods=["GET", "POST"])
@login_required
def evaluation_page(project_id, evaluation_name, metaid, group, owner, past_date):
    # receive all the data and insert them into xlsx
    # group id, evaluation name, date time is constant
    row_to_insert = []
    group_id = group
    submit_type = request.form['submit_button']
    date = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    row_to_insert.append(group_id)
    row_to_insert.append(evaluation_name)
    # if edit, create a new row with owner = current user
    if submit_type == 'edit':
        row_to_insert.append(current_user.username)
    else:
        row_to_insert.append(owner)
    row_to_insert.append(date)
    student_list = request.form.getlist('student')
    if len(student_list) > 0:
        row_to_insert.append("|".join(student_list))
    else:
        row_to_insert.append(" ")

    # The rest are variables from TW
    # get project by project_id
    project = Permission.query.filter_by(project_id=project_id).first()
    path_to_load_project = "{}/{}/{}".format(base_directory, project.owner, project.project)
    with open("{}/TW.json".format(path_to_load_project), 'r')as f:
        json_data = json.loads(f.read(), strict=False)
    for category in json_data['category']:
        category_name = category['name']
        for section in category['section']:
            section_name = section['name']
            section_name = '{}|{}'.format(category_name, section_name)
            if section['type'] == 'radio':
                value = request.form.get(section_name, " ")
                row_to_insert.append(value)
            elif section['type'] == 'checkbox':
                value = request.form.getlist(section_name)
                if len(value) != 0:
                    value = '|'.join(value)
                else:
                    value = " "
                row_to_insert.append(value)
            else:
                # text don't need to be saved
                print('to be continued')
    path_to_evaluation_file = "{}/evaluation.xlsx".format(path_to_load_project)
    evaluation_workbook = load_workbook(path_to_evaluation_file)
    evaluation_worksheet = evaluation_workbook['eva']
    # change the last update by append the current user according to submit type
    if submit_type == 'update':
        index = int(
            select_index_by_group_eva_owner_date(evaluation_name, group_id, owner, past_date, evaluation_worksheet))
        # delete the old row by index
        last_comment = select_by_col_name('comment', evaluation_worksheet)[index - 2]
        comment = request.form.get('comment', " ")
        if comment != " " and last_comment != " ":
            comment = "{}|{}".format(last_comment, comment)
        else:
            comment = last_comment
        row_to_insert.append(comment)
        last_update = select_by_col_name('last_updates', evaluation_worksheet)[index - 2]
        row_to_insert.append(last_update)
        evaluation_worksheet.delete_rows(index, 1)
        evaluation_worksheet.append(row_to_insert)

    elif submit_type == 'create':
        comment = request.form.get('comment', " ")
        row_to_insert.append(comment)
        last_update = current_user.username
        row_to_insert.append(last_update)
        evaluation_worksheet.append(row_to_insert)

    elif submit_type == 'edit':
        comment = request.form.get('comment', " ")
        row_to_insert.append(comment)
        last_update = current_user.username
        row_to_insert.append(last_update)
        evaluation_worksheet.append(row_to_insert)
    elif submit_type == 'overwrite':
        index = int(
            select_index_by_group_eva_owner_date(evaluation_name, group_id, owner, past_date, evaluation_worksheet))
        last_comment = select_by_col_name('comment', evaluation_worksheet)[index - 2]
        comment = request.form.get('comment', " ")
        if comment != " " and last_comment != " ":
            comment = "{}|{}".format(last_comment, comment)
        else:
            comment = last_comment
        row_to_insert.append(comment)
        last_update = select_by_col_name('last_updates', evaluation_worksheet)[index - 2]
        last_update = "{}|{}".format(last_update, current_user.username)
        row_to_insert.append(last_update)
        evaluation_worksheet.delete_rows(index, 1)
        evaluation_worksheet.append(row_to_insert)

    # save the workbook
    evaluation_workbook.save(path_to_evaluation_file)

    # change the last edit
    evaluation_in_database = Evaluation.query.filter_by(project_name=project.project, project_owner=project.owner,
                                                        eva_name=evaluation_name).first()
    evaluation_in_database.last_edit = current_user.username
    db.session.commit()
    msg = "The grade has been updated successfully"
    return redirect(
        url_for('jump_to_evaluation_page', project_id=project_id, evaluation_name=evaluation_name, metaid=metaid,
                group=group, owner=owner, msg=msg))


@app.route('/download_page/<string:project_id>/<string:evaluation_name>/<string:group>/<string:type>',
           methods=['GET', 'POST'])
@login_required
# send all grades in the given evaluation
def download_page(project_id, evaluation_name, group, type):
    # get project by project_id
    project = Permission.query.filter_by(project_id=project_id).first()
    path_to_load_project = "{}/{}/{}".format(base_directory, project.owner, project.project)
    path_to_evaluation_file = "{}/evaluation.xlsx".format(path_to_load_project)
    eva_workbook = load_workbook(path_to_evaluation_file)
    group_worksheet = eva_workbook['group']
    eva_worksheet = eva_workbook['eva']
    students_worksheet = eva_workbook['students']

    with open("{}/TW.json".format(path_to_load_project), 'r')as f:
        json_data = json.loads(f.read(), strict=False)

    if type == "normal":
        temp_eva = select_row_by_group_id("eva_name", evaluation_name, eva_worksheet)
        temp_eva_in_group = [x for x in temp_eva if x['group_id'] == group]

        students_in_one_group = get_students_by_group(group_worksheet, students_worksheet)[group]
        msg = "Downloaded grade group{}".format(group)

        return render_template("download_page.html", project=project, json_data=json_data, group=group, msg=msg,
                               evaluation_name=evaluation_name, students=students_in_one_group,
                               grades=temp_eva_in_group)


# download xlsx file or html file
@app.route('/download/<string:project_id>/<string:evaluation_name>', methods=['GET', 'POST'])
@login_required
def download(project_id, evaluation_name):
    project = Permission.query.filter_by(project_id=project_id).first()
    path_to_load_project = "{}/{}/{}".format(base_directory, current_user.username, project.project)
    path_to_evaluation_file = "{}/evaluation.xlsx".format(path_to_load_project)
    eva_workbook = load_workbook(path_to_evaluation_file)
    group_worksheet = eva_workbook['group']
    eva_worksheet = eva_workbook['eva']
    students_worksheet = eva_workbook['students']
    students = get_students_by_group(group_worksheet, students_worksheet)
    with open("{}/TW.json".format(path_to_load_project), 'r')as f:
        json_data = json.loads(f.read(), strict=False)
    group_col = []
    for col_item in list(group_worksheet.iter_cols())[0]:
        if col_item.value != "groupid":
            group_col.append(col_item.value)
    if evaluation_name == "all":
        try:
            filename = "{}_{}.xlsx".format(project.project, evaluation_name)
            return send_file(path_to_evaluation_file, attachment_filename=filename)
            # msg = "Successfully downloaded"
            # return redirect(url_for('load_project', project_name=project_name, msg=msg))
        except Exception as e:
            print(str(e))
    else:
        # response = urllib.request.urlopen(url_for('jump_to_evaluation_page', project_name=project_name, evaluation_name=evaluation_name))
        # html = response.read()
        # print(html)
        new_row = {}
        first_row = list(eva_worksheet.iter_rows())[0]
        for tag in first_row:
            new_row[tag.value] = ""
        # grab the data that used for this evaluation
        owner = request.form['owner']
        temp_eva = select_row_by_group_id("eva_name", evaluation_name, eva_worksheet)
        eva_to_edit = {}
        for group in group_col:
            for row in temp_eva:
                if str(group) == str(row['group_id']) and str(owner) == str(row['owner']):
                    eva_to_edit[str(group)] = row
        # store the data to a html file and send out(download)
        msg = ""
        path_to_html = "{}/{}_{}.html".format(path_to_load_project, project.project, evaluation_name)
        # remove the old file in case duplicated file existence
        if os.path.exists(path_to_html):
            os.remove(path_to_html)
        with open(path_to_html, 'w') as f:
            f.write(render_template("evaluation_page.html", project=project, json_data=json_data, group_col=group_col,
                                    msg=msg, evaluation_name=evaluation_name, edit_data=eva_to_edit, owner=owner,
                                    students=students))
        return send_file(path_to_html, as_attachment=True)


# def jump_to_evaluation_page(project_id, evaluation_name, group, msg):
@app.route('/sendEmail/<string:project_id>/<string:evaluation_name>', methods=['GET', 'POST'])
@login_required
def sendEmail(project_id, evaluation_name):
    project = Permission.query.filter_by(project_id=project_id).first()
    path_to_load_project = "{}/{}/{}".format(base_directory, project.owner, project.project)
    path_to_evaluation_file = "{}/evaluation.xlsx".format(path_to_load_project)
    eva_workbook = load_workbook(path_to_evaluation_file)
    group_worksheet = eva_workbook['group']
    eva_worksheet = eva_workbook['eva']
    students_worksheet = eva_workbook['students']
    with open("{}/TW.json".format(path_to_load_project), 'r')as f:
        json_data = json.loads(f.read(), strict=False)
    # data of groups
    group_col = []
    for col_item in list(group_worksheet.iter_cols())[0]:
        if col_item.value != "groupid":
            group_col.append(col_item.value)

    # #request data from EmailForm
    # useremail = request.form['useremail']
    # userpassword = request.form['userpassword']

    # #set up server and log in email (by sendmail python)
    # try:
    #     server = smtplib.SMTP(host='smtp-mail.outlook.com', port=587)
    #     server.starttls()
    #     server.login(useremail, userpassword)
    #     #send out emails by group
    #     for group in group_col:
    #         students_email = select_students_by_group(group, group_worksheet)
    #         index_of_group = int(select_index_by_group_eva_owner(evaluation_name, group, owner, eva_worksheet))
    #         grade_map = select_map_by_index(index_of_group, eva_worksheet)
    #         students_in_one_group = get_students_by_group(group_worksheet, students_worksheet)[group]
    #         #load download_page.html and store it to 'part' which will be attached to message in mail
    #         path_to_html = "{}/{}_{}_{}.html".format(path_to_load_project, project.project, evaluation_name, group)
    #         #write the download page html and automatically stored in local project
    #         msg = "Downloaded grade group{}".format(group)
    #         with open(path_to_html, 'w') as f:
    #             f.write(render_template("download_page.html", project=project, json_data=json_data,
    #                                     group=group, msg=msg, evaluation_name=evaluation_name,
    #                                     edit_data=grade_map, owner=owner, students=students_in_one_group))
    #         #load the download page to message
    #
    #         with open(path_to_html, 'rb')as f:
    #             file_data = f.read()
    #             file_name = "{}_{}_{}.html".format(project.project, evaluation_name, group)
    #         for email in students_email:
    #             # create an instance of message
    #             if email is not None:
    #                 message = EmailMessage()
    #                 message['From'] = useremail
    #                 message['To'] = email
    #                 message['Subject'] = "{}|{}|{}".format(project.project, evaluation_name, group)
    #                 text = "send from {}, {} for group{}".format(project.project, evaluation_name, group)
    #                 message.set_content(text)
    #                 # message.attach(part)
    #                 message.add_attachment(file_data, maintype='html', subtype='html', filename=file_name)
    #
    #                 server.send_message(message)
    #         #remove the html file after sending email
    #         #in case of duplicated file existence
    #         if os.path.exists(path_to_html):
    #             os.remove(path_to_html)
    #     server.close()
    try:
        from_email = "from=runqzhao@uiowa.edu"
        for group in group_col:
            students_email = select_students_by_group(group, group_worksheet)
            # grade_of_group = select_row_by_group_id(group)
            # students_in_one_group = get_students_by_group(group_worksheet, students_worksheet)[group]
            # load download_page.html and store it to 'part' which will be attached to message in mail
            path_to_html = "{}/{}_{}_{}.html".format(path_to_load_project, project.project, evaluation_name, group)
            file_name = "{}_{}_{}.pdf".format(project.project, evaluation_name, group)
            path_to_pdf = "{}/{}_{}_{}.pdf".format(path_to_load_project, project.project, evaluation_name, group)
            # #write the download page html and automatically stored in local project
            subject = "grade: project{}, evaluation{}, group{}".format(project.project, evaluation_name, group)
            with open(path_to_html, 'w') as f:
                f.write(download_page(project.project_id, evaluation_name, group, "normal"))
            with open(path_to_html, 'r') as f:
                pdf = HTML2PDF()
                pdf.add_page()
                pdf.write_html(f.read())
                pdf.output(path_to_pdf)
            # load the download page to message
            index = 0
            for email in students_email:
                # create an instance of message
                if email is not None:
                    # send by linux mail
                    # mail_linux_command = ""
                    subject += str(index)
                    index += 1
                    with open(path_to_pdf, "r") as file_to_pdf:
                        subprocess.call(["mail", "-s", subject, "-S", from_email, "-a", file_name, email],
                                        stdin=file_to_pdf)
        msg = "Emails send out Successfully"
    except Exception as e:
        print('Something went wrong' + str(e))
        msg = "Something went wrong"

    # remove the html file after sending email
    # in case of duplicated file existence

    if os.path.exists(path_to_html):
        os.remove(path_to_html)
    # if os.path.exists(path_to_pdf):
    #     os.remove(path_to_pdf)
    return redirect(url_for('project_profile', project_id=project_id))


@app.route('/account', methods=['GET', 'POST'])
@login_required
def account():
    # if the notification is "permission"
    notifications = Notification.query.filter_by(to_user=current_user.username).all()
    return render_template('account.html', notifications=notifications)


@app.route('/notification_receiver/<string:notification_id>', methods=['GET', 'POST'])
@login_required
def notification_receiver(notification_id):
    response = request.form['response']
    notification = Notification.query.filter_by(notification_id=notification_id).first()
    permission = Permission.query.filter_by(project_id=notification.appendix).first()
    authority = permission.status.split('|')[1]
    if response == 'Decline':
        db.session.delete(permission)
        db.session.commit()
    else:
        permission.status = authority
        db.session.commit()
    notification.status = 'read'
    db.session.commit()
    return redirect(url_for('account'))


@app.route('/student_dashboard')
@login_required
def student_dashboard():
    return render_template('student_dashboard.html', name=current_user.username)


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))


# def sendmail(from_email, password, to_email, subject, message):
#     msg = MIMEMultipart()
#     msg['From'] = from_email
#     msg['To'] = to_email
#     msg['Subject'] = subject
#     msg.attach(MIMEText(message, 'plain'))
#     try:
#         server = smtplib.SMTP(host='smtp-mail.outlook.com', port=587)
#         server.starttls()
#         server.login(from_email, password)
#         server.send_message(msg)
#         server.quit()
#         return "successfully"
#     except Exception as e:
#         print('Something went wrong' + str(e))
#         return "Something went wrong"

# select all elements by the column name
def select_by_col_name(col_name, worksheet):
    first_row = list(worksheet.iter_rows())[0]
    index_of_col = -1
    for variables in first_row:
        if variables.value == col_name:
            index_of_col = first_row.index(variables)
    if index_of_col == -1:
        return "col_name unmatched"
    else:
        col = list(worksheet.iter_cols())[index_of_col]
        data_by_col = []
        for data in range(1, len(col)):
            data_by_col.append(col[data].value)
        return data_by_col


# index is the row index in xsl
def select_map_by_index(index, worksheet):
    list_of_rows = list(worksheet.iter_rows())
    # row_to_return = []
    # for item in list_of_rows[index-1]:
    #     row_to_return.append(item.value)
    row_selected = list_of_rows[index - 1]
    tags = list_of_rows[0]
    map_to_return = {}
    for tag_index in range(0, len(tags)):
        tag_value = tags[tag_index].value
        map_to_return[tag_value] = row_selected[tag_index].value
    return map_to_return


# index is the row index in xsl
def select_row_by_index(index, worksheet):
    list_of_rows = list(worksheet.iter_rows())
    row_to_return = []
    for item in list_of_rows[index - 1]:
        row_to_return.append(item.value)
    return row_to_return


# select by the groupid and eva name
def select_index_by_group_eva(eva_name, group_id, worksheet):
    evaluation_list = select_by_col_name("eva_name", worksheet)
    group_list = select_by_col_name("group_id", worksheet)
    list_of_result = []
    for index in range(0, len(evaluation_list)):
        if evaluation_list[index] == eva_name and group_list[index] == group_id:
            list_of_result.append(index + 2)
    return list_of_result


# group_id and eva_name and owner are primary key in eva worksheet
def select_index_by_group_eva_owner(eva_name, group_id, owner, worksheet):
    evaluation_list = select_by_col_name("eva_name", worksheet)
    group_list = select_by_col_name("group_id", worksheet)
    owner_list = select_by_col_name("owner", worksheet)
    list_of_result = []
    for index in range(0, len(evaluation_list)):
        if evaluation_list[index] == eva_name and group_list[index] == group_id and owner_list[index] == owner:
            list_of_result.append(index + 2)
    return list_of_result


# find the unique index by eva_name, group id, owner, date
def select_index_by_group_eva_owner_date(eva_name, group_id, owner, date, worksheet):
    evaluation_list = select_by_col_name("eva_name", worksheet)
    group_list = select_by_col_name("group_id", worksheet)
    owner_list = select_by_col_name("owner", worksheet)
    date_list = select_by_col_name("date", worksheet)
    for index in range(0, len(evaluation_list)):
        if evaluation_list[index] == eva_name and group_list[index] == group_id and owner_list[
            index] == owner and date == date_list[index]:
            return index + 2
    return "nothing find"


def select_row_by_group_id(col_name, col_value, worksheet):
    # we suppose that the return rows are multiple
    rows_selected = []
    rows_in_worksheet = list(worksheet.iter_rows())
    # index in data_by_col == index in row
    data_by_col = select_by_col_name(col_name, worksheet)
    for index_data in range(0, len(data_by_col)):
        if data_by_col[index_data] == col_value:
            map_to_append = {}
            # tranfer the item to item.value
            for item in rows_in_worksheet[index_data + 1]:
                # create a map contains keys-values : like "group_id" = 1
                index_of_item = rows_in_worksheet[index_data + 1].index(item)
                tag_of_item = rows_in_worksheet[0][index_of_item].value
                map_to_append[tag_of_item] = item.value
            # record each row by its group_id
            # for example : 1:{group_id = 1, date= x/y/z, eva_name = eva3, .....}.
            rows_selected.append(map_to_append)
    # if nothing matched, return the empty array;
    return rows_selected


# select all students in the group
# implement in group worksheet
def select_students_by_group(group, worksheet):
    students = []
    groups = select_by_col_name("groupid", worksheet)
    index_of_group = groups.index(group) + 2
    row_by_index = list(worksheet.iter_rows())[index_of_group - 1]
    # the first element in row_by_index is groupid
    for student in row_by_index[1:]:
        students.append(student.value)
    return students


# generate an empty map with group name and evaluation name
def new_map_generator(group, eva_name, worksheet):
    map_to_return = {}
    tags_of_item = list(worksheet.iter_rows())[0]
    for tag in tags_of_item:
        if tag.value == 'group_id':
            map_to_return[tag.value] = group
        elif tag.value == 'eva_name':
            map_to_return[tag.value] = eva_name
        else:
            map_to_return[tag.value] = ''

    return map_to_return


def new_row_generator(group, students, eva_name, worksheet):
    row_to_return = []
    tags_of_item = list(worksheet.iter_rows())[0]
    for tag in tags_of_item:
        if tag.value == 'group_id':
            row_to_return.append(group)
        elif tag.value == 'eva_name':
            row_to_return.append(eva_name)
        elif tag.value == 'date':
            date = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            row_to_return.append(date)
        elif tag.value == 'owner':
            row_to_return.append(current_user.username)
        elif tag.value == 'students':
            students_string = ",".join(students)
            row_to_return.append(students_string)
        elif tag.value == 'last_updates':
            row_to_return.append(current_user.username)
        else:
            row_to_return.append(" ")
    return row_to_return


def get_students_by_group(group_worksheet, students_worksheet):
    return_map = {}
    for group in list(group_worksheet.iter_rows())[1:]:
        return_map[group[0].value] = []
        for email in group[1:]:
            if email.value is not None:
                index = select_by_col_name('Email', students_worksheet).index(email.value)
                student_name = select_by_col_name('Student', students_worksheet)[index]
                student_couple = [email.value, student_name]
                return_map[group[0].value].append(student_couple)
    return return_map


# After login===============================================================================================================================

if __name__ == '__main__':
    # db.create_all() # only run it the first time
    app.run(debug=True)

    # token: MFFt4RjpXNMh1c_T1AQj
