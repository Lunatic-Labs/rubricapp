from flask import Flask, render_template, redirect, url_for, request, send_file, jsonify, send_from_directory, safe_join, abort
from flask_bootstrap import Bootstrap
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, BooleanField
from flask_wtf.file import FileField, FileAllowed, FileRequired
from wtforms.validators import InputRequired, Email, Length, ValidationError
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
#from selenium import webdriver;
from filelock import Timeout, FileLock
from datetime import datetime, date
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
import os
import subprocess
# shutil used to delete whole directory(folder)
import shutil
import openpyxl
from openpyxl import load_workbook
import uuid
import time
import json
import sys
from fpdf import FPDF, HTMLMixin
import platform
from django.utils.safestring import mark_safe
from django.template import Library
import time
import json
import smtplib
from email.message import EmailMessage
from concurrent.futures import ThreadPoolExecutor, as_completed
import zipfile
from os.path import basename
import glob



# from dataBase import *

register = Library()



# file directory
# requirement of two arguments: file address of app.py and fire address of root directory.
files_dir = None
if len(sys.argv) > 1:
    files_dir = sys.argv[1]
elif platform.node() in ['rubric.cs.uiowa.edu', 'rubric-dev.cs.uiowa.edu']:
    files_dir = "/var/www/wsgi-scripts/rubric"
else:
    print(
        "Requires argument: path to put files and database (suggestion is `pwd` when already in directory containing app.py)")
    sys.exit(1)

app = Flask(__name__)
app.config['SECRET_KEY'] = 'Thisissupposedtobesecret!'
if platform.node() in ['rubric.cs.uiowa.edu', 'rubric-dev.cs.uiowa.edu']:
    dbpass = None
    with open ("{}/dbpass".format(files_dir), 'r') as f:
        dbpass = f.readline().rstrip()

    dbuser = None
    with open ("{}/dbuser".format(files_dir), 'r') as f:
        dbuser = f.readline().rstrip()

    app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql://{0}:{1}@127.0.0.1/rubric'.format(dbuser, dbpass)
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///{}/account.db'.format(files_dir)

bootstrap = Bootstrap(app)
db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# tables in database; each class match to a table in database
#   *size of username, project_id, owner, project_name should be consistent in different tables.
#   *password is encrypted

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    #use username or email to login
    username = db.Column(db.String(30), unique=True, nullable=False)
    email = db.Column(db.String(255), unique=True, nullable=False)
    password = db.Column(db.String(80), nullable=False)
    #role in university; ex. instructor or ta
    role = db.Column(db.String(20), nullable=True)
    University = db.Column(db.String(255), nullable=True)
    #self introduction
    description = db.Column(db.String(255), nullable=True)


class Permission(UserMixin, db.Model):
    #project_id is made up with projectname, owner, shareto
    project_id = db.Column(db.String(255), primary_key=True)
    owner = db.Column(db.String(30), nullable=False)
    shareTo = db.Column(db.String(30), nullable=False)
    #project is project name
    project = db.Column(db.String(150), nullable=False)
    status = db.Column(db.String(50), nullable=False)


class Project(UserMixin, db.Model):
    project_name = db.Column(db.String(150), primary_key=True)
    owner = db.Column(db.String(30), primary_key=True)
    project_status = db.Column(db.String(50), nullable=False)
    description = db.Column(db.String(255), nullable=True)


class Evaluation(UserMixin, db.Model):
    #sharer of project can also create evaluation (or not allowed); still undecided
    eva_name = db.Column(db.String(150), primary_key=True)
    project_name = db.Column(db.String(150), primary_key=True)
    project_owner = db.Column(db.String(30), primary_key=True)
    owner = db.Column(db.String(30), nullable=False)
    description = db.Column(db.String(255), nullable=True)
    last_edit = db.Column(db.String(30), nullable=True)


# not using this table right now
# designed to send messages among users
class Notification(UserMixin, db.Model):
    notification_id = db.Column(db.Integer, primary_key=True)
    from_user = db.Column(db.String(30), nullable=False)
    to_user = db.Column(db.String(50), nullable=False)
    message_type = db.Column(db.String(50), nullable=False)
    message_content = db.Column(db.String(255), nullable=True)
    status = db.Column(db.String(50), nullable=False)
    time = db.Column(db.String(50), nullable=False)
    appendix = db.Column(db.String(255), nullable=True)


# besides uploading rubric, we also offer default rubric
class DefaultRubric(UserMixin, db.Model):
    json_name = db.Column(db.String(150), primary_key=True)
    json_description = db.Column(db.String(500), nullable=True)
    json_owner = db.Column(db.String(30), nullable=True)


# sending emails usually takes a long time; this table record information of the process of email sending
class EmailSendingRecord(UserMixin, db.Model):
    project_name = db.Column(db.String(150), primary_key=True)
    project_owner = db.Column(db.String(30), primary_key=True)
    eva_name = db.Column(db.String(150), primary_key=True)
    num_of_tasks = db.Column(db.Integer, nullable=True)                 #total number of students
    num_of_finished_tasks = db.Column(db.Integer, nullable=True)
    time_email_sent = db.Column(db.TEXT, nullable = True)


# Settings of Directory ======================================================================================================

# SET THE BASE DIRECTORY
os.chdir(files_dir)
base_directory = os.getcwd()
home_directory = base_directory
base_directory = base_directory + "/users"

# login manager is a extension library for login system including login_required
# login_required
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


class LoginForm(FlaskForm):
    email = StringField('email', validators=[InputRequired(), Email(message='Invalid email'), Length(max=255)])
    password = PasswordField('password', validators=[InputRequired(), Length(min=8, max=80)])
    remember = BooleanField('remember me')


class RegisterForm(FlaskForm):
    email = StringField('email', validators=[InputRequired(), Email(message='Invalid email'), Length(max=255)])
    password = PasswordField('password', validators=[InputRequired(), Length(min=8, max=80)], description="password size between 8-80")

@register.filter(is_safe=True)
def js(obj):
    return mark_safe(json.dumps(obj))


#Validator is function which checks whether the information is correct before proceed;
# NameValidator is used in login
class NameValidator(object):
    @login_required
    def __call__(self, form, field):
        duplicate_project_name = Project.query.filter_by(project_name=field.data,
                                                         owner=current_user.username).first()
        # print(field.data)
        # print(current_user.username)
        # print(duplicate_project_name)
        if duplicate_project_name is not None:
            raise ValidationError("The project name has been used before")

# check whether primary keys are in the student file;
class validate_project_student_file(object):
    @login_required
    def __call__(self, form, field):
        try:
            path_to_current_user = "{}/{}".format(base_directory, current_user.username)
            path_to_student_file_stored = "{}/".format(path_to_current_user)
            student_file_filename = "student.xlsx"
            field.data.save(path_to_student_file_stored + student_file_filename)
            student_file_workbook = load_workbook(path_to_student_file_stored + student_file_filename)
            student_file_worksheet = student_file_workbook['Sheet1']
            find_Student = True if 'Student' in [x.value for x in
                                                 list(student_file_worksheet.iter_rows())[0]] else False
            find_Email = True if 'Email' in [x.value for x in list(student_file_worksheet.iter_rows())[0]] else False
            find_group = True if 'group' in [x.value for x in list(student_file_worksheet.iter_rows())[0]] else False
            find_meta_group = True if 'meta' in [x.value for x in list(student_file_worksheet.iter_rows())[0]] else False
            if find_group is False:
                # os.remove(path_to_student_file_stored)
                raise ValidationError("Can not find group")
            elif find_Student is False:
                raise ValidationError("Can not find Student")
            elif find_Email is False:
                raise ValidationError("Can not find Email")
            # os.remove(path_to_student_file_stored+student_file_filename)
            elif find_meta_group is False:
                raise ValidationError("Can not find meta - group")
        except Exception as e:
            raise ValidationError(e)


# check whether primary keys are in the rubric json file
class validate_project_json_file(object):
    @login_required
    def __call__(self, form, field):
        try:
            path_to_current_user = "{}/{}".format(base_directory, current_user.username)
            path_to_json_file_stored = "{}/".format(path_to_current_user)
            json_file_filename = "TW.json"
            field.data.save(path_to_json_file_stored + json_file_filename)
            myLock = FileLock((path_to_json_file_stored + json_file_filename) + '.lock', timeout=5)
            with myLock:
                with open(path_to_json_file_stored + json_file_filename, 'r')as f:
                    json_data = json.loads(f.read(), strict=False)

            if 'name' in json_data.keys() and 'category' in json_data.keys():
                for category in json_data['category']:
                    if 'name' in category.keys() and 'section' in category.keys():
                        category_name = (category['name'])
                        for section in category['section']:
                            if 'name' in section.keys() and 'type' in section.keys() and 'values' in section.keys():
                                for value in section['values']:
                                    if 'name' not in value.keys() or 'desc' not in value.keys():
                                        raise ValidationError("lack of NAME or DESC in json file")

                            else:
                                raise ValidationError("lack of NAME or TYPE or VALUES in json file")
                    else:
                        raise ValidationError("lack of NAME or SECTIONS in json file")
            else:
                raise ValidationError("lack of NAME or CATEGORY in json file")
        except Exception as e:
            raise ValidationError(e)
        # os.remove(path_to_json_file_stored+ json_file_filename)


# flaskform for wtf
class ProjectForm(FlaskForm):
    project_name = StringField('Project Name',
                               validators=[InputRequired(), Length(min=3, max=150), NameValidator()], description="3-150 characters")
    project_description = StringField('Description', validators=[Length(min=0, max=255)], description="0-255 characters")
    student_file = FileField('Roster', validators=[InputRequired(), validate_project_student_file()])
    json_file = FileField('Rubric', validators=[InputRequired(), validate_project_json_file()])

@app.route('/')
def index():
    return render_template('index.html')


#log in function; Access User table
@app.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()

    # login validator
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.email.data).first()
        if user:
            if check_password_hash(user.password, form.password.data):
                login_user(user, remember=form.remember.data)
                # instructor jump to instructor page, student jump to student page
                # if(user.instructor == "1"):
                return redirect(url_for('instructor_project')) # jacky: after login, users are directed to the Rubric page, instead of Overview page

            else:
                return render_template('login.html', msg="password not correct", form=form)
        else:
            return render_template('login.html', msg="user doesn't exist", form=form)

    return render_template('login.html', msg="", form=form)

#sign up function; Access User table
@app.route('/signup', methods=['GET', 'POST'])
def signup():
    form = RegisterForm()

    # signup validator
    if form.validate_on_submit():
        # check if the user and email has existed in the database
        check_email = User.query.filter_by(email=form.email.data).first()
        if check_email:
            return render_template('signup.html', form=form, msg="Warning !!! The email has been used")
        else:
            hashed_password = generate_password_hash(form.password.data, method='sha256')
            # In issue 28, we changed username to be email, we saved the username section as we don't need to change the table
            new_user = User(username=form.email.data, email=form.email.data, password=hashed_password)
            db.session.add(new_user)
            db.session.commit()

            # After making sure that the new user is created, the user's private folder can be created by using the user name

            path_to_user_folder = "{}/{}".format(base_directory, new_user.username)
            os.mkdir(path_to_user_folder)

            return redirect(url_for('login'))

    return render_template('signup.html', form=form, msg="")


# home page
@app.route('/instructor_dashboard')
@login_required
def instructor_dashboard():
    # Load all projects to instructor_dashboard
    # Find all projects in User's private folder by using current user

    path_to_current_user = "{}/{}".format(base_directory, current_user.username)
    project_list = [x.project for x in Permission.query.filter_by(owner=current_user.username, shareTo=current_user.username).all()]
    project_len = len(project_list)

    return render_template('instructor_dashboard.html', name=current_user.username, project_list=project_list,
                           project_len=project_len)

# Manage Rubrics: showing all private rubric;
@app.route('/project_profile_jumptool', methods=["POST", "GET"])
@login_required
def project_profile_jumptool():
    # a jump tool before load project
    # display the information of projects (title and desc) and its recent evaluations

    path_to_current_user = "{}/{}".format(base_directory, current_user.username)
    # list of evaluation & list of groups relating to one project
    project_set_map = {}
    # search projects in database by username
    project_list = Permission.query.filter_by(owner=current_user.username, shareTo=current_user.username).all()
    project_information_map = {}
    for project in project_list:
        path_to_evaluation_file = "{}/{}/{}/evaluation.xlsx".format(base_directory, current_user.username,
                                                                    project.project)
        evaluation_workbook = openpyxl.load_workbook(path_to_evaluation_file)
        evaluation_worksheet = evaluation_workbook['eva']
        group_worksheet = evaluation_workbook['group']
        group_col = []
        for col_item in list(group_worksheet.iter_cols())[0]:
            if col_item.value != "groupid":
                group_col.append(col_item.value)
        set_of_eva = Evaluation.query.filter_by(project_name=project.project, project_owner=current_user.username).all()
        project_set_map[project.project] = (group_col, set_of_eva)
        project_information_map[project.project] = Project.query.filter_by(project_name=project.project,
                                                                           owner=project.owner).first()

    return render_template("project_profile_jumptool.html", project_list=project_list, project_set_map=project_set_map,
                           project_information_map=project_information_map)

# Manage Rubrics: show single rubric and its evaluation grading status(graded group, ungraded group, grade table, )
@app.route('/project_profile/<string:project_id>/<string:msg>', methods=["POST", "GET"])
@login_required
def project_profile(project_id, msg):
    """
    It controls the project_profile.html page, it collects a list of all the evaluations and get them displayed on the web
    page
    :param project_id: project id
    :param msg: it is used to make sure the page is correctly loaded, if not, the error message will be displayed in the
    error box in the web page
    :return: a rendered web page with dictionaries map the info of current project
    """
    # show each grade in this project and divided into eva s
    project = Permission.query.filter_by(project_id=project_id).first()

    list_of_shareTo_permission = [x for x in
                                  Permission.query.filter_by(project=project.project, owner=current_user.username).all()
                                  if x.shareTo != current_user.username]

    path_to_evaluation_file = "{}/{}/{}/evaluation.xlsx".format(base_directory, current_user.username, project.project)
    evaluation_workbook = openpyxl.load_workbook(path_to_evaluation_file)
    evaluation_worksheet = evaluation_workbook['eva']
    group_worksheet = evaluation_workbook['group']
    meta_worksheet = evaluation_workbook['meta']
    list_of_eva = select_by_col_name('eva_name', evaluation_worksheet)
    set_of_eva = set(list_of_eva)
    # get all groups and its owners
    dic_of_eva = {}

    dic_of_choosen = {}
    set_of_meta = set(select_by_col_name('metaid', meta_worksheet))

    meta_group_map_list = []
    for group_index in range(2, len(list(meta_worksheet.iter_rows())) + 1):
        meta_group_map_list.append(select_map_by_index(group_index, meta_worksheet))

    for eva in set_of_eva:
        all_groups_choosen = set()
        all_groups_not_choosen = set()
        all_groups = set()
        choosen = {}
        for meta in set_of_meta:
            choosen[meta] = set()
        total = {}
        notchoosen = {}
        for meta in set_of_meta:

            notchoosen[meta] = set([x['groupid'] for x in meta_group_map_list if str(x['metaid']) == str(meta)])
            total[meta] = set([x['groupid'] for x in meta_group_map_list if str(x['metaid']) == str(meta)])
        # update 9/13: simple profile
        dic_of_eva[eva] = []
        temp_eva = select_row_by_group_id("eva_name", eva, evaluation_worksheet)
        for eva_row in temp_eva:
            dic_of_eva[eva].append(eva_row)
            for (key, value) in eva_row.items():
                if (key != "group_id") and (key != "eva_name") and (key != "owner") and (key != "date") and (key != "students") and (key != "last_updates"):
                    if (value is not None) and (value != " ") and (value != ""):
                        metaid = [x[0] for x in list(total.items()) if eva_row["group_id"] in x[1]][0]
                        choosen[metaid].add(eva_row["group_id"])
                        notchoosen[metaid].discard(eva_row["group_id"])
        for meta in set_of_meta:
            for choosen_i in choosen[meta]:
                all_groups_choosen.add(choosen_i)
            for notchoosen_j in notchoosen[meta]:
                all_groups_not_choosen.add(notchoosen_j)
            for all_k in total[meta]:
                all_groups.add(all_k)
        dic_of_choosen[eva] = (choosen, notchoosen, total, all_groups_choosen, all_groups_not_choosen, all_groups)

    tags = [x.value for x in list(evaluation_worksheet.iter_rows())[0]]

    # group management
    management_groups = []
    rows_got_from_group_worksheet = list(group_worksheet.iter_rows())
    for row in rows_got_from_group_worksheet:
        management_groups.append([x.value for x in row])

    records = EmailSendingRecord.query.filter_by(project_name=project.project, project_owner=current_user.username).all()
    if records is not None:
        sending_info_dict = {}
        for record in records:
            sending_info_dict[record.eva_name] = [record.num_of_tasks, record.num_of_finished_tasks, record.time_email_sent]
    else:
        sending_info_dict = {}

    return render_template("project_profile.html", dic_of_eva=dic_of_eva, meta_list=set_of_meta,
                           list_of_shareTo_permission=list_of_shareTo_permission, management_groups=management_groups,
                           tags=tags, project=project, set_of_eva=list(set_of_eva), dic_of_choosen=dic_of_choosen,
                           msg=msg, sending_info_dict=sending_info_dict)


@app.route('/management_group/<string:project_id>', methods=['GET', 'POST'])
@login_required
def managment_group(project_id):
    project = Permission.query.filter_by(project_id=project_id).first()
    path_to_evaluation_xlsx = "{}/{}/{}/evaluation.xlsx".format(base_directory, current_user.username, project.project)
    evaluation_workbook = openpyxl.load_workbook(path_to_evaluation_xlsx)
    group_worksheet = evaluation_workbook['group']
    for row_index in range(1, len(list(group_worksheet.iter_rows()))):
        for col_index in range(1, len(list(group_worksheet.iter_cols()))):
            student_email = request.form.get((list(group_worksheet.iter_cols())[0][row_index].value + str(col_index)),
                                             " ")

            # group management detector
            # if the given value is None, inserted should also be None
            if student_email == " " or student_email == "None":
                group_worksheet.cell(row_index + 1, col_index + 1).value = None
            else:
                group_worksheet.cell(row_index + 1, col_index + 1).value = student_email
    evaluation_workbook.save(path_to_evaluation_xlsx)
    return redirect(url_for("project_profile", project_id=project_id, msg="success"))


@app.route('/delete_eva/<string:project_id>/<string:evaluation>/<string:group>/<string:grader>/<string:datetime>',
           methods=['GET', 'POST'])
@login_required
def delete_eva(project_id, evaluation, group, grader, datetime):
    project = Permission.query.filter_by(project_id=project_id).first()
    path_to_evaluation_xlsx = "{}/{}/{}/evaluation.xlsx".format(base_directory, current_user.username, project.project)
    evaluation_workbook = openpyxl.load_workbook(path_to_evaluation_xlsx)
    evaluation_worksheet = evaluation_workbook['eva']
    group_worksheet = evaluation_workbook['group']
    allgroups = select_by_col_name('groupid', group_worksheet)
    students_worksheet = evaluation_workbook['students']

    index = int(select_index_by_group_eva_owner_date(evaluation, group, grader, datetime, evaluation_worksheet))
    evaluation_worksheet.delete_rows(index, 1)

    # check whether all group have at least one empty grade in this evaluation
    group_col_in_eva = set(select_by_col_name('group', evaluation_worksheet))
    empty_group = [x for x in allgroups if x not in group_col_in_eva]

    students = get_students_by_group(group_worksheet, students_worksheet)

    for empty in empty_group:
        students_name = []
        # couple is [email, student_name]
        for student_couple in students[str(group)]:
            students_name.append(student_couple[1])
        empty_row = new_row_generator(str(group), students_name, evaluation, evaluation_worksheet)
        evaluation_worksheet.append(empty_row)
    evaluation_workbook.save(path_to_evaluation_xlsx)
    return redirect(url_for("project_profile", project_id=project_id, msg="success"))


@app.route('/delete_project/<string:project_id>', methods=['GET', 'POST'])
@login_required
def delete_project(project_id):
    """
    Delete a project from database
    :param project_id: project id
    :return: rerender current page
    """
    project = Permission.query.filter_by(project_id=project_id).first()
    permission_to_delete = Permission.query.filter_by(project=project.project).all()
    path_to_current_project = "{}/{}/{}".format(base_directory, current_user.username, project.project)
    if os.path.exists(path_to_current_project):
        shutil.rmtree(path_to_current_project)

        # after delete the folder, delete all the permissions that were send from the project
        for permission in permission_to_delete:
            db.session.delete(permission)
            db.session.commit()

        # delete the project in project table
        project_in_database = Project.query.filter_by(project_name=project.project, owner=project.owner).first()
        db.session.delete(project_in_database)
        db.session.commit()
        msg = "project has been deleted successfully"
    else:
        msg = "something went wrong with the project"

    return redirect(url_for("project_profile_jumptool", project_id=project_id))


@app.route('/update_permission/<string:project_id>/<string:project_id_full>', methods=["GET", "POST"])
@login_required
def update_permission(project_id, project_id_full):
    try:
        submit = request.form['submit']
        if submit == 'update':
            authority = request.form['authority']
            query = Permission.query.filter_by(project_id=project_id).first()
            query.status = authority
            db.session.commit()
            msg = "successfully updated authority"
        else:
            query = Permission.query.filter_by(project_id=project_id).first()
            db.session.delete(query)
            db.session.commit()
            msg = "successfully delete permission"
    except Exception as e:
        msg = "failure to update authority, {}".format(e)

    return redirect(url_for("project_profile", project_id=project_id_full, msg=msg))


@app.route('/create_permission/<string:project_id>', methods=["GET", "POST"])
@login_required
def create_permission(project_id):
    """
    This is being used in project_profile.html, which creates permission to a another user to share the rubric. The func-
    ction first search the typed in username, if the user exist, it creates a permission in Permission table, otherwise
    it returns to current page with error messages displayed
    :param project_id: current project id
    :return: It depends on the validity of typed in username
    """
    try:
        username = request.form.get('username', " ")
        authority = "overwrite"
        pending_authority = "pending|{}".format(authority)
        account_user = User.query.filter_by(username=username).first()
        if username != current_user.username:
            if account_user is not None:
                    # create permission:
                project = Permission.query.filter_by(project_id=project_id).first()
                permission_projectid = "{}{}{}{}".format(current_user.username, username, project.project, authority)
                permission_existed = Permission.query.filter_by(project_id=permission_projectid).first()
                if permission_existed:
                    return redirect(url_for("project_profile", project_id=project_id, msg="Permission existed!"))
                else:
                    new_permission = Permission(project_id=permission_projectid, owner=current_user.username, shareTo=username,
                                                project=project.project, status=pending_authority)
                    db.session.add(new_permission)
                    db.session.commit()

                    return redirect(url_for("project_profile", project_id=project_id, msg="Permission successfully created!"))
            else:
                return redirect(url_for("project_profile", project_id=project_id, msg="User not found!"))
        else:
            return redirect(url_for("project_profile", project_id=project_id, msg="You cannot give permission to yourself!"))

    except:
        return redirect(url_for("project_profile", project_id=project_id, msg="fail to create permission"))


@app.route('/instructor_project', methods=["POST", "GET"])
@login_required
def instructor_project():
    """
    Load All project and shared project from database for the current user
    :return: a rendered template with all the projects the current user has
    """
    list_of_all_projects = Permission.query.filter_by(shareTo=current_user.username).all()
    list_of_personal_projects = Permission.query.filter_by(owner=current_user.username,
                                                           shareTo=current_user.username).all()
    list_of_shared_project = []
    for project in list_of_all_projects:
        flag = True
        for personal_project in list_of_personal_projects:
            if project.project_id == personal_project.project_id:
                flag = False
        if flag:
            list_of_shared_project.append(project)

    list_of_personal_project_database = {}
    list_of_shared_project_database = {}
    # load the description of project
    for personal_project in list_of_personal_projects:
        project_in_project_db = Project.query.filter_by(project_name=personal_project.project,
                                                        owner=personal_project.owner).first()
        list_of_personal_project_database[project_in_project_db.project_name] = project_in_project_db
    for shared_project in list_of_shared_project:
        project_in_project_db = Project.query.filter_by(project_name=shared_project.project,
                                                        owner=shared_project.owner).first()
        list_of_shared_project_database[project_in_project_db.project_name] = project_in_project_db
    return render_template('instructor_project.html', personal_project_list=list_of_personal_projects,
                           shared_project_list=list_of_shared_project,
                           list_of_personal_project_database=list_of_personal_project_database,
                           list_of_shared_project_database=list_of_shared_project_database)


@app.route('/create_project', methods=["POST", "GET"])
@login_required
def create_project():
    """
    # Request from file by WTF
    # Create a new project folder under 'path_to_current_user'
    # save files in new folder and build a evaluation doc depending on json file

    :return:
    """

    path_to_current_user = "{}/{}".format(base_directory, current_user.username)
    path_to_student_file = "{}/student.xlsx".format(path_to_current_user)
    path_to_json_file = "{}/TW.json".format(path_to_current_user)
    form = ProjectForm()
    try:
        if form.validate_on_submit():
            # create project folder
            path_to_current_user_project = "{}/{}/{}".format(base_directory, current_user.username, form.project_name.data)
            os.mkdir(path_to_current_user_project)

            path_to_student_file_stored = "{}/student.xlsx".format(path_to_current_user_project)
            shutil.move(path_to_student_file, path_to_student_file_stored)
            path_to_json_file_stored = "{}/TW.json".format(path_to_current_user_project)
            shutil.move(path_to_json_file, path_to_json_file_stored)
            # creating evaluation doc based on grading criteria json file

            # copy student sheet to evaluation doc
            student_file_workbook = openpyxl.load_workbook(path_to_student_file_stored)
            student_file_worksheet = student_file_workbook['Sheet1']

            # create group file depending on student file
            list_of_group = select_by_col_name('group', student_file_worksheet)
            set_of_group = set(list_of_group)

            # Fixing a bug where a None element was found. Is this safe?
            set_of_group.discard(None)

            # create a group workbook
            path_to_group_file = "{}/group.xlsx".format(path_to_current_user_project)
            group_workbook = openpyxl.Workbook()
            group_file_worksheet = group_workbook.create_sheet('Sheet1')
            meta_file_worksheet = group_workbook.create_sheet('Sheet2')
            # all student information map
            student_map_list = []
            for student_index in range(2, len(list(student_file_worksheet.iter_rows())) + 1):
                student_map_list.append(select_map_by_index(student_index, student_file_worksheet))
            # insert group columns
            group_file_worksheet.cell(1, 1).value = 'groupid'
            meta_file_worksheet.cell(1, 1).value = 'groupid'
            meta_file_worksheet.cell(1, 2).value = 'metaid'
            start_index = 2
            max_num_students_pergroup = 0
            for group in set_of_group:
                group_file_worksheet.cell(start_index, 1).value = group
                student_emails = [x['Email'] for x in student_map_list if x['group'] == group]
                if len(student_emails) > max_num_students_pergroup:
                    max_num_students_pergroup = len(student_emails)
                meta_file_worksheet.cell(start_index, 1).value = group
                meta_group = [x['meta'] for x in student_map_list if x['group'] == group][0]
                meta_file_worksheet.cell(start_index, 2).value = meta_group
                for insert_index in range(2, len(student_emails) + 2):
                    group_file_worksheet.cell(start_index, insert_index).value = student_emails[insert_index - 2]
                start_index += 1
            for index in range(1,max_num_students_pergroup+1):
                group_file_worksheet.cell(1, 1+index).value = ("student" + str(index))
            group_workbook.save(path_to_group_file)

            path_to_evaluation = "{}/evaluation.xlsx".format(path_to_current_user_project)
            evaluation_workbook = openpyxl.Workbook()
            evaluation_group = evaluation_workbook.create_sheet('group')
            evaluation_meta = evaluation_workbook.create_sheet('meta')
            evaluation_student = evaluation_workbook.create_sheet('students')
            copy_all_worksheet(evaluation_group, group_file_worksheet)
            copy_all_worksheet(evaluation_meta, meta_file_worksheet)
            copy_all_worksheet(evaluation_student, student_file_worksheet)
            # create EVA depending on the json file
            evaluation_eva = evaluation_workbook.create_sheet('eva')
            # open json file and load json
            myLock = FileLock(path_to_json_file_stored+'.lock', timeout = 5)
            with myLock:
                with open(path_to_json_file_stored, 'r')as f:
                    json_data = json.loads(f.read(), strict=False)
            # The group id, eva_name, date are defults
            tags_to_append = ['group_id', 'eva_name', 'owner', 'date', 'students']
            for category in json_data['category']:
                category_name = (category['name'])
                for section in category['section']:
                    # instructors don't care about the text value, the text values will only be send to students.
                    if section['type'] != 'text':
                        value_to_append = "{}|{}".format(category_name, section['name'])
                        tags_to_append.append(value_to_append)
            tags_to_append.append("comment")
            tags_to_append.append("last_updates")
            evaluation_eva.append(tags_to_append)

            evaluation_workbook.save(path_to_evaluation)

            # create permission to owener himself
            project_id = "{}{}{}{}".format(current_user.username, current_user.username, form.project_name.data, 'full')
            self_permission = Permission(project_id=project_id, owner=current_user.username, shareTo=current_user.username,
                                         project=form.project_name.data, status='full')
            db.session.add(self_permission)
            db.session.commit()

            # create the project in database
            project_to_add = Project(project_name=form.project_name.data, project_status='public',
                                     owner=current_user.username, description=form.project_description.data)
            db.session.add(project_to_add)
            db.session.commit()

            return redirect(url_for("instructor_project"))


        else:
            if os.path.exists(path_to_student_file):
                os.remove(path_to_student_file)
            if os.path.exists(path_to_json_file):
                os.remove(path_to_json_file)
            return render_template('create_project.html', form=form, alert="")

    except Exception as e:
        if os.path.exists(path_to_student_file):
            os.remove(path_to_student_file)
        if os.path.exists(path_to_json_file):
            os.remove(path_to_json_file)
        return render_template("create_project.html", form=form, alert=e)


def copy_all_worksheet(copy_to, copy_from):
    for row in range(0, len(list(copy_from.iter_rows()))):
        for col in range(0, len(list(copy_from.iter_cols()))):
            copy_to.cell(row=row + 1, column=col + 1).value = copy_from.cell(row=row + 1, column=col + 1).value


@app.route('/create_project_by_share/<string:project_id>', methods=["POST", "GET"])
@login_required
def create_project_by_share(project_id):
    """

    :param project_id:
    :return:
    """
    new_project_name = request.form['project_name']
    duplicate_project_name = Project.query.filter_by(project_name=new_project_name, owner=current_user.username).first()
    if duplicate_project_name is not None:
        return redirect(url_for('account', msg="This rubric name has been used before"))
    path_to_current_user_project = "{}/{}/{}".format(base_directory, current_user.username, new_project_name)
    # copy json file:
    project = Permission.query.filter_by(project_id=project_id).first()
    if project is not None:
        owner = project.owner
        project_name = project.project
        path_to_json_file = "{}/{}/{}/TW.json".format(base_directory, owner, project_name)#use project name and project owner info to locate the path of json
        path_to_json_file_stored = "{}/TW.json".format(path_to_current_user_project)
        if os.path.exists(path_to_json_file):
            os.mkdir(path_to_current_user_project)
            shutil.copy2(path_to_json_file, path_to_json_file_stored)
        else:
            return redirect('account', msg="the rubric you were trying to copy has been deleted")
    else:
        return redirect('account', msg="the rubric you were trying to copy has been deleted")

    new_project_desc = request.form['project_desc']
    student_file = request.files['student_file']
    path_to_student_file_stored = "{}/student.xlsx".format(path_to_current_user_project)
    student_file.save(path_to_student_file_stored)

    #check if the student file is valid:
    student_file_workbook = load_workbook(path_to_student_file_stored)
    student_file_worksheet = student_file_workbook['Sheet1']
    find_Student = True if 'Student' in [x.value for x in list(student_file_worksheet.iter_rows())[0]] else False
    find_Email = True if 'Email' in [x.value for x in list(student_file_worksheet.iter_rows())[0]] else False
    find_group = True if 'group' in [x.value for x in list(student_file_worksheet.iter_rows())[0]] else False
    find_meta_group = True if 'meta' in [x.value for x in list(student_file_worksheet.iter_rows())[0]] else False

    if find_Student is False:
        return redirect('account', msg="no Student column in student file!")
    if find_Email is False:
        return redirect('account', msg="no Email column in student file!")
    if find_group is False:
        return redirect('account', msg="no group column in student file!")
    if find_meta_group is False:
        return redirect('account', msg="no meta group column in student file!")

    #create project:
    # create group file depending on student file
    list_of_group = select_by_col_name('group', student_file_worksheet)
    set_of_group = set(list_of_group)

    # Fixing a bug where a None element was found. Is this safe?
    set_of_group.discard(None)

    # create a group workbook
    path_to_group_file = "{}/group.xlsx".format(path_to_current_user_project)
    group_workbook = openpyxl.Workbook()
    group_file_worksheet = group_workbook.create_sheet('Sheet1')
    meta_file_worksheet = group_workbook.create_sheet('Sheet2')
    # all student information map
    student_map_list = []
    for student_index in range(2, len(list(student_file_worksheet.iter_rows())) + 1):
        student_map_list.append(select_map_by_index(student_index, student_file_worksheet))
    # insert group columns
    group_file_worksheet.cell(1, 1).value = 'groupid'
    meta_file_worksheet.cell(1, 1).value = 'groupid'
    meta_file_worksheet.cell(1, 2).value = 'metaid'
    start_index = 2
    for group in set_of_group:
        group_file_worksheet.cell(start_index, 1).value = group
        student_emails = [x['Email'] for x in student_map_list if x['group'] == group]
        meta_file_worksheet.cell(start_index, 1).value = group
        meta_group = [x['meta'] for x in student_map_list if x['group'] == group][0]
        meta_file_worksheet.cell(start_index, 2).value = meta_group
        for insert_index in range(2, len(student_emails) + 2):
            group_file_worksheet.cell(start_index, insert_index).value = student_emails[insert_index - 2]
        start_index += 1
    group_workbook.save(path_to_group_file)

    path_to_evaluation = "{}/evaluation.xlsx".format(path_to_current_user_project)
    evaluation_workbook = openpyxl.Workbook()
    evaluation_group = evaluation_workbook.create_sheet('group')
    evaluation_meta = evaluation_workbook.create_sheet('meta')
    evaluation_student = evaluation_workbook.create_sheet('students')
    copy_all_worksheet(evaluation_group, group_file_worksheet)
    copy_all_worksheet(evaluation_meta, meta_file_worksheet)
    copy_all_worksheet(evaluation_student, student_file_worksheet)
    # create EVA depending on the json file
    evaluation_eva = evaluation_workbook.create_sheet('eva')
    # open json file and load json
    myLock = FileLock(path_to_json_file_stored+'.lock', timeout = 5)
    with myLock:
        with open(path_to_json_file_stored, 'r')as f:
            json_data = json.loads(f.read(), strict=False)
    # The group id, eva_name, date are defults
    tags_to_append = ['group_id', 'eva_name', 'owner', 'date', 'students']
    for category in json_data['category']:
        category_name = (category['name'])
        for section in category['section']:
            # instructors don't care about the text value, the text values will only be send to students.
            if section['type'] != 'text':
                value_to_append = "{}|{}".format(category_name, section['name'])
                tags_to_append.append(value_to_append)
    tags_to_append.append("comment")
    tags_to_append.append("last_updates")
    evaluation_eva.append(tags_to_append)

    evaluation_workbook.save(path_to_evaluation)

    # create permission to owner himself
    project_id = "{}{}{}{}".format(current_user.username, current_user.username, new_project_name, 'full')
    self_permission = Permission(project_id=project_id, owner=current_user.username, shareTo=current_user.username,
                                 project=new_project_name, status='full')
    db.session.add(self_permission)
    db.session.commit()

    # create the project in database
    project_id = "{}{}".format(current_user.username, new_project_name)
    project_to_add = Project(project_name=new_project_name, project_status='public',
                             owner=current_user.username, description=new_project_desc)
    db.session.add(project_to_add)
    db.session.commit()

    return redirect(url_for("instructor_project"))


@app.route('/create_project_by_share_name_and_owner/<string:type>/<string:project_name>/<string:project_owner>', methods=["POST", "GET"])
@login_required
def create_project_by_share_name_and_owner(type, project_name, project_owner):

    new_project_name = request.form['project_name']
    duplicate_project_name = Project.query.filter_by(project_name=new_project_name, owner=current_user.username).first()
    if duplicate_project_name is not None:
        msg = "This rubric name has been used before"
    path_to_current_user_project = "{}/{}/{}".format(base_directory, current_user.username, new_project_name)
    if type == "Share":
        path_to_json_file = "{}/{}/{}/TW.json".format(base_directory, project_owner, project_name)#jacky: use project name and project owner info to locate the path of json?
    else:
        path_to_json_file = "{}/{}/{}".format(home_directory, "Default", project_name)
    path_to_json_file_stored = "{}/TW.json".format(path_to_current_user_project)
    if os.path.exists(path_to_json_file):
        os.mkdir(path_to_current_user_project)
        shutil.copy2(path_to_json_file, path_to_json_file_stored)
    else:
        return redirect(url_for('account', msg="the rubric you were trying to copy has been deleted"))
    new_project_desc = request.form['project_desc']
    student_file = request.files['student_file']
    path_to_student_file_stored = "{}/student.xlsx".format(path_to_current_user_project)
    student_file.save(path_to_student_file_stored)

    #check if the student file is valid:
    student_file_workbook = load_workbook(path_to_student_file_stored)
    student_file_worksheet = student_file_workbook['Sheet1']
    find_Student = True if 'Student' in [x.value for x in list(student_file_worksheet.iter_rows())[0]] else False
    find_Email = True if 'Email' in [x.value for x in list(student_file_worksheet.iter_rows())[0]] else False
    find_group = True if 'group' in [x.value for x in list(student_file_worksheet.iter_rows())[0]] else False
    find_meta_group = True if 'meta' in [x.value for x in list(student_file_worksheet.iter_rows())[0]] else False

    if find_Student is False:
        return redirect('account', msg="no Student column in student file!")
    if find_Email is False:
        return redirect('account', msg="no Email column in student file!")
    if find_group is False:
        return redirect('account', msg="no group column in student file!")
    if find_meta_group is False:
        return redirect('account', msg="no meta group column in student file!")

    #create project:
    # create group file depending on student file
    list_of_group = select_by_col_name('group', student_file_worksheet)
    set_of_group = set(list_of_group)
    # create a group workbook
    path_to_group_file = "{}/group.xlsx".format(path_to_current_user_project)
    group_workbook = openpyxl.Workbook()
    group_file_worksheet = group_workbook.create_sheet('Sheet1')
    meta_file_worksheet = group_workbook.create_sheet('Sheet2')
    # all student information map
    student_map_list = []
    for student_index in range(2, len(list(student_file_worksheet.iter_rows())) + 1):
        student_map_list.append(select_map_by_index(student_index, student_file_worksheet))
    # insert group columns
    group_file_worksheet.cell(1, 1).value = 'groupid'
    meta_file_worksheet.cell(1, 1).value = 'groupid'
    meta_file_worksheet.cell(1, 2).value = 'metaid'
    start_index = 2
    for group in set_of_group:
        group_file_worksheet.cell(start_index, 1).value = group
        student_emails = [x['Email'] for x in student_map_list if x['group'] == group]
        meta_file_worksheet.cell(start_index, 1).value = group
        meta_group = [x['meta'] for x in student_map_list if x['group'] == group][0]
        meta_file_worksheet.cell(start_index, 2).value = meta_group
        for insert_index in range(2, len(student_emails) + 2):
            group_file_worksheet.cell(start_index, insert_index).value = student_emails[insert_index - 2]
        start_index += 1
    group_workbook.save(path_to_group_file)

    path_to_evaluation = "{}/evaluation.xlsx".format(path_to_current_user_project)
    evaluation_workbook = openpyxl.Workbook()
    evaluation_group = evaluation_workbook.create_sheet('group')
    evaluation_meta = evaluation_workbook.create_sheet('meta')
    evaluation_student = evaluation_workbook.create_sheet('students')
    copy_all_worksheet(evaluation_group, group_file_worksheet)
    copy_all_worksheet(evaluation_meta, meta_file_worksheet)
    copy_all_worksheet(evaluation_student, student_file_worksheet)
    # create EVA depending on the json file
    evaluation_eva = evaluation_workbook.create_sheet('eva')
    # open json file and load json
    myLock = FileLock(path_to_json_file_stored+'.lock', timeout = 5)
    with myLock:
        with open(path_to_json_file_stored, 'r')as f:
            json_data = json.loads(f.read(), strict=False)
    # The group id, eva_name, date are defults
    tags_to_append = ['group_id', 'eva_name', 'owner', 'date', 'students']
    for category in json_data['category']:
        category_name = (category['name'])
        for section in category['section']:
            # instructors don't care about the text value, the text values will only be send to students.
            if section['type'] != 'text':
                value_to_append = "{}|{}".format(category_name, section['name'])
                tags_to_append.append(value_to_append)
    tags_to_append.append("comment")
    tags_to_append.append("last_updates")
    evaluation_eva.append(tags_to_append)

    evaluation_workbook.save(path_to_evaluation)

    # create permission to owener himself
    project_id = "{}{}{}{}".format(current_user.username, current_user.username, new_project_name, 'full')
    self_permission = Permission(project_id=project_id, owner=current_user.username, shareTo=current_user.username,
                                 project=new_project_name, status='full')
    db.session.add(self_permission)
    db.session.commit()

    # create the project in database
    # project_id = "{}{}".format(current_user.username, new_project_name)
    project_to_add = Project(project_name=new_project_name, project_status='public',
                             owner=current_user.username, description=new_project_desc)
    db.session.add(project_to_add)
    db.session.commit()

    return redirect(url_for("instructor_project"))

@app.route('/load_project/<string:project_id>/<string:msg>', methods=["GET"])
@login_required
def load_project(project_id, msg):
    # get project by project_id
    project = Permission.query.filter_by(project_id=project_id).first()
    path_to_evaluation_xlsx = "{}/{}/{}/evaluation.xlsx".format(base_directory, project.owner, project.project)
    evaluation_workbook = openpyxl.load_workbook(path_to_evaluation_xlsx)
    meta_worksheet = evaluation_workbook['meta']
    set_of_eva = Evaluation.query.filter_by(project_name=project.project, project_owner=project.owner).all()
    set_of_meta = set(select_by_col_name('metaid', meta_worksheet))

    return render_template("project.html", project=project, data_of_eva_set=set_of_eva, set_of_meta=set_of_meta,
                           msg=msg, useremail=current_user.email)

@app.route('/create_evaluation/<string:project_id>', methods=['GET', 'POST'])
@login_required
def create_evaluation(project_id):
    # get project by id
    project = Permission.query.filter_by(project_id=project_id).first()
    # load group columns and evaluation worksheet
    evaluation_name = request.form['evaluation_name']
    evaluation_name_find_in_db = Evaluation.query.filter_by(project_owner=current_user.username,
                                                            project_name=project.project,
                                                            eva_name=evaluation_name).first()
    # give each new evaluation a default value of sending record
    record_existence = EmailSendingRecord.query.filter_by(project_name=project.project,
                                                          project_owner=current_user.username,
                                                          eva_name=evaluation_name).first()
    if record_existence is None:
        new_record = EmailSendingRecord(project_name=project.project,
                                        project_owner=current_user.username,
                                        eva_name=evaluation_name,
                                        num_of_tasks=0,
                                        num_of_finished_tasks=0)
        db.session.add(new_record)
        db.session.commit()

    if evaluation_name_find_in_db is None:
        evaluation_desc = request.form.get('evaluation_description', " ")
        path_to_load_project = "{}/{}/{}".format(base_directory, current_user.username, project.project)
        path_to_evaluation_file = "{}/evaluation.xlsx".format(path_to_load_project)
        eva_workbook = load_workbook(path_to_evaluation_file)
        group_worksheet = eva_workbook['group']
        eva_worksheet = eva_workbook['eva']
        meta_worksheet = eva_workbook['meta']
        students_worksheet = eva_workbook['students']

        group_col = []
        for col_item in list(group_worksheet.iter_cols())[0]:
            if col_item.value != "groupid":
                group_col.append(col_item.value)

        # get all students by students
        students = get_students_by_group(group_worksheet, students_worksheet)
        # create a empty row for each group in the new evaluation
        for group in group_col:
            students_name = []
            # couple is [email, student_name]
            for student_couple in students[str(group)]:
                students_name.append(student_couple[1])
            row_to_insert = new_row_generator(str(group), students_name, evaluation_name, eva_worksheet)
            eva_worksheet.append(row_to_insert)
        eva_workbook.save(path_to_evaluation_file)

        # create evaluation in database:
        evaluation_to_add = Evaluation(eva_name=evaluation_name, project_name=project.project,
                                       project_owner=project.owner, owner=current_user.username,
                                       description=evaluation_desc)
        db.session.add(evaluation_to_add)
        db.session.commit()
        msg = "New Evaluation has been created successfully"

        set_of_meta = set(select_by_col_name('metaid', meta_worksheet))
        return redirect(
            url_for('jump_to_evaluation_page', project_id=project.project_id, evaluation_name=evaluation_name, metaid=set_of_meta.pop(), group="***None***", msg=msg))

    else:
        print(evaluation_name_find_in_db)
        return redirect(url_for('load_project', project_id=project_id, msg="The evaluation_name has been used before"))


@app.route(
    '/jump_to_evaluation_page/<string:project_id>/<string:evaluation_name>/<string:metaid>/<string:group>/<string:msg>',
    methods=["GET", "POST"])
@login_required
def jump_to_evaluation_page(project_id, evaluation_name, metaid, group, msg):
    # get project by project_id
    project = Permission.query.filter_by(project_id=project_id).first()
    # prepare the json data and group numbers before it jumps to evaluation page
    path_to_load_project = "{}/{}/{}".format(base_directory, project.owner, project.project)
    myLock = FileLock(path_to_load_project+'.lock', timeout = 5)
    with myLock:
        with open("{}/TW.json".format(path_to_load_project), 'r')as f:
            json_data = json.loads(f.read(), strict=False)
    excelLock = FileLock("{}/evaluation.xlsx".format(path_to_load_project) + '.lock', timeout=5)
    with excelLock:
        eva_workbook = load_workbook("{}/evaluation.xlsx".format(path_to_load_project))
        group_worksheet = eva_workbook['group']
        students_worksheet = eva_workbook['students']
        meta_worksheet = eva_workbook['meta']

        # data of meta groups
        meta_group_map_list = []
        for group_index in range(2, len(list(meta_worksheet.iter_rows())) + 1):
            meta_group_map_list.append(select_map_by_index(group_index, meta_worksheet))
        group_col = [x['groupid'] for x in meta_group_map_list if str(x['metaid']) == str(metaid)]
        # if only click on meta group, by default choose its first group
        if group == "***None***":
            group = group_col[0]
        # check if evaluation exists in the worksheet
        eva_worksheet = eva_workbook['eva']

        # Transform ROWS in worksheet to DICTIONARY
        new_row = {}
        first_row = list(eva_worksheet.iter_rows())[0]
        for tag in first_row:
            new_row[tag.value] = ""

        temp_eva = select_row_by_group_id("eva_name", evaluation_name, eva_worksheet)

        # dictionary contains all data
        eva_to_edit = {}
        # list contains only owners
        owner_list = []

        #first, convert string to time and then pick the latest update which committed by this user.
        previous_max_date = datetime.min
        active_tab_tuple = ()
        for row in temp_eva:
            if str(group) == str(row['group_id']):
                owner_per_row = str(row['owner'])
                date = str(row['date'])
                date_datetime = datetime.strptime(date, "%Y-%m-%d_%H-%M-%S")
                # tuple will be unique in this evaluation
                tuple = (owner_per_row, date)
                if owner_per_row == current_user.username and date_datetime > previous_max_date:
                    previous_max_date = date_datetime
                    active_tab_tuple = tuple
                owner_list.append(tuple)
                eva_to_edit[tuple] = row

        if len(active_tab_tuple) == 0:
            active_tab_tuple = owner_list[0]

        students = get_students_by_group(group_worksheet, students_worksheet)
        for temp_group in group_col:
            index = select_index_by_group_eva(evaluation_name, temp_group, eva_worksheet)
            students_attendence = (eva_worksheet.cell(index[0], 5).value).split("|")
            print(students_attendence)
            for temp_student in students[temp_group]:
                if temp_student[1] in students_attendence:
                    students[temp_group][students[temp_group].index(temp_student)].append(1)
                else:
                    students[temp_group][students[temp_group].index(temp_student)].append(0)

        path_to_evaluation_file = "{}/{}/{}/evaluation.xlsx".format(base_directory, project.owner, project.project)
        evaluation_workbook = openpyxl.load_workbook(path_to_evaluation_file)
        evaluation_worksheet = evaluation_workbook['eva']
        meta_worksheet = evaluation_workbook['meta']
        list_of_eva = select_by_col_name('eva_name', evaluation_worksheet)
        set_of_eva = set(list_of_eva)
        dic_of_eva = {}

        dic_of_choosen = {}
        set_of_meta = set(select_by_col_name('metaid', meta_worksheet))

        meta_group_map_list = []
        for group_index in range(2, len(list(meta_worksheet.iter_rows())) + 1):
            meta_group_map_list.append(select_map_by_index(group_index, meta_worksheet))

        for eva in set_of_eva:
            all_groups_choosen = set()
            all_groups_not_choosen = set()
            all_groups = set()
            choosen = {}
            for meta in set_of_meta:
                choosen[meta] = set()
            total = {}
            notchoosen = {}
            for meta in set_of_meta:
                notchoosen[meta] = set([x['groupid'] for x in meta_group_map_list if str(x['metaid']) == str(meta)])
                total[meta] = set([x['groupid'] for x in meta_group_map_list if str(x['metaid']) == str(meta)])
            # update 9/13: simple profile
            dic_of_eva[eva] = []
            temp_eva = select_row_by_group_id("eva_name", eva, evaluation_worksheet)
            for eva_row in temp_eva:
                dic_of_eva[eva].append(eva_row)
                for (key, value) in eva_row.items():
                    if (key != "group_id") and (key != "eva_name") and (key != "owner") and (key != "date") and (
                            key != "students") and (key != "last_updates"):
                        if (value is not None) and (value != " ") and (value != ""):
                            meta = [x[0] for x in list(total.items()) if eva_row["group_id"] in x[1]][0]
                            choosen[meta].add(eva_row["group_id"])
                            notchoosen[meta].discard(eva_row["group_id"])
            for meta in set_of_meta:
                for choosen_i in choosen[meta]:
                    all_groups_choosen.add(choosen_i)
                for notchoosen_j in notchoosen[meta]:
                    all_groups_not_choosen.add(notchoosen_j)
                for all_k in total[meta]:
                    all_groups.add(all_k)
            dic_of_choosen[eva] = [choosen, notchoosen, total, all_groups_choosen, all_groups_not_choosen, all_groups]
        listOfGroups = dic_of_choosen[evaluation_name][0][metaid]
        listOfGroupss = []
        if listOfGroups == set():
            listOfGroups = listOfGroupss
        else:
            for i in listOfGroups:
                listOfGroupss.append(i)
            listOfGroups=listOfGroupss
        return render_template("evaluation_page.html", project=project, json_data=json_data, group=group, metaid=metaid,
                               group_col=group_col, set_of_meta=set_of_meta, msg=msg, evaluation_name=evaluation_name,
                               edit_data=eva_to_edit, owner_list=owner_list, students=students,
                               current_user=current_user.username,
                               active_tab_tuple=active_tab_tuple,
                               listOfGroups = listOfGroups)


def dumper(obj):
    try:
        return obj.toJSON()
    except:
        return obj.__dict__
@app.route(
    '/evaluation_receiver/<project_id>/<string:evaluation_name>/<string:metaid>/<string:group>/<string:owner>/<string:past_date>',
    methods=["GET", "POST"])
@login_required
def evaluation_page(project_id, evaluation_name, metaid, group, owner, past_date):
    # receive all the data and insert them into xlsx
    # group id, evaluation name, date time is constant
    row_to_insert = []
    group_id = group
    submit_type = request.form['submit_button']
    date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    row_to_insert.append(group_id)
    row_to_insert.append(evaluation_name)
    # if edit, create a new row with owner = current user
    if submit_type == 'edit':
        row_to_insert.append(current_user.username)
    else:
        row_to_insert.append(owner)
    row_to_insert.append(date)
    student_list = request.form.getlist('student')
    if len(student_list) > 0:
        row_to_insert.append("|".join(student_list))
    else:
        row_to_insert.append(" ")

    # The rest are variables from TW
    # get project by project_id
    project = Permission.query.filter_by(project_id=project_id).first()
    path_to_load_project = "{}/{}/{}".format(base_directory, project.owner, project.project)
    myLock = FileLock(path_to_load_project+'.lock', timeout = 5)
    with myLock:
        with open("{}/TW.json".format(path_to_load_project), 'r')as f:
            json_data = json.loads(f.read(), strict=False)
    for category in json_data['category']:
        category_name = category['name']
        for section in category['section']:
            section_name = section['name']
            section_name = '{}{}|{}|{}'.format(owner, past_date, category_name, section_name)
            if section['type'] == 'radio':
                if submit_type == 'create':
                    value = request.form.get('{}|{}'.format(category_name, section['name']), " ")
                    row_to_insert.append(value)
                else:
                    value = request.form.get(section_name, " ")
                    row_to_insert.append(value)
            elif section['type'] == 'checkbox':
                if submit_type == 'create':
                    value = request.form.getlist('{}|{}'.format(category_name, section['name']))
                    if len(value) != 0:
                        value = '|'.join(value)
                    else:
                        value = " "
                    row_to_insert.append(value)
                else:
                    value = request.form.getlist(section_name)
                    if len(value) != 0:
                        value = '|'.join(value)
                    else:
                        value = " "
                    row_to_insert.append(value)
            else:
                # text don't need to be saved
                print('to be continued')
    path_to_evaluation_file = "{}/evaluation.xlsx".format(path_to_load_project)
    excelLock = FileLock(path_to_evaluation_file + '.lock', timeout=5)
    with excelLock:
        evaluation_workbook = load_workbook(path_to_evaluation_file)
        evaluation_worksheet = evaluation_workbook['eva']
        # change the last update by append the current user according to submit type
        if submit_type == 'update':
            index = int(
                select_index_by_group_eva_owner_date(evaluation_name, group_id, owner, past_date, evaluation_worksheet))
            # delete the old row by index
            last_comment = select_by_col_name('comment', evaluation_worksheet)[index - 2]
            comment = request.form.get('{}{}|comment'.format(owner, past_date), " ")
            if comment != " " and last_comment != " ":
                comment = "{}|{}".format(last_comment, comment)
            else:
                comment = last_comment
            row_to_insert.append(comment)
            last_update = select_by_col_name('last_updates', evaluation_worksheet)[index - 2]
            row_to_insert.append(last_update)
            evaluation_worksheet.delete_rows(index, 1)
            evaluation_worksheet.append(row_to_insert)

        elif submit_type == 'create':
            comment = request.form.get('comment', " ")
            row_to_insert.append(comment)
            last_update = current_user.username
            row_to_insert.append(last_update)
            evaluation_worksheet.append(row_to_insert)

        elif submit_type == 'edit':
            comment = request.form.get('{}{}|comment'.format(owner, past_date), " ")
            row_to_insert.append(comment)
            last_update = current_user.username
            row_to_insert.append(last_update)
            evaluation_worksheet.append(row_to_insert)
        elif submit_type == 'overwrite':
            index = int(
                select_index_by_group_eva_owner_date(evaluation_name, group_id, owner, past_date, evaluation_worksheet))
            last_comment = select_by_col_name('comment', evaluation_worksheet)[index - 2]
            comment = request.form.get('{}{}|comment'.format(owner, past_date), " ")
            if comment != " " and last_comment != " ":
                comment = "{}|{}".format(last_comment, comment)
            else:
                comment = last_comment
            row_to_insert.append(comment)
            last_update = select_by_col_name('last_updates', evaluation_worksheet)[index - 2]
            last_update = "{}|{}".format(last_update, current_user.username)
            row_to_insert.append(last_update)
            evaluation_worksheet.delete_rows(index, 1)
            evaluation_worksheet.append(row_to_insert)

        # save the workbook
        evaluation_workbook.save(path_to_evaluation_file)

        # change the last edit
        evaluation_in_database = Evaluation.query.filter_by(project_name=project.project, project_owner=project.owner,
                                                            eva_name=evaluation_name).first()
        evaluation_in_database.last_edit = current_user.username
        db.session.commit()
        msg = "The grade has been updated successfully"

        return redirect(
            url_for('jump_to_evaluation_page', project_id=project_id, evaluation_name=evaluation_name, metaid=metaid,
                    group=group, owner=owner, msg=msg))


@app.route(
    '/evaluation_commit/<project_id>/<string:evaluation_name>/<string:metaid>/<string:group>/<string:owner>/<string:past_date>/<string:category>',
    methods=["POST"])
@login_required
def evaluation_commit(project_id, evaluation_name, metaid, group, owner, past_date, category):
    try:
        # receive all the data and insert them into xlsx
        # group id, evaluation name, date time is constant
        row_to_insert = []
        group_id = group
        #date = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        # The rest are variables from TW
        # get project by project_id
        project = Permission.query.filter_by(project_id=project_id).first()
        path_to_load_project = "{}/{}/{}".format(base_directory, project.owner, project.project)
        myLock = FileLock(path_to_load_project+'.lock', timeout = 5)
        with myLock:
            with open("{}/TW.json".format(path_to_load_project), 'r')as f:
                json_data = json.loads(f.read(), strict=False)
        category_to_commit = [x for x in json_data['category'] if x['name'] == category][0]
        category_name = category_to_commit['name']
        for section in category_to_commit['section']:
            section_name = section['name']
            section_name = '{}{}|{}|{}'.format(owner, past_date, category_name, section_name)
            if section['type'] == 'radio':
                value = request.form.get(section_name, " ")
                row_to_insert.append(value)
            elif section['type'] == 'checkbox':
                value = request.form.getlist(section_name)
                if len(value) != 0:
                    value = '|'.join(value)
                else:
                    value = " "
                row_to_insert.append(value)
            else:
                # text don't need to be saved
                print('to be continued')
        path_to_evaluation_file = "{}/evaluation.xlsx".format(path_to_load_project)
        excelLock = FileLock(path_to_evaluation_file + '.lock', timeout=5)
        with excelLock:
            evaluation_workbook = load_workbook(path_to_evaluation_file)
            evaluation_worksheet = evaluation_workbook['eva']
            # change the last update by append the current user according to submit type

            index = int(
                select_index_by_group_eva_owner_date(evaluation_name, group_id, owner, past_date, evaluation_worksheet))

            last_update = select_by_col_name('last_updates', evaluation_worksheet)[index - 2]
            if last_update != current_user.username:
                last_update = "{}|{}".format(last_update, current_user.username)
            # count the index of category
            # 1,2,3 unchanged
            #evaluation_worksheet.cell(index, 4).value = date
            start_point = 6 + (list(x['name'] for x in json_data['category']).index(category))*len(row_to_insert)

            for i in range(0, len(row_to_insert)):
                # f.write(str(index) + " " + str(i) + str(row_to_insert[i]) + '\n')
                evaluation_worksheet.cell(index, start_point + i).value = row_to_insert[i]


            # save the workbook
            evaluation_workbook.save(path_to_evaluation_file)

            # change the last edit
            evaluation_in_database = Evaluation.query.filter_by(project_name=project.project, project_owner=project.owner,
                                                                eva_name=evaluation_name).first()
            evaluation_in_database.last_edit = current_user.username
            db.session.commit()
            msg = "The grade has been updated successfully"

            return jsonify({'success': 'success'})
    except Exception as e:
        return jsonify({'error': e})

@app.route(
    '/attendence_commit/<project_id>/<string:evaluation_name>/<string:metaid>/<string:group>',
    methods=["POST"])
@login_required
def attendence_commit(project_id, evaluation_name, metaid, group):
    project = Permission.query.filter_by(project_id=project_id).first()
    path_to_load_project = "{}/{}/{}".format(base_directory, project.owner, project.project)
    path_to_evaluation_file = "{}/evaluation.xlsx".format(path_to_load_project)
    attendence = request.form.getlist("student")
    if len(attendence) != 0:
        attendence_string = "|".join(attendence)
    else:
        attendence_string = " "
    excelLock = FileLock(path_to_evaluation_file + '.lock', timeout=5)
    with excelLock:
        evaluation_workbook = load_workbook(path_to_evaluation_file)
        evaluation_worksheet = evaluation_workbook['eva']
        # change the last update by append the current user according to submit type

        #change attendency in every row which belongs to that group
        list_of_index = select_index_by_group_eva(evaluation_name, group, evaluation_worksheet)
        for index in list_of_index:
            evaluation_worksheet.cell(index, 5).value = attendence_string

        # save the workbook
        evaluation_workbook.save(path_to_evaluation_file)

        # change the last edit
        evaluation_in_database = Evaluation.query.filter_by(project_name=project.project, project_owner=project.owner,
                                                            eva_name=evaluation_name).first()
        evaluation_in_database.last_edit = current_user.username
        db.session.commit()
        msg = "The grade has been updated successfully"

        return jsonify({'success': 'success'})

@app.route('/download_page/<string:project_id>/<string:evaluation_name>/<string:group>/<string:type>/<string:show_score>',
           methods=['GET', 'POST'])
@login_required
# send all grades in the given evaluation
def download_page(project_id, evaluation_name, group, type, show_score):
    """
    Creating an grade report on top of the download_page.html template, it is used in sendEmail function to create grade
    reports for each group
    :param project_id: current project id
    :param evaluation_name: current name of evaluation
    :param group: current group id
    :param type:
    :param show_score: whether or not displaying the score
    :return: creating a html grade report
    """
    # get project by project_id
    project = Permission.query.filter_by(project_id=project_id).first()
    path_to_load_project = "{}/{}/{}".format(base_directory, project.owner, project.project)
    path_to_evaluation_file = "{}/evaluation.xlsx".format(path_to_load_project)
    eva_workbook = load_workbook(path_to_evaluation_file)
    group_worksheet = eva_workbook['group']
    eva_worksheet = eva_workbook['eva']
    students_worksheet = eva_workbook['students']
    myLock = FileLock(path_to_load_project+'.lock', timeout = 5)
    with myLock:
        with open("{}/TW.json".format(path_to_load_project), 'r')as f:
            json_data = json.loads(f.read(), strict=False)
    if type == "normal":
        temp_eva = select_row_by_group_id("eva_name", evaluation_name, eva_worksheet)
        temp_eva_in_group = [x for x in temp_eva if x['group_id'] == group]

        students_in_one_group = get_students_by_group(group_worksheet, students_worksheet)[group]
        msg = "Downloaded grade group{}".format(group)

        return render_template("download_page.html", project=project, json_data=json_data, group=group, msg=msg,
                               evaluation_name=evaluation_name, students=students_in_one_group,
                               grades=temp_eva_in_group, show_score=show_score)


# download xlsx file or html file
@app.route('/download/<string:project_id>/<string:evaluation_name>', methods=['GET', 'POST'])
@login_required
def download(project_id, evaluation_name):

    today = date.today().strftime("%b_%d_%Y")

    project = Permission.query.filter_by(project_id=project_id).first()
    path_to_load_project = "{}/{}/{}".format(base_directory, current_user.username, project.project)
    path_to_evaluation_file = "{}/evaluation.xlsx".format(path_to_load_project)
    eva_workbook = load_workbook(path_to_evaluation_file)
    group_worksheet = eva_workbook['group']
    eva_worksheet = eva_workbook['eva']
    students_worksheet = eva_workbook['students']
    students = get_students_by_group(group_worksheet, students_worksheet)
    myLock = FileLock(path_to_load_project+'.lock', timeout = 5)
    with myLock:
        with open("{}/TW.json".format(path_to_load_project), 'r')as f:
            json_data = json.loads(f.read(), strict=False)
    group_col = []
    for col_item in list(group_worksheet.iter_cols())[0]:
        if col_item.value != "groupid":
            group_col.append(col_item.value)
    filename = "{}_{}_{}.xlsx".format(project.project, evaluation_name, today)
    if evaluation_name == "all_eva":
        try:
            return send_file(path_to_evaluation_file, cache_timeout=1, attachment_filename=filename, as_attachment=True)
            # msg = "Successfully downloaded"
            # return redirect(url_for("project_profile_jumptool", project_id=project_id, msg=msg))
        except Exception as e:
            print(str(e))
    else:
        # response = urllib.request.urlopen(url_for('jump_to_evaluation_page', project_name=project_name, evaluation_name=evaluation_name))
        # html = response.read()
        # print(html)
        new_row = {}
        first_row = list(eva_worksheet.iter_rows())[0]
        for tag in first_row:
            new_row[tag.value] = ""
        # grab the data that used for this evaluation
        owner = request.form['owner']
        temp_eva = select_row_by_group_id("eva_name", evaluation_name, eva_worksheet)
        eva_to_edit = {}
        for group in group_col:
            for row in temp_eva:
                if str(group) == str(row['group_id']) and str(owner) == str(row['owner']):
                    eva_to_edit[str(group)] = row
        # store the data to a html file and send out(download)
        msg = ""
        path_to_html = "{}/{}_{}.html".format(path_to_load_project, project.project, evaluation_name)
        # remove the old file in case duplicated file existence
        if os.path.exists(path_to_html):
            os.remove(path_to_html)
        myLock = FileLock(path_to_html+'.lock', timeout = 5)
        with myLock:
            with open(path_to_html, 'w') as f:
                f.write(render_template("evaluation_page.html", project=project, json_data=json_data, group_col=group_col,
                                        msg=msg, evaluation_name=evaluation_name, edit_data=eva_to_edit, owner=owner,
                                        students=students))
        return send_file(path_to_html, cache_timeout=1, as_attachment=True)

@app.route('/downloadRubric/<string:type>/<string:name>/<string:owner>', methods= ['GET', 'POST'])
@login_required
def downloadRubric(type, name, owner):
    if type == "default":
        path_to_default_json = "{}/{}/{}".format(home_directory, "Default", name)
        return send_file(path_to_default_json, attachment_filename=name, as_attachment=True)




@app.route('/sendEmail/<string:project_id>/<string:evaluation_name>/<string:show_score>', methods=['GET', 'POST'])
@login_required
def sendEmail(project_id, evaluation_name, show_score):
    """
    This function sends emails to students, part of it uses threadpool to accelerate the process, parameters are passed
    from project_profile.html
    :param project_id: current project's id, it is used to locate the current project in Permission table
    :param evaluation_name: current evaluation name
    :param show_score: a option which user can choose to send out the grade report with scores displayed or not
    :return: return to current page
    """
    project = Permission.query.filter_by(project_id=project_id).first()
    path_to_load_project = "{}/{}/{}".format(base_directory, project.owner, project.project)

    # make a 'evaluation_feedback' folder that stores all evaluations.
    path_to_project_includeEvaluationFolder = "{}/{}/{}/evaluation_feedback".format(base_directory, project.owner, project.project)    
    if not os.path.exists(path_to_project_includeEvaluationFolder):
            os.mkdir(path_to_project_includeEvaluationFolder)
        
   

    path_to_evaluation_file = "{}/evaluation.xlsx".format(path_to_load_project)
    eva_workbook = load_workbook(path_to_evaluation_file)
    group_worksheet = eva_workbook['group']
    eva_worksheet = eva_workbook['eva']
    students_worksheet = eva_workbook['students']
    with open("{}/TW.json".format(path_to_load_project), 'r')as f:
        json_data = json.loads(f.read(), strict=False)
    # data of groups

    group_col = []
    for col_item in list(group_worksheet.iter_cols())[0]:
        if col_item.value != "groupid":
            group_col.append(col_item.value)
    from_email = "Rubric@uiowa.edu"
    students_emails = select_by_col_name("Email", students_worksheet)
    total_num_of_email = len(students_emails)

    current_record = EmailSendingRecord.query.filter_by(project_name=project.project,
                                    project_owner=current_user.username, eva_name=evaluation_name).first()
    current_record.num_of_finished_tasks = -1

    generate_HTML_files(project_id, evaluation_name, group_col,show_score)
    now = datetime.now()
    current_time = now.strftime("%m/%d/%Y at %I:%M:%S")
    current_record = EmailSendingRecord.query.filter_by(project_name=project.project,
                                    project_owner=current_user.username, eva_name=evaluation_name).first()

    current_record.time_email_sent = current_time

    db.session.commit()
    return redirect(url_for('project_profile', project_id=project_id, msg="success"))



# This function serves as a helper function for the 'Send Email' & 'Download Feedback' buttons in the '/project_profile/...' route.
# It generates html files of the rubric. It's implemented so that feedbacks can be previewed (downloaded) before being sent (emailed) with out
# rendering an error.
def generate_HTML_files(project_id,evaluation_name,group_col,show_score):

    project = Permission.query.filter_by(project_id=project_id).first()
    project_path = "{}/{}/{}".format(base_directory, project.owner, project.project)

    isSendEmail = False

    path_to_project_includeEvaluationFolder = "{}/{}/{}/evaluation_feedback".format(base_directory, project.owner, project.project)    
    if not os.path.exists(path_to_project_includeEvaluationFolder):
            os.mkdir(path_to_project_includeEvaluationFolder)

    path_to_evaluation_file = "{}/evaluation.xlsx".format(project_path)
    eva_workbook = load_workbook(path_to_evaluation_file)
    group_worksheet = eva_workbook['group']

    current_record = EmailSendingRecord.query.filter_by(project_name=project.project,
                                    project_owner=current_user.username, eva_name=evaluation_name).first()

    if current_record is None:
        new_record = EmailSendingRecord(project_name=project.project,
                                        project_owner=current_user.username,
                                        eva_name=evaluation_name)
        db.session.add(new_record)
        db.session.commit()
        current_record = EmailSendingRecord.query.filter_by(project_name=project.project,
                                                        project_owner=current_user.username,
                                                        eva_name=evaluation_name).first()
    students_worksheet = eva_workbook['students']
    total_num_of_email = len(select_by_col_name("Email", students_worksheet))

    if current_record.num_of_finished_tasks == -1:
        isSendEmail = True
        current_record.num_of_finished_tasks = 0
        current_record.num_of_tasks = total_num_of_email
        db.session.commit()  



    for group in group_col:
            students_email = select_students_by_group(group, group_worksheet)
            file_name = "{}_{}_{}.html".format(project.project, evaluation_name, group)
            
            path_to_evalutaion_name = "{}/{}/{}/evaluation_feedback/{}".format(base_directory, project.owner, project.project,evaluation_name)    #make directory to place each evaluation under a file named after eva_name
            if not os.path.exists(path_to_evalutaion_name):
                os.mkdir(path_to_evalutaion_name)
            path_to_html = "{}/{}".format(path_to_evalutaion_name, file_name)
            if isSendEmail is True:
                print("FALSE - coming_from_downloadFeedback")
                current_record.num_of_finished_tasks += len(students_email)
            if os.path.exists(path_to_html):
                os.remove(path_to_html)
            with open(path_to_html, 'w') as f:
                f.write(download_page(project.project_id, evaluation_name, group, "normal", show_score))
    db.session.commit()






def send_emails_to_students(group, project, evaluation_name, from_email, path_to_html, students_email, current_record):
    """
    It was extracted from "sendEmail" function so that threadpool can be easily applied, it interacts with the server
    and send out emails, and it prints debugging messages at last
    :param group: current group name
    :param project: current project name
    :param evaluation_name: current evaluation name
    :param from_email: defined in sendEmail function
    :param path_to_html: it points to the grade report html file of the current group, all group memebers share the same
    html grade report
    :param students_email: a list of emails of students in current group
    :param current_record: sending record (row) in the EmailSendingRecord database, contains three parameters
    :return: X
    """
    subject = "grade: {}, {}, {}".format(project.project, evaluation_name, group)
    print('\n\n' + str(subject))
    try:
        index = 0;
        for email in students_email:
            # create an instance of message
            if email is not None:
                subject += str(index)
                index += 1
                myLock = FileLock(path_to_html+'.lock')
                with open(path_to_html, "r") as file_to_html:
                    #subprocess.call(["mail", "-s", subject, "-r", from_email, "-a", path_to_html, email])
                    subprocess.call(["mailx", "-s", "\"TestingGG\"", "elhadie540@hotmail.com"])
                    dateTimeObj = datetime.now()
                    timestampStr = dateTimeObj.strftime("%d-%b-%Y (%H:%M:%S.%f)")
                    current_record.num_of_finished_tasks += 1
                    print("Sent the email to " + email + " at " + timestampStr)
    except Exception as e:
        print('Something went wrong' + str(e))
        msg = ""
        msg += str(e)
        msg += "\n"
        # remove the html file after sending email
        # in case of duplicated file existence
    print("Number of finished tasks is: " + str(current_record.num_of_finished_tasks))
    print("sent to email: " + current_record.last_email)
    if os.path.exists(path_to_html):
        os.remove(path_to_html)
        # if os.path.exists(path_to_pdf):
        #    os.remove(path_to_pdf)
    # return redirect(url_for('project_profile', project_id=project_id, msg="success"))
    # db.session.commit()


@app.route('/account/<string:msg>', methods=['GET', 'POST'])
@login_required
def account(msg):
    #if msg is success, msg = ""
    if msg == "success":
        msg = ""

    #load default json files
    json_list = DefaultRubric.query.all()
    json_data_of_all_default_rubric = {}
    for json_file in json_list:
        path_to_this_json = "{}/{}/{}".format(home_directory, "Default", json_file.json_name)
        with open(path_to_this_json, 'r') as file:
            json_data_of_current_json_file = json.loads(file.read(), strict=False)
        json_data_of_all_default_rubric[json_file.json_name] = json_data_of_current_json_file
    return render_template('account.html', msg=msg, default_json_list=json_list, json_data_of_all_default_rubric=json_data_of_all_default_rubric)


@app.route('/search_project', methods=['POST'])
@login_required
def search_project():
    """
    In the account.html, user will active this function when they search rubric by rubric name
    :return: a list of rubric that has the typed name
    """
# flag_project = True
    #load default json files
    json_list = DefaultRubric.query.all()
    json_data_of_all_default_rubric = {}
    for json_file in json_list:
        path_to_this_json = "{}/{}/{}".format(home_directory, "Default", json_file.json_name)
        with open(path_to_this_json, 'r') as file:
            json_data_of_current_json_file = json.loads(file.read(), strict=False)
        json_data_of_all_default_rubric[json_file.json_name] = json_data_of_current_json_file

    #search project
    project_name = request.form.get('project_name')
    project_items = Project.query.filter_by(project_name = project_name).first()
    if project_items:
        list_of_project = Project.query.filter_by(project_name=project_name).all()
        json_data_of_all_project = {}
        for project in list_of_project:
            path_to_this_project_json = "{}/{}/{}/TW.json".format(base_directory, project.owner, project.project_name)
            with open(path_to_this_project_json, 'r') as file:
                json_data_of_curr_project = json.loads(file.read(), strict=False)
            json_data_of_all_project[project.project_name + project.owner] = json_data_of_curr_project
        return render_template('account.html', msg="", list_of_projects = list_of_project, json_data = json_data_of_all_project, default_json_list=json_list, json_data_of_all_default_rubric=json_data_of_all_default_rubric)
    else:
        return render_template('account.html', msg="can't find this rubirc", project_name=project_name)


@app.route('/search_account', methods=['GET', 'POST'])
@login_required
def search_account():
    """
    search by username
    :return: a list of project the user has
    """
    #load default json files
    json_list = DefaultRubric.query.all()
    json_data_of_all_default_rubric = {}
    for json_file in json_list:
        path_to_this_json = "{}/{}/{}".format(home_directory, "Default", json_file.json_name)
        with open(path_to_this_json, 'r') as file:
            json_data_of_current_json_file = json.loads(file.read(), strict=False)
        json_data_of_all_default_rubric[json_file.json_name] = json_data_of_current_json_file

    #search by account
    account_username = request.form.get('account_username')
    account_user = User.query.filter_by(username=account_username).first()
    if account_user is not None:
        list_of_all_projects = Permission.query.filter_by(shareTo=account_username).all()
        list_of_personal_projects = Permission.query.filter_by(owner=account_username,
                                                               shareTo=account_username).all()
        list_of_shared_project = []
        for project in list_of_all_projects:
            flag = True
            for personal_project in list_of_personal_projects:
                if project.project_id == personal_project.project_id:
                    flag = False
            if flag:
                list_of_shared_project.append(project)

        list_of_personal_project_database = {}
        list_of_shared_project_database = {}
        # load the description of project
        # load json data
        json_data = {}
        project_eva = {}
        for personal_project in list_of_personal_projects:
            project_in_project_db = Project.query.filter_by(project_name=personal_project.project,
                                                            owner=personal_project.owner).first()
            list_of_personal_project_database[project_in_project_db.project_name] = project_in_project_db
            evaluations = Evaluation.query.filter_by(project_name=personal_project.project).all()
            evaluations_names = [x.eva_name for x in evaluations]
            project_eva[personal_project.project] = evaluations_names
            path_to_this_project_json = "{}/{}/{}/TW.json".format(base_directory, personal_project.owner, personal_project.project)
            json_data[personal_project.project_id] = {}
            myLock = FileLock(path_to_this_project_json+'.lock', timeout = 5)
            with myLock:
                with open(path_to_this_project_json, 'r')as f:
                    this_json_data = json.loads(f.read(), strict=False)
                    json_data [personal_project.project_id] = this_json_data
        for shared_project in list_of_shared_project:
            project_in_project_db = Project.query.filter_by(project_name=shared_project.project,
                                                            owner=shared_project.owner).first()
            list_of_shared_project_database[project_in_project_db.project_name] = project_in_project_db
            evaluations = Evaluation.query.filter_by(project_name=personal_project.project).all()
            evaluations_names = [x.eva_name for x in evaluations]
            project_eva[personal_project.project] = evaluations_names
            json_data[shared_project.project_id] = {}
            path_to_this_project_json = "{}/{}/{}/TW.json".format(base_directory, shared_project.owner, shared_project.project)
            myLock = FileLock(path_to_this_project_json+'.lock', timeout = 5)
            with myLock:
                with open(path_to_this_project_json, 'r')as f:
                    this_json_data = json.loads(f.read(), strict=False)
                    json_data [shared_project.project_id] = this_json_data
        return render_template('account.html', msg="", personal_project_list=list_of_personal_projects,
                           shared_project_list=list_of_shared_project,
                           list_of_personal_project_database=list_of_personal_project_database,
                           list_of_shared_project_database=list_of_shared_project_database, project_eva=project_eva, json_data=json_data, default_json_list=json_list, json_data_of_all_default_rubric=json_data_of_all_default_rubric, flag_2=False)
    else:
        msg = "Can't find this user"
        return render_template('account.html', msg=msg)


@app.route('/notification_receiver/<string:notification_id>', methods=['GET', 'POST'])
@login_required
def notification_receiver(notification_id):
    response = request.form['response']
    notification = Notification.query.filter_by(notification_id=notification_id).first()
    permission = Permission.query.filter_by(project_id=notification.appendix).first()
    authority = permission.status.split('|')[1]
    if response == 'Decline':
        db.session.delete(permission)
        db.session.commit()
    else:
        permission.status = authority
        db.session.commit()
    notification.status = 'read'
    db.session.commit()
    return redirect(url_for('account'))


@app.route('/student_dashboard')
@login_required
def student_dashboard():
    return render_template('student_dashboard.html', name=current_user.username)


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

# This route is for the 'Download Feedback' button on the '/project_profile/...' route. It first generates html files of the 
# rubrics, in case the emails haven't been 'sent'. This gives the instructor a preview of what their sending.
@app.route('/downloadFeedBack/<string:project_id>/<string:evaluation_name>/<string:show_score>', methods=['GET', 'POST'])
@login_required
def downloadFeedBack(project_id, evaluation_name, show_score):
    project = Permission.query.filter_by(project_id=project_id).first()         #find project
    
    project_path = "{}/{}/{}".format(base_directory, project.owner, project.project)

    path_to_evaluation_file = "{}/evaluation.xlsx".format(project_path)
    eva_workbook = load_workbook(path_to_evaluation_file)
    group_worksheet = eva_workbook['group']

    group_col = []
    for col_item in list(group_worksheet.iter_cols())[0]:
        if col_item.value != "groupid":
            group_col.append(col_item.value)
  

    generate_HTML_files(project_id, evaluation_name, group_col, show_score)
    

    evaluation_name_find_in_db = Evaluation.query.filter_by(project_owner=current_user.username, project_name=project.project).all()    #look up all evaluations associated with project

    currentDir = os.getcwd();
    os.chdir(currentDir+"/users/{}".format(project.owner))                  # change current directory to /users/username/project/
    currentZipFiles = []
    for zips in glob.glob("feedback_*.zip"):
        currentZipFiles.append(zips)

    for x in currentZipFiles:
        if os.path.exists(os.getcwd()+"/"+str(x)):
            os.remove(os.getcwd()+"/"+str(x))

    os.chdir(currentDir)                                                    # change directory back to what it was previously
    

    pathUP = "{}/{}".format(base_directory, project.owner)    #make directory to place each evaluation under a file named after eva_name

    now = datetime.now().time()
    today = date.today().strftime("%b_%d_%Y")
    current_time = now.strftime("%I_%M_%S")
    filename = "{}_{}_{}.zip".format(project.project,evaluation_name,today)
    fullPath = "{}/{}".format(pathUP,filename)
    if os.path.exists(fullPath):
        os.remove(fullPath)

    
    zipObj = zipfile.ZipFile(fullPath,'w')
    at_least_one_match = False
    path_to_evalutaion_name = "{}/{}/{}/evaluation_feedback/{}".format(base_directory, project.owner, project.project,evaluation_name)    #make directory to place each evaluation under a file named after eva_name

    for roots, _, files in os.walk(path_to_evalutaion_name):
        for file in files:
            filepath = os.path.join(roots,file)
            at_least_one_match = True
            zipObj.write(filepath, os.path.basename(file))
    zipObj.close()

    if not at_least_one_match:
        raise Exception("EVALUATION HAS NOT BEEN SENT. OR IT DOESN'T EXIST IN DB")
    return send_file(fullPath,cache_timeout=1, as_attachment=True)
    #return redirect(url_for('project_profile', project_id=project_id, msg="success"))




# select all elements by the column name
def select_by_col_name(col_name, worksheet):
    first_row = list(worksheet.iter_rows())[0]
    index_of_col = -1
    for variables in first_row:
        if variables.value == col_name:
            index_of_col = first_row.index(variables)
    if index_of_col == -1:
        return "col_name unmatched"
    else:
        col = list(worksheet.iter_cols())[index_of_col]
        data_by_col = []
        for data in range(1, len(col)):
            data_by_col.append(col[data].value)
        return data_by_col


# index is the row index in xsl
def select_map_by_index(index, worksheet):
    list_of_rows = list(worksheet.iter_rows())
    # row_to_return = []
    # for item in list_of_rows[index-1]:
    #     row_to_return.append(item.value)
    row_selected = list_of_rows[index - 1]
    tags = list_of_rows[0]
    map_to_return = {}
    for tag_index in range(0, len(tags)):
        tag_value = tags[tag_index].value
        map_to_return[tag_value] = row_selected[tag_index].value
    return map_to_return


# index is the row index in xsl
def select_row_by_index(index, worksheet):
    list_of_rows = list(worksheet.iter_rows())
    row_to_return = []
    for item in list_of_rows[index - 1]:
        row_to_return.append(item.value)
    return row_to_return


# select by the groupid and eva name
def select_index_by_group_eva(eva_name, group_id, worksheet):
    evaluation_list = select_by_col_name("eva_name", worksheet)
    group_list = select_by_col_name("group_id", worksheet)
    list_of_result = []
    for index in range(0, len(evaluation_list)):
        if evaluation_list[index] == eva_name and group_list[index] == group_id:
            list_of_result.append(index + 2)
    return list_of_result


# group_id and eva_name and owner are primary key in eva worksheet
def select_index_by_group_eva_owner(eva_name, group_id, owner, worksheet):
    evaluation_list = select_by_col_name("eva_name", worksheet)
    group_list = select_by_col_name("group_id", worksheet)
    owner_list = select_by_col_name("owner", worksheet)
    list_of_result = []
    for index in range(0, len(evaluation_list)):
        if evaluation_list[index] == eva_name and group_list[index] == group_id and owner_list[index] == owner:
            list_of_result.append(index + 2)
    return list_of_result


# find the unique index by eva_name, group id, owner, date
def select_index_by_group_eva_owner_date(eva_name, group_id, owner, date, worksheet):
    evaluation_list = select_by_col_name("eva_name", worksheet)
    group_list = select_by_col_name("group_id", worksheet)
    owner_list = select_by_col_name("owner", worksheet)
    date_list = select_by_col_name("date", worksheet)
    for index in range(0, len(evaluation_list)):
        if evaluation_list[index] == eva_name and group_list[index] == group_id and owner_list[
            index] == owner and date == date_list[index]:
            return index + 2
    return "nothing find"


def select_row_by_group_id(col_name, col_value, worksheet):
    # we suppose that the return rows are multiple
    rows_selected = []
    rows_in_worksheet = list(worksheet.iter_rows())
    # index in data_by_col == index in row
    data_by_col = select_by_col_name(col_name, worksheet)
    for index_data in range(0, len(data_by_col)):
        if data_by_col[index_data] == col_value:
            map_to_append = {}
            # tranfer the item to item.value
            for item in rows_in_worksheet[index_data + 1]:
                # create a map contains keys-values : like "group_id" = 1
                index_of_item = rows_in_worksheet[index_data + 1].index(item)
                tag_of_item = rows_in_worksheet[0][index_of_item].value
                map_to_append[tag_of_item] = item.value
            # record each row by its group_id
            # for example : 1:{group_id = 1, date= x/y/z, eva_name = eva3, .....}.
            rows_selected.append(map_to_append)
    # if nothing matched, return the empty array;
    return rows_selected


# select all students in the group
# implement in group worksheet
def select_students_by_group(group, worksheet):
    students = []
    groups = select_by_col_name("groupid", worksheet)
    index_of_group = groups.index(group) + 2
    row_by_index = list(worksheet.iter_rows())[index_of_group - 1]
    # the first element in row_by_index is groupid
    for student in row_by_index[1:]:
        students.append(student.value)
    return students


# generate an empty map with group name and evaluation name
def new_map_generator(group, eva_name, worksheet):
    map_to_return = {}
    tags_of_item = list(worksheet.iter_rows())[0]
    for tag in tags_of_item:
        if tag.value == 'group_id':
            map_to_return[tag.value] = group
        elif tag.value == 'eva_name':
            map_to_return[tag.value] = eva_name
        else:
            map_to_return[tag.value] = ''

    return map_to_return


def new_row_generator(group, students, eva_name, worksheet):
    row_to_return = []
    tags_of_item = list(worksheet.iter_rows())[0]
    for tag in tags_of_item:
        if tag.value == 'group_id':
            row_to_return.append(group)
        elif tag.value == 'eva_name':
            row_to_return.append(eva_name)
        elif tag.value == 'date':
            date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            row_to_return.append(date)
        elif tag.value == 'owner':
            row_to_return.append(current_user.username)
        elif tag.value == 'students':
            students_string = ",".join(students)
            row_to_return.append(students_string)
        elif tag.value == 'last_updates':
            row_to_return.append(current_user.username)
        else:
            row_to_return.append(" ")
    return row_to_return


def get_students_by_group(group_worksheet, students_worksheet):
    return_map = {}
    for group in list(group_worksheet.iter_rows())[1:]:
        return_map[group[0].value] = []
        for email in group[1:]:
            if email.value is not None:
                index = select_by_col_name('Email', students_worksheet).index(email.value)
                student_name = select_by_col_name('Student', students_worksheet)[index]
                student_couple = [email.value, student_name]
                return_map[group[0].value].append(student_couple)
    return return_map



# After login===============================================================================================================================

# this variable is expected by the wsgi server
application = app

if __name__ == '__main__':
    # db.create_all() # only run it the first time
    app.run(debug=True)

    # token: MFFt4RjpXNMh1c_T1AQj
