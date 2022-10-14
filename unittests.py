from app import *
import os
import openpyxl
import shutil
from flask import Flask
import random
import testProject
import unittest
import collections

class select_by_col_name_Test(unittest.TestCase):
    def test_check_studentxlsx_student_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_student_file_stored = "{}/student.xlsx".format(path_to_current_user_project)
        student_file_workbook = openpyxl.load_workbook(path_to_student_file_stored)
        student_file_worksheet = student_file_workbook['Sheet1']
        
        self.assertEqual(select_by_col_name("Student",student_file_worksheet),['Mccray, Maja', 'Guerrero, Fateh', 'Mcphee, Pearce', 'Michael, Olivia', 'Austin, Saarah', 'Velasquez, Rex', 'Coles, Rivka', 'Moody, Lyla', 'Morales, Conah', 'Cotton, Libbie', 'Ahmad, Rocky', 'Gardner, Shakira', 'Lopez, Cecilia', 'Aguilar, Sumayyah', 'Almond, Shae', 'Patton, Naima', "O'Moore, Kitty", 'Bowes, Aryan', 'Kirkpatrick, Arwen', 'Kendall, Jakob', 'Dowling, Safiyyah', 'Hull, Ciaran', 'Goddard, Rafael', 'Whittington, Aden', 'Mosley, Korban', 'Meyer, Sienna', 'Boyce, Bradlee', 'Patterson, Lilly', 'Wickens, Hebe', 'England, Kenzo', 'Parry, Sulayman', 'Baldwin, Ismail', 'Werner, Lillia', 'Emerson, Hiba', 'Casey, Gabriela', 'Johnston, Moshe', 'Thorpe, Shana', 'Lowry, Om', 'Lawson, Nikhil', 'Haney, Mara'])
        self.assertNotEqual(select_by_col_name("Student",student_file_worksheet),[])

    def test_check_studentxlsx_email_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_student_file_stored = "{}/student.xlsx".format(path_to_current_user_project)
        student_file_workbook = openpyxl.load_workbook(path_to_student_file_stored)
        student_file_worksheet = student_file_workbook['Sheet1']
        
        self.assertEqual(select_by_col_name("Email",student_file_worksheet),['rubricapp-c0@mailinator.com', 'rubricapp-c1@mailinator.com', 'rubricapp-c2@mailinator.com', 'rubricapp-c3@mailinator.com', 'rubricapp-c4@mailinator.com', 'rubricapp-c5@mailinator.com', 'rubricapp-c6@mailinator.com', 'rubricapp-c7@mailinator.com', 'rubricapp-c8@mailinator.com', 'rubricapp-c9@mailinator.com', 'rubricapp-c10@mailinator.com', 'rubricapp-c11@mailinator.com', 'rubricapp-c12@mailinator.com', 'rubricapp-c13@mailinator.com', 'rubricapp-c14@mailinator.com', 'rubricapp-c15@mailinator.com', 'rubricapp-c16@mailinator.com', 'rubricapp-c17@mailinator.com', 'rubricapp-c18@mailinator.com', 'rubricapp-c19@mailinator.com', 'rubricapp-c20@mailinator.com', 'rubricapp-c21@mailinator.com', 'rubricapp-c22@mailinator.com', 'rubricapp-c23@mailinator.com', 'rubricapp-c24@mailinator.com', 'rubricapp-c25@mailinator.com', 'rubricapp-c26@mailinator.com', 'rubricapp-c27@mailinator.com', 'rubricapp-c28@mailinator.com', 'rubricapp-c29@mailinator.com', 'rubricapp-c30@mailinator.com', 'rubricapp-c31@mailinator.com', 'rubricapp-c32@mailinator.com', 'rubricapp-c33@mailinator.com', 'rubricapp-c34@mailinator.com', 'rubricapp-c35@mailinator.com', 'rubricapp-c36@mailinator.com', 'rubricapp-c37@mailinator.com', 'rubricapp-c38@mailinator.com', 'rubricapp-c39@mailinator.com'])
        self.assertNotEqual(select_by_col_name("Email",student_file_worksheet),[])
        #print(select_by_col_name("Email",student_file_worksheet))
    
    def test_check_studentxlsx_group_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_student_file_stored = "{}/student.xlsx".format(path_to_current_user_project)
        student_file_workbook = openpyxl.load_workbook(path_to_student_file_stored)
        student_file_worksheet = student_file_workbook['Sheet1']
        
        self.assertEqual(select_by_col_name("group",student_file_worksheet),['H', 'H', 'H', 'H', 'He', 'He', 'He', 'He', 'Li', 'Li', 'Li', 'Li', 'Be', 'Be', 'Be', 'Be', 'B', 'B', 'B', 'B', 'C', 'C', 'C', 'C', 'N', 'N', 'N', 'N', 'O', 'O', 'O', 'O', 'F', 'F', 'F', 'F', 'Ne', 'Ne', 'Ne', 'Ne'])
        self.assertNotEqual(select_by_col_name("group",student_file_worksheet),[])
        #print(select_by_col_name("meta",student_file_worksheet))
    
    def test_check_studentxlsx_meta_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_student_file_stored = "{}/student.xlsx".format(path_to_current_user_project)
        student_file_workbook = openpyxl.load_workbook(path_to_student_file_stored)
        student_file_worksheet = student_file_workbook['Sheet1']
        
        self.assertEqual(select_by_col_name("meta",student_file_worksheet),['a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'c', 'c', 'c', 'c'])
        self.assertNotEqual(select_by_col_name("meta",student_file_worksheet),[])
    
    def test_check_groupxlsx_groupid_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_group_file_stored = "{}/group.xlsx".format(path_to_current_user_project)
        group_file_workbook = openpyxl.load_workbook(path_to_group_file_stored)
        group_file_worksheet1 = group_file_workbook['Sheet1']
        group_file_worksheet2 = group_file_workbook['Sheet2']
        
        self.assertNotEqual(select_by_col_name("groupid",group_file_worksheet1),[])
        self.assertNotEqual(select_by_col_name("groupid",group_file_worksheet2),[])

        self.assertEqual(collections.Counter(select_by_col_name("groupid",group_file_worksheet1)),collections.Counter(['Li', 'C', 'Be', 'N', 'B', 'F', 'H', 'Ne', 'He', 'O']))
        self.assertEqual(collections.Counter(select_by_col_name("groupid",group_file_worksheet1)),collections.Counter(['Li', 'C', 'Be', 'N', 'B', 'F', 'H', 'Ne', 'He', 'O']))
        #print(select_by_col_name("groupid",group_file_worksheet1))
    
    def test_check_groupxlsx_metaid_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_group_file_stored = "{}/group.xlsx".format(path_to_current_user_project)
        group_file_workbook = openpyxl.load_workbook(path_to_group_file_stored)
        group_file_worksheet2 = group_file_workbook['Sheet2']

        self.assertNotEqual(select_by_col_name("metaid",group_file_worksheet2),[])
        self.assertEqual(collections.Counter(select_by_col_name("metaid",group_file_worksheet2)),collections.Counter(['b', 'a', 'a', 'b', 'b', 'a', 'a', 'b', 'c', 'a']))
        #print(select_by_col_name("metaid",group_file_worksheet2))

    def test_check_groupxlsx_student1_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_group_file_stored = "{}/group.xlsx".format(path_to_current_user_project)
        group_file_workbook = openpyxl.load_workbook(path_to_group_file_stored)
        group_file_worksheet1 = group_file_workbook['Sheet1']

        self.assertNotEqual(select_by_col_name("student1",group_file_worksheet1),[])
        self.assertEqual(collections.Counter(select_by_col_name("student1",group_file_worksheet1)),collections.Counter(['rubricapp-c16@mailinator.com', 'rubricapp-c32@mailinator.com', 'rubricapp-c4@mailinator.com', 'rubricapp-c8@mailinator.com', 'rubricapp-c36@mailinator.com', 'rubricapp-c12@mailinator.com', 'rubricapp-c28@mailinator.com', 'rubricapp-c0@mailinator.com', 'rubricapp-c20@mailinator.com', 'rubricapp-c24@mailinator.com']))
        #print(select_by_col_name("student1",group_file_worksheet1))
    
    def test_check_groupxlsx_student2_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_group_file_stored = "{}/group.xlsx".format(path_to_current_user_project)
        group_file_workbook = openpyxl.load_workbook(path_to_group_file_stored)
        group_file_worksheet1 = group_file_workbook['Sheet1']

        self.assertNotEqual(select_by_col_name("student2",group_file_worksheet1),[])
        self.assertEqual(collections.Counter(select_by_col_name("student2",group_file_worksheet1)),collections.Counter(['rubricapp-c5@mailinator.com', 'rubricapp-c29@mailinator.com', 'rubricapp-c9@mailinator.com', 'rubricapp-c21@mailinator.com', 'rubricapp-c25@mailinator.com', 'rubricapp-c33@mailinator.com', 'rubricapp-c1@mailinator.com', 'rubricapp-c13@mailinator.com', 'rubricapp-c17@mailinator.com', 'rubricapp-c37@mailinator.com']))
        #print(select_by_col_name("student2",group_file_worksheet1))

    def test_check_groupxlsx_student3_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_group_file_stored = "{}/group.xlsx".format(path_to_current_user_project)
        group_file_workbook = openpyxl.load_workbook(path_to_group_file_stored)
        group_file_worksheet1 = group_file_workbook['Sheet1']

        self.assertNotEqual(select_by_col_name("student3",group_file_worksheet1),[])
        self.assertEqual(collections.Counter(select_by_col_name("student3",group_file_worksheet1)),collections.Counter(['rubricapp-c6@mailinator.com', 'rubricapp-c14@mailinator.com', 'rubricapp-c22@mailinator.com', 'rubricapp-c38@mailinator.com', 'rubricapp-c10@mailinator.com', 'rubricapp-c34@mailinator.com', 'rubricapp-c26@mailinator.com', 'rubricapp-c30@mailinator.com', 'rubricapp-c18@mailinator.com', 'rubricapp-c2@mailinator.com']))
        #print(select_by_col_name("student3",group_file_worksheet1))
    
    def test_check_groupxlsx_student4_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_group_file_stored = "{}/group.xlsx".format(path_to_current_user_project)
        group_file_workbook = openpyxl.load_workbook(path_to_group_file_stored)
        group_file_worksheet1 = group_file_workbook['Sheet1']

        self.assertNotEqual(select_by_col_name("student4",group_file_worksheet1),[])
        self.assertEqual(collections.Counter(select_by_col_name("student4",group_file_worksheet1)),collections.Counter(['rubricapp-c35@mailinator.com', 'rubricapp-c39@mailinator.com', 'rubricapp-c11@mailinator.com', 'rubricapp-c23@mailinator.com', 'rubricapp-c27@mailinator.com', 'rubricapp-c31@mailinator.com', 'rubricapp-c15@mailinator.com', 'rubricapp-c3@mailinator.com', 'rubricapp-c7@mailinator.com', 'rubricapp-c19@mailinator.com']))
        #print(select_by_col_name("student4",group_file_worksheet1))
    
    def test_check_evaluationxlsx_groupid_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_evaluation_file_stored = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_file_workbook = openpyxl.load_workbook(path_to_evaluation_file_stored)
        evaluation_file_worksheet4 = evaluation_file_workbook['eva']

        self.assertNotEqual(select_by_col_name("group_id",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("group_id",evaluation_file_worksheet4),[])
        #print(select_by_col_name("group_id",evaluation_file_worksheet4))

    def test_check_evaluationxlsx_evaname_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_evaluation_file_stored = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_file_workbook = openpyxl.load_workbook(path_to_evaluation_file_stored)
        evaluation_file_worksheet4 = evaluation_file_workbook['eva']

        self.assertNotEqual(select_by_col_name("eva_name",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("eva_name",evaluation_file_worksheet4),[])
        #print(select_by_col_name("group_id",evaluation_file_worksheet4))
    
    def test_check_evaluationxlsx_owner_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_evaluation_file_stored = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_file_workbook = openpyxl.load_workbook(path_to_evaluation_file_stored)
        evaluation_file_worksheet4 = evaluation_file_workbook['eva']

        self.assertNotEqual(select_by_col_name("owner",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("owner",evaluation_file_worksheet4),[])
        #print(select_by_col_name("group_id",evaluation_file_worksheet4))

    def test_check_evaluationxlsx_date_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_evaluation_file_stored = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_file_workbook = openpyxl.load_workbook(path_to_evaluation_file_stored)
        evaluation_file_worksheet4 = evaluation_file_workbook['eva']

        self.assertNotEqual(select_by_col_name("date",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("date",evaluation_file_worksheet4),[])
        #print(select_by_col_name("group_id",evaluation_file_worksheet4))

    def test_check_evaluationxlsx_students_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_evaluation_file_stored = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_file_workbook = openpyxl.load_workbook(path_to_evaluation_file_stored)
        evaluation_file_worksheet4 = evaluation_file_workbook['eva']

        self.assertNotEqual(select_by_col_name("students",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("students",evaluation_file_worksheet4),[])
        #print(select_by_col_name("group_id",evaluation_file_worksheet4))

    def test_check_evaluationxlsx_F_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_evaluation_file_stored = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_file_workbook = openpyxl.load_workbook(path_to_evaluation_file_stored)
        evaluation_file_worksheet4 = evaluation_file_workbook['eva']

        self.assertNotEqual(select_by_col_name("Evaluating|Determined the significance or relevance of information/data needed for the task.",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("Evaluating|Determined the significance or relevance of information/data needed for the task.",evaluation_file_worksheet4),[])
        #print(select_by_col_name("group_id",evaluation_file_worksheet4))

    def test_check_evaluationxlsx_F_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_evaluation_file_stored = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_file_workbook = openpyxl.load_workbook(path_to_evaluation_file_stored)
        evaluation_file_worksheet4 = evaluation_file_workbook['eva']

        self.assertNotEqual(select_by_col_name("Evaluating|Determined the significance or relevance of information/data needed for the task.",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("Evaluating|Determined the significance or relevance of information/data needed for the task.",evaluation_file_worksheet4),[])
        #print(select_by_col_name("group_id",evaluation_file_worksheet4))

    def test_check_evaluationxlsx_G_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_evaluation_file_stored = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_file_workbook = openpyxl.load_workbook(path_to_evaluation_file_stored)
        evaluation_file_worksheet4 = evaluation_file_workbook['eva']

        self.assertNotEqual(select_by_col_name("Evaluating|Observed Characteristics",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("Evaluating|Observed Characteristics",evaluation_file_worksheet4),[])
        #print(select_by_col_name("group_id",evaluation_file_worksheet4))

    def test_check_evaluationxlsx_H_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_evaluation_file_stored = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_file_workbook = openpyxl.load_workbook(path_to_evaluation_file_stored)
        evaluation_file_worksheet4 = evaluation_file_workbook['eva']

        self.assertNotEqual(select_by_col_name("Interpreting|Provided meaning to data, made inferences and predictions from data, or extracted patterns from data.",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("Interpreting|Provided meaning to data, made inferences and predictions from data, or extracted patterns from data.",evaluation_file_worksheet4),[])
        #print(select_by_col_name("group_id",evaluation_file_worksheet4))

    def test_check_evaluationxlsx_I_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_evaluation_file_stored = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_file_workbook = openpyxl.load_workbook(path_to_evaluation_file_stored)
        evaluation_file_worksheet4 = evaluation_file_workbook['eva']

        self.assertNotEqual(select_by_col_name("Interpreting|Observed Characteristics",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("Interpreting|Observed Characteristics",evaluation_file_worksheet4),[])
        #print(select_by_col_name("group_id",evaluation_file_worksheet4))

    def test_check_evaluationxlsx_J_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_evaluation_file_stored = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_file_workbook = openpyxl.load_workbook(path_to_evaluation_file_stored)
        evaluation_file_worksheet4 = evaluation_file_workbook['eva']

        self.assertNotEqual(select_by_col_name("Manipulating or Transforming (Extent)|Observed Characteristics.",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("Manipulating or Transforming (Extent)|Observed Characteristics",evaluation_file_worksheet4),[])
        #print(select_by_col_name("group_id",evaluation_file_worksheet4))

    def test_check_evaluationxlsx_K_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_evaluation_file_stored = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_file_workbook = openpyxl.load_workbook(path_to_evaluation_file_stored)
        evaluation_file_worksheet4 = evaluation_file_workbook['eva']

        self.assertNotEqual(select_by_col_name("Manipulating or Transforming (Extent)|Observed Characteristics",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("Manipulating or Transforming (Extent)|Observed Characteristics",evaluation_file_worksheet4),[])
        #print(select_by_col_name("group_id",evaluation_file_worksheet4))

    def test_check_evaluationxlsx_L_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_evaluation_file_stored = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_file_workbook = openpyxl.load_workbook(path_to_evaluation_file_stored)
        evaluation_file_worksheet4 = evaluation_file_workbook['eva']

        self.assertNotEqual(select_by_col_name("Manipulating or Transforming (Accuracy)|Converted information/data from one form to another.",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("Manipulating or Transforming (Accuracy)|Converted information/data from one form to another.",evaluation_file_worksheet4),[])
        #print(select_by_col_name("group_id",evaluation_file_worksheet4))
    
    def test_check_evaluationxlsx_M_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_evaluation_file_stored = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_file_workbook = openpyxl.load_workbook(path_to_evaluation_file_stored)
        evaluation_file_worksheet4 = evaluation_file_workbook['eva']

        self.assertNotEqual(select_by_col_name("Manipulating or Transforming (Accuracy)|Observed Characteristics",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("Manipulating or Transforming (Accuracy)|Observed Characteristics",evaluation_file_worksheet4),[])
        #print(select_by_col_name("group_id",evaluation_file_worksheet4))

    def test_check_evaluationxlsx_O_col(self):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, email, project_name)
        path_to_evaluation_file_stored = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_file_workbook = openpyxl.load_workbook(path_to_evaluation_file_stored)
        evaluation_file_worksheet4 = evaluation_file_workbook['eva']

        self.assertNotEqual(select_by_col_name("last_updates",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("last_updates",evaluation_file_worksheet4),[])
        #print(select_by_col_name("group_id",evaluation_file_worksheet4))



            


if __name__ == '__main__':
    
    flaskapp = create_app()
    email = "testing@gmail.com"
    project_name = "Project " + str(random.randint(0,1000))+ str(random.randint(0,1000))
    description_name = "Description" + str(random.randint(0,1000))
    with flaskapp.app_context():
        testProject.create_test_project(email,project_name,description_name)
        unittest.main(argv=['first-arg-is-ignored'], exit=False)

    
        #s = select_by_col_name_Test(email,project_name)
        #s.print_col()
        