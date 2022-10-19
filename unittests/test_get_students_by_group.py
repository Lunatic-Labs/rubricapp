import sys
sys.path.append('..')
import unittest
from app import create_app, get_students_by_group
from openpyxl import load_workbook
import random
from createTestProject import *
from enum import Enum
import email
from re import X

class StudEmails(Enum):
    H =  ["rubricapp-c0@mailinator.com",  "rubricapp-c1@mailinator.com",  "rubricapp-c2@mailinator.com",  "rubricapp-c3@mailinator.com" ]
    He = ["rubricapp-c4@mailinator.com",  "rubricapp-c5@mailinator.com",  "rubricapp-c6@mailinator.com",  "rubricapp-c7@mailinator.com" ]
    Li = ["rubricapp-c8@mailinator.com",  "rubricapp-c9@mailinator.com",  "rubricapp-c10@mailinator.com", "rubricapp-c11@mailinator.com"]
    Be = ["rubricapp-c12@mailinator.com", "rubricapp-c13@mailinator.com", "rubricapp-c14@mailinator.com", "rubricapp-c15@mailinator.com"]
    B =  ["rubricapp-c16@mailinator.com", "rubricapp-c17@mailinator.com", "rubricapp-c18@mailinator.com", "rubricapp-c19@mailinator.com"]
    C =  ["rubricapp-c20@mailinator.com", "rubricapp-c21@mailinator.com", "rubricapp-c22@mailinator.com", "rubricapp-c23@mailinator.com"]
    N =  ["rubricapp-c24@mailinator.com", "rubricapp-c25@mailinator.com", "rubricapp-c26@mailinator.com", "rubricapp-c27@mailinator.com"]
    O =  ["rubricapp-c28@mailinator.com", "rubricapp-c29@mailinator.com", "rubricapp-c30@mailinator.com", "rubricapp-c31@mailinator.com"]
    F =  ["rubricapp-c32@mailinator.com", "rubricapp-c33@mailinator.com", "rubricapp-c34@mailinator.com", "rubricapp-c35@mailinator.com"]
    Ne = ["rubricapp-c36@mailinator.com", "rubricapp-c37@mailinator.com", "rubricapp-c38@mailinator.com", "rubricapp-c39@mailinator.com"]

class StudNames(Enum):
    H =  ["Mccray, Maja",      "Guerrero, Fateh",   "Mcphee, Pearce",     "Michael, Olivia"  ]
    He = ["Austin, Saarah",    "Velasquez, Rex",    "Coles, Rivka",       "Moody, Lyla"      ]
    Li = ["Morales, Conah",    "Cotton, Libbie",    "Ahmad, Rocky",       "Gardner, Shakira" ]
    Be = ["Lopez, Cecilia",    "Aguilar, Sumayyah", "Almond, Shae",       "Patton, Naima"    ]
    B =  ["O'Moore, Kitty",    "Bowes, Aryan",      "Kirkpatrick, Arwen", "Kendall, Jakob"   ]
    C =  ["Dowling, Safiyyah", "Hull, Ciaran",      "Goddard, Rafael",    "Whittington, Aden"]
    N =  ["Mosley, Korban",    "Meyer, Sienna",     "Boyce, Bradlee",     "Patterson, Lilly" ]
    O =  ["Wickens, Hebe",     "England, Kenzo",    "Parry, Sulayman",    "Baldwin, Ismail"  ]
    F =  ["Werner, Lillia",    "Emerson, Hiba",     "Casey, Gabriela",    "Johnston, Moshe"  ]
    Ne = ["Thorpe, Shana",     "Lowry, Om",         "Lawson, Nikhil",     "Haney, Mara"      ]


def checkEmail(self, index):
    self.assertEqual(len(self.students[str(self.group_col[index])]), 4)
    answer = [] 
    for emails in StudEmails:
        if (self.group_col[index]) == emails.name:
            answer = emails.value
    count = 0
    for student_couple in self.students[str(self.group_col[index])]:
        self.assertEqual(student_couple[0], answer[count])
        count += 1

def checkName(self, index):
    self.assertEqual(len(self.students[str(self.group_col[index])]), 4)
    answer = [] 
    for names in StudNames:
        if (self.group_col[index]) == names.name:
            answer = names.value
    count = 0
    for student_couple in self.students[str(self.group_col[index])]:
        self.assertEqual(student_couple[1], answer[count])
        count += 1


class TestGetStudentsByGroup(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        
        base_directory = os.getcwd()+"/users"
        if not os.path.exists(base_directory+"/test@gmail.com"):
            os.mkdir(base_directory+"/test@gmail.com")

        flask_app = create_app()
        with flask_app.app_context():
            cls.projectName = "Test pName" + str(random.getrandbits(12)) + str(random.getrandbits(12)) + str(random.getrandbits(12))
            create_test_project("test@gmail.com", cls.projectName)

        path_to_load_project = "{}/{}/{}".format(base_directory, "test@gmail.com", cls.projectName)

        path_to_evaluation_file = "{}/evaluation.xlsx".format(path_to_load_project)
        eva_workbook = load_workbook(path_to_evaluation_file)
        cls.group_worksheet = eva_workbook['group']
        cls.student_worksheet = eva_workbook['students']
        cls.students = get_students_by_group(cls.group_worksheet, cls.student_worksheet)

        cls.group_col = []
        for col_item in list(cls.group_worksheet.iter_cols())[0]:
            if col_item.value != "groupid":
                cls.group_col.append(col_item.value)  

    @classmethod
    def tearDownClass(cls):
        delete_project("test@gmail.com", cls.projectName)


    def test_gsbg_index_0_email(self):
        checkEmail(self, 0)

    def test_gsbg_index_1_email(self):
        checkEmail(self, 1)

    def test_gsbg_index_2_email(self):
        checkEmail(self, 2)

    def test_gsbg_index_3_email(self):
        checkEmail(self, 3)

    def test_gsbg_index_4_email(self):
        checkEmail(self, 4)

    def test_gsbg_index_5_email(self):
        checkEmail(self, 5)

    def test_gsbg_index_6_email(self):
        checkEmail(self, 6)

    def test_gsbg_index_7_email(self):
        checkEmail(self, 7)

    def test_gsbg_index_8_email(self):
        checkEmail(self, 8)

    def test_gsbg_index_9_email(self):
        checkEmail(self, 9)

# --- Checking Names from this point onward  ---

    def test_gsbg_index_0_name(self):
        checkName(self, 0)

    def test_gsbg_index_1_name(self):
        checkName(self, 1)

    def test_gsbg_index_2_name(self):
        checkName(self, 2)

    def test_gsbg_index_3_name(self):
        checkName(self, 3)

    def test_gsbg_index_4_name(self):
        checkName(self, 4)

    def test_gsbg_index_5_name(self):
        checkName(self, 5)

    def test_gsbg_index_6_name(self):
        checkName(self, 6)

    def test_gsbg_index_7_name(self):
        checkName(self, 7)

    def test_gsbg_index_8_name(self):
        checkName(self, 8)

    def test_gsbg_index_9_name(self):
        checkName(self, 9)
