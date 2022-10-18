import sys
sys.path.append('..')
from app import *
import os
import openpyxl
import shutil
from flask import Flask
import random
from createTestProject import *
import unittest
import collections

def makeEvalNames(evalnameList):
    for i in range(40):
        evalnameList.insert(i,"EvalName" + str(i+2))

evalnameList = []
makeEvalNames(evalnameList)

class select_map_by_index_Test(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        
        base_directory = os.getcwd()+"/users"
        if not os.path.exists(base_directory+"/test@gmail.com"):
            os.mkdir(base_directory+"/test@gmail.com")

        flask_app = create_app()
        with flask_app.app_context():
            cls.projectName = "Test pName" + str(random.getrandbits(12)) + str(random.getrandbits(12)) + str(random.getrandbits(12))
            create_test_project("test@gmail.com", cls.projectName)
            createEvaluation("test@gmail.com", cls.projectName,evalnameList)

    @classmethod
    def tearDownClass(cls):    
        delete_project("test@gmail.com", cls.projectName)
    
    @classmethod
    def studentxlsx_sheet1_setup(cls):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, "test@gmail.com", cls.projectName)
        path_to_student_file_stored = "{}/student.xlsx".format(path_to_current_user_project)
        student_file_workbook = openpyxl.load_workbook(path_to_student_file_stored)
        return student_file_workbook['Sheet1']

    @classmethod
    def groupxlsx_sheet1_setup(cls):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, "test@gmail.com", cls.projectName)
        path_to_group_file_stored = "{}/group.xlsx".format(path_to_current_user_project)
        group_file_workbook = openpyxl.load_workbook(path_to_group_file_stored)
        return group_file_workbook['Sheet1']

    @classmethod
    def groupxlsx_sheet2_setup(cls):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, "test@gmail.com", cls.projectName)
        path_to_group_file_stored = "{}/group.xlsx".format(path_to_current_user_project)
        group_file_workbook = openpyxl.load_workbook(path_to_group_file_stored)
        return group_file_workbook['Sheet2']

    @classmethod
    def evaluationxlsx_meta_setup(cls):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, "test@gmail.com", cls.projectName)
        path_to_evaluation_file_stored = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_file_workbook = openpyxl.load_workbook(path_to_evaluation_file_stored)
        return evaluation_file_workbook['meta']

    @classmethod
    def evaluationxlsx_students_setup(cls):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, "test@gmail.com", cls.projectName)
        path_to_evaluation_file_stored = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_file_workbook = openpyxl.load_workbook(path_to_evaluation_file_stored)
        return evaluation_file_workbook['students']
    
    @classmethod
    def evaluationxlsx_group_setup(cls):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, "test@gmail.com", cls.projectName)
        path_to_evaluation_file_stored = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_file_workbook = openpyxl.load_workbook(path_to_evaluation_file_stored)
        return evaluation_file_workbook['group']

    def test_studentxlsx_map_row2(self):
        student_file_worksheet = select_map_by_index_Test.studentxlsx_sheet1_setup()
        self.assertEqual(select_map_by_index(3,student_file_worksheet),{'Student': 'Guerrero, Fateh', 'Email': 'rubricapp-c1@mailinator.com', 'group': 'H', 'meta': 'a'})

    def test_studentxlsx_map_row1(self):
        student_file_worksheet = select_map_by_index_Test.studentxlsx_sheet1_setup()
        self.assertEqual(select_map_by_index(2,student_file_worksheet),{'Student': 'Mccray, Maja', 'Email': 'rubricapp-c0@mailinator.com', 'group': 'H', 'meta': 'a'})

    def test_studentxlsx_map_row41(self):
        group_file_worksheet = select_map_by_index_Test.studentxlsx_sheet1_setup()
        self.assertEqual(select_map_by_index(41,group_file_worksheet),{'Student': 'Haney, Mara', 'Email': 'rubricapp-c39@mailinator.com', 'group': 'Ne', 'meta': 'c'})

    def test_groupxlsx_map_row(self):
        group_file_worksheet = select_map_by_index_Test.groupxlsx_sheet1_setup()
        for i in range(2,12):
            v = group_file_worksheet.cell(i,1)
            
            if v.value == "O":
                self.assertEqual(select_map_by_index(i,group_file_worksheet),{'groupid': 'O', 'student1': 'rubricapp-c28@mailinator.com', 'student2': 'rubricapp-c29@mailinator.com', 'student3': 'rubricapp-c30@mailinator.com', 'student4': 'rubricapp-c31@mailinator.com'}) 
            elif v.value == "C":
                self.assertEqual(select_map_by_index(i,group_file_worksheet),{'groupid': 'C', 'student1': 'rubricapp-c20@mailinator.com', 'student2': 'rubricapp-c21@mailinator.com', 'student3': 'rubricapp-c22@mailinator.com', 'student4': 'rubricapp-c23@mailinator.com'})
            elif v.value == "Li":
                self.assertEqual(select_map_by_index(i,group_file_worksheet),{'groupid': 'Li', 'student1': 'rubricapp-c8@mailinator.com', 'student2': 'rubricapp-c9@mailinator.com', 'student3': 'rubricapp-c10@mailinator.com', 'student4': 'rubricapp-c11@mailinator.com'})
            elif v.value == "F":
                self.assertEqual(select_map_by_index(i,group_file_worksheet),{'groupid': 'F', 'student1': 'rubricapp-c32@mailinator.com', 'student2': 'rubricapp-c33@mailinator.com', 'student3': 'rubricapp-c34@mailinator.com', 'student4': 'rubricapp-c35@mailinator.com'})
            elif v.value == "B":
                self.assertEqual(select_map_by_index(i,group_file_worksheet),{'groupid': 'B', 'student1': 'rubricapp-c16@mailinator.com', 'student2': 'rubricapp-c17@mailinator.com', 'student3': 'rubricapp-c18@mailinator.com', 'student4': 'rubricapp-c19@mailinator.com'})
            elif v.value == "Be":
                self.assertEqual(select_map_by_index(i,group_file_worksheet),{'groupid': 'Be', 'student1': 'rubricapp-c12@mailinator.com', 'student2': 'rubricapp-c13@mailinator.com', 'student3': 'rubricapp-c14@mailinator.com', 'student4': 'rubricapp-c15@mailinator.com'})
            elif v.value == "He":
                self.assertEqual(select_map_by_index(i,group_file_worksheet),{'groupid': 'He', 'student1': 'rubricapp-c4@mailinator.com', 'student2': 'rubricapp-c5@mailinator.com', 'student3': 'rubricapp-c6@mailinator.com', 'student4': 'rubricapp-c7@mailinator.com'})
            elif v.value == "Ne":
                self.assertEqual(select_map_by_index(i,group_file_worksheet),{'groupid': 'Ne', 'student1': 'rubricapp-c36@mailinator.com', 'student2': 'rubricapp-c37@mailinator.com', 'student3': 'rubricapp-c38@mailinator.com', 'student4': 'rubricapp-c39@mailinator.com'})
            elif v.value == "H":
                self.assertEqual(select_map_by_index(i,group_file_worksheet),{'groupid': 'H', 'student1': 'rubricapp-c0@mailinator.com', 'student2': 'rubricapp-c1@mailinator.com', 'student3': 'rubricapp-c2@mailinator.com', 'student4': 'rubricapp-c3@mailinator.com'})
            elif v.value == "N":
                self.assertEqual(select_map_by_index(i,group_file_worksheet),{'groupid': 'N', 'student1': 'rubricapp-c24@mailinator.com', 'student2': 'rubricapp-c25@mailinator.com', 'student3': 'rubricapp-c26@mailinator.com', 'student4': 'rubricapp-c27@mailinator.com'})  

    def test_groupxlsx_map_row2(self):
        group_file_worksheet = select_map_by_index_Test.groupxlsx_sheet2_setup()

        for i in range(2,12):
            v = group_file_worksheet.cell(i,1)
            
            if v.value == "O":
                self.assertEqual(select_map_by_index(i,group_file_worksheet),{'groupid': 'O', 'metaid': 'b'}) 
            elif v.value == "C":
                self.assertEqual(select_map_by_index(i,group_file_worksheet),{'groupid': 'C', 'metaid': 'b'})
            elif v.value == "Li":
                self.assertEqual(select_map_by_index(i,group_file_worksheet),{'groupid': 'Li', 'metaid': 'a'})
            elif v.value == "F":
                self.assertEqual(select_map_by_index(i,group_file_worksheet),{'groupid': 'F', 'metaid': 'b'})
            elif v.value == "B":
                self.assertEqual(select_map_by_index(i,group_file_worksheet),{'groupid': 'B', 'metaid': 'a'})
            elif v.value == "Be":
                self.assertEqual(select_map_by_index(i,group_file_worksheet),{'groupid': 'Be', 'metaid': 'a'})
            elif v.value == "He":
                self.assertEqual(select_map_by_index(i,group_file_worksheet),{'groupid': 'He', 'metaid': 'a'})
            elif v.value == "Ne":
                self.assertEqual(select_map_by_index(i,group_file_worksheet),{'groupid': 'Ne', 'metaid': 'c'})
            elif v.value == "H":
                self.assertEqual(select_map_by_index(i,group_file_worksheet),{'groupid': 'H', 'metaid': 'a'})
            elif v.value == "N":
                self.assertEqual(select_map_by_index(i,group_file_worksheet),{'groupid': 'N', 'metaid': 'b'}) 

    def test_evalxlsx_map_row(self):
        evaluation_file_worksheet = select_map_by_index_Test.evaluationxlsx_group_setup()

        for i in range(2,12):
            v = evaluation_file_worksheet.cell(i,1)
            
            if v.value == "O":
                self.assertEqual(select_map_by_index(i,evaluation_file_worksheet),{'groupid': 'O', 'student1': 'rubricapp-c28@mailinator.com', 'student2': 'rubricapp-c29@mailinator.com', 'student3': 'rubricapp-c30@mailinator.com', 'student4': 'rubricapp-c31@mailinator.com'}) 
            elif v.value == "C":
                self.assertEqual(select_map_by_index(i,evaluation_file_worksheet),{'groupid': 'C', 'student1': 'rubricapp-c20@mailinator.com', 'student2': 'rubricapp-c21@mailinator.com', 'student3': 'rubricapp-c22@mailinator.com', 'student4': 'rubricapp-c23@mailinator.com'})
            elif v.value == "Li":
                self.assertEqual(select_map_by_index(i,evaluation_file_worksheet),{'groupid': 'Li', 'student1': 'rubricapp-c8@mailinator.com', 'student2': 'rubricapp-c9@mailinator.com', 'student3': 'rubricapp-c10@mailinator.com', 'student4': 'rubricapp-c11@mailinator.com'})
            elif v.value == "F":
                self.assertEqual(select_map_by_index(i,evaluation_file_worksheet),{'groupid': 'F', 'student1': 'rubricapp-c32@mailinator.com', 'student2': 'rubricapp-c33@mailinator.com', 'student3': 'rubricapp-c34@mailinator.com', 'student4': 'rubricapp-c35@mailinator.com'})
            elif v.value == "B":
                self.assertEqual(select_map_by_index(i,evaluation_file_worksheet),{'groupid': 'B', 'student1': 'rubricapp-c16@mailinator.com', 'student2': 'rubricapp-c17@mailinator.com', 'student3': 'rubricapp-c18@mailinator.com', 'student4': 'rubricapp-c19@mailinator.com'})
            elif v.value == "Be":
                self.assertEqual(select_map_by_index(i,evaluation_file_worksheet),{'groupid': 'Be', 'student1': 'rubricapp-c12@mailinator.com', 'student2': 'rubricapp-c13@mailinator.com', 'student3': 'rubricapp-c14@mailinator.com', 'student4': 'rubricapp-c15@mailinator.com'})
            elif v.value == "He":
                self.assertEqual(select_map_by_index(i,evaluation_file_worksheet),{'groupid': 'He', 'student1': 'rubricapp-c4@mailinator.com', 'student2': 'rubricapp-c5@mailinator.com', 'student3': 'rubricapp-c6@mailinator.com', 'student4': 'rubricapp-c7@mailinator.com'})
            elif v.value == "Ne":
                self.assertEqual(select_map_by_index(i,evaluation_file_worksheet),{'groupid': 'Ne', 'student1': 'rubricapp-c36@mailinator.com', 'student2': 'rubricapp-c37@mailinator.com', 'student3': 'rubricapp-c38@mailinator.com', 'student4': 'rubricapp-c39@mailinator.com'})
            elif v.value == "H":
                self.assertEqual(select_map_by_index(i,evaluation_file_worksheet),{'groupid': 'H', 'student1': 'rubricapp-c0@mailinator.com', 'student2': 'rubricapp-c1@mailinator.com', 'student3': 'rubricapp-c2@mailinator.com', 'student4': 'rubricapp-c3@mailinator.com'})
            elif v.value == "N":
                self.assertEqual(select_map_by_index(i,evaluation_file_worksheet),{'groupid': 'N', 'student1': 'rubricapp-c24@mailinator.com', 'student2': 'rubricapp-c25@mailinator.com', 'student3': 'rubricapp-c26@mailinator.com', 'student4': 'rubricapp-c27@mailinator.com'}) 

    def test_evalxlsx_student_map_row2(self):
        student_file_worksheet = select_map_by_index_Test.evaluationxlsx_students_setup()
        
        self.assertEqual(select_map_by_index(3,student_file_worksheet),{'Student': 'Guerrero, Fateh', 'Email': 'rubricapp-c1@mailinator.com', 'group': 'H', 'meta': 'a'})

    def test_evalxlsx_student_map_row1(self):
        student_file_worksheet = select_map_by_index_Test.evaluationxlsx_students_setup()

        self.assertEqual(select_map_by_index(2,student_file_worksheet),{'Student': 'Mccray, Maja', 'Email': 'rubricapp-c0@mailinator.com', 'group': 'H', 'meta': 'a'})

    def test_evalxlsx_student_map_row41(self):
        group_file_worksheet = select_map_by_index_Test.evaluationxlsx_students_setup()

        self.assertEqual(select_map_by_index(41,group_file_worksheet),{'Student': 'Haney, Mara', 'Email': 'rubricapp-c39@mailinator.com', 'group': 'Ne', 'meta': 'c'})
   
    def test_evalxlsx_meta_map_row(self):
        eval_file_worksheet = select_map_by_index_Test.evaluationxlsx_meta_setup()

        for i in range(2,12):
            v = eval_file_worksheet.cell(i,1)
            
            if v.value == "O":
                self.assertEqual(select_map_by_index(i,eval_file_worksheet),{'groupid': 'O', 'metaid': 'b'}) 
            elif v.value == "C":
                self.assertEqual(select_map_by_index(i,eval_file_worksheet),{'groupid': 'C', 'metaid': 'b'})
            elif v.value == "Li":
                self.assertEqual(select_map_by_index(i,eval_file_worksheet),{'groupid': 'Li', 'metaid': 'a'})
            elif v.value == "F":
                self.assertEqual(select_map_by_index(i,eval_file_worksheet),{'groupid': 'F', 'metaid': 'b'})
            elif v.value == "B":
                self.assertEqual(select_map_by_index(i,eval_file_worksheet),{'groupid': 'B', 'metaid': 'a'})
            elif v.value == "Be":
                self.assertEqual(select_map_by_index(i,eval_file_worksheet),{'groupid': 'Be', 'metaid': 'a'})
            elif v.value == "He":
                self.assertEqual(select_map_by_index(i,eval_file_worksheet),{'groupid': 'He', 'metaid': 'a'})
            elif v.value == "Ne":
                self.assertEqual(select_map_by_index(i,eval_file_worksheet),{'groupid': 'Ne', 'metaid': 'c'})
            elif v.value == "H":
                self.assertEqual(select_map_by_index(i,eval_file_worksheet),{'groupid': 'H', 'metaid': 'a'})
            elif v.value == "N":
                self.assertEqual(select_map_by_index(i,eval_file_worksheet),{'groupid': 'N', 'metaid': 'b'})         
