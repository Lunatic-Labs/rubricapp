import sys
import flask_login
sys.path.append('..')
# from app import select_by_col_name, select_map_by_index, json, FileLock, current_user, new_row_generator, get_students_by_group,datetime
from filelock import FileLock
# json, FileLock, current_user
import json
from functions import select_by_col_name, select_map_by_index, new_row_generator, get_students_by_group, datetime
from flask_login import utils
import os
import openpyxl
import shutil

os.chdir("..")
base_directory = os.getcwd()
home_directory = base_directory
base_directory = base_directory + "/users"
project_name = []
date = [0]

def create_test_project(email, projectName):


    path_to_sample_roster = "{}/core/sample_file/rosters/sample_roster.xlsx".format(home_directory)
    path_to_sample_json = "{}/core/sample_file/rubrics/information_processing/information_processing.json".format(home_directory)
        
    # create project folder
    path_to_current_user_project = "{}/{}/{}".format(base_directory, email, projectName)
    os.mkdir(path_to_current_user_project)

    path_to_student_file_stored = "{}/student.xlsx".format(path_to_current_user_project)
    shutil.copy(path_to_sample_roster, path_to_student_file_stored)
    path_to_json_file_stored = "{}/TW.json".format(path_to_current_user_project)
    shutil.copy(path_to_sample_json, path_to_json_file_stored)
    # creating evaluation doc based on grading criteria json file

    # copy student sheet to evaluation doc
    student_file_workbook = openpyxl.load_workbook(path_to_student_file_stored)
    student_file_worksheet = student_file_workbook['Sheet1']

    # create group file depending on student file
    list_of_group = select_by_col_name('group', student_file_worksheet)
    set_of_group = set(list_of_group)

    # Fixing a bug where a None element was found. Is this safe?
    set_of_group.discard(None)

    # create a group workbook
    path_to_group_file = "{}/group.xlsx".format(path_to_current_user_project)
    group_workbook = openpyxl.Workbook()
    group_file_worksheet = group_workbook.create_sheet('Sheet1')
    meta_file_worksheet = group_workbook.create_sheet('Sheet2')
    # all student information map
    student_map_list = []
    for student_index in range(2, len(list(student_file_worksheet.iter_rows())) + 1):
        student_map_list.append(select_map_by_index(
            student_index, student_file_worksheet))
    # insert group columns
    group_file_worksheet.cell(1, 1).value = 'groupid'
    meta_file_worksheet.cell(1, 1).value = 'groupid'
    meta_file_worksheet.cell(1, 2).value = 'metaid'
    start_index = 2
    max_num_students_pergroup = 0

    list_of_group = list(dict.fromkeys(list_of_group))
    for group in list_of_group:##change
        group_file_worksheet.cell(start_index, 1).value = group
        student_emails = [x['Email']
                        for x in student_map_list if x['group'] == group]
        if len(student_emails) > max_num_students_pergroup:
            max_num_students_pergroup = len(student_emails)
        meta_file_worksheet.cell(start_index, 1).value = group
        meta_group = [x['meta']
                    for x in student_map_list if x['group'] == group][0]
        meta_file_worksheet.cell(start_index, 2).value = meta_group
        for insert_index in range(2, len(student_emails) + 2):
            group_file_worksheet.cell(
                start_index, insert_index).value = student_emails[insert_index - 2]
        start_index += 1
    
    for index in range(1, max_num_students_pergroup+1):
        group_file_worksheet.cell(1, 1+index).value = ("student" + str(index) )
    group_workbook.save(path_to_group_file)

    path_to_evaluation = "{}/evaluation.xlsx".format(
        path_to_current_user_project)
    evaluation_workbook = openpyxl.Workbook()
    evaluation_group = evaluation_workbook.create_sheet('group')
    evaluation_meta = evaluation_workbook.create_sheet('meta')
    evaluation_student = evaluation_workbook.create_sheet('students')
    copy_all_worksheet(evaluation_group, group_file_worksheet)
    copy_all_worksheet(evaluation_meta, meta_file_worksheet)
    copy_all_worksheet(evaluation_student, student_file_worksheet)
    # create EVA depending on the json file
    evaluation_eva = evaluation_workbook.create_sheet('eva')
    # open json file and load json
    myLock = FileLock(path_to_json_file_stored+'.lock', timeout=5)
    with myLock:
        with open(path_to_json_file_stored, 'r')as f:
            json_data = json.loads(f.read(), strict=False)
    # The group id, eva_name, date are defults
    tags_to_append = ['group_id', 'eva_name',
                    'owner', 'date', 'students']
    for category in json_data['category']:
        category_name = (category['name'])
        for section in category['section']:
            # instructors don't care about the text value, the text values will only be send to students.
            if section['type'] != 'text':
                value_to_append = "{}|{}".format(
                    category_name, section['name'])
                tags_to_append.append(value_to_append)
    tags_to_append.append("comment")
    tags_to_append.append("last_updates")
    evaluation_eva.append(tags_to_append)

    evaluation_workbook.save(path_to_evaluation)
    flask_login.current_user = 'Guest'

def createEvaluation(email,projectName,evaluation_name):
    path_to_load_project = "{}/{}/{}".format(base_directory, email, projectName)
    path_to_evaluation_file = "{}/evaluation.xlsx".format(path_to_load_project)
    eva_workbook = openpyxl.load_workbook(path_to_evaluation_file)
    group_worksheet = eva_workbook['group']
    eva_worksheet = eva_workbook['eva']
    meta_worksheet = eva_workbook['meta']
    students_worksheet = eva_workbook['students']

    group_col = []
    for col_item in list(group_worksheet.iter_cols())[0]:
        if col_item.value != "groupid":
            group_col.append(col_item.value)

    # get all students by students
    students = get_students_by_group(group_worksheet, students_worksheet)

    group_col = list(dict.fromkeys(group_col))

    # create a empty row for each group in the new evaluation
    
    j = 0
    for group in group_col:
        students_name = []
                # couple is [email, student_name]
        i=0
        for student_couple in students[str(group)]:
            students_name.append(student_couple[1])
            row_to_insert = new_row_generator(str(group), students_name[i], evaluation_name[j], eva_worksheet)
            date[0] = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            eva_worksheet.append(row_to_insert)
            i+=1
            j+=1
    eva_workbook.save(path_to_evaluation_file)

def copy_all_worksheet(copy_to, copy_from):
    for row in range(0, len(list(copy_from.iter_rows()))):
        for col in range(0, len(list(copy_from.iter_cols()))):
            copy_to.cell(row=row + 1, column=col +
                        1).value = copy_from.cell(row=row + 1, column=col + 1).value


def delete_project(email, projectName):

    path_to_current_project = "{}/{}/{}".format(base_directory, email, projectName)

    if os.path.exists(path_to_current_project):
        shutil.rmtree(path_to_current_project)