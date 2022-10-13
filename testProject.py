
from app import select_by_col_name, select_map_by_index, Permission, Project, json, db, FileLock, base_directory, home_directory
import os
import openpyxl
import shutil


def create_test_project(email, projectName, projectDescription):
    
    # get find sample roster and rubric
    path_to_sample_roster = "{}/sample_file/rosters/sample_roster.xlsx".format(home_directory)
    path_to_sample_json = "{}/sample_file/rubrics/information_processing/information_processing.json".format(home_directory)
        
    # create project folder
    path_to_current_user_project = "{}/{}/{}".format(base_directory, email, projectName)
    os.mkdir(path_to_current_user_project)

    path_to_student_file_stored = "{}/student.xlsx".format(path_to_current_user_project)
    shutil.copy(path_to_sample_roster, path_to_student_file_stored)
    path_to_json_file_stored = "{}/TW.json".format(path_to_current_user_project)
    shutil.copy(path_to_sample_json, path_to_json_file_stored)
    # creating evaluation doc based on grading criteria json file

    # copy student sheet to evaluation doc
    student_file_workbook = openpyxl.load_workbook(path_to_student_file_stored)
    student_file_worksheet = student_file_workbook['Sheet1']

    # create group file depending on student file
    list_of_group = select_by_col_name('group', student_file_worksheet)
    set_of_group = set(list_of_group)

    # Fixing a bug where a None element was found. Is this safe?
    set_of_group.discard(None)

    # create a group workbook
    path_to_group_file = "{}/group.xlsx".format(path_to_current_user_project)
    group_workbook = openpyxl.Workbook()
    group_file_worksheet = group_workbook.create_sheet('Sheet1')
    meta_file_worksheet = group_workbook.create_sheet('Sheet2')
    # all student information map
    student_map_list = []
    for student_index in range(2, len(list(student_file_worksheet.iter_rows())) + 1):
        student_map_list.append(select_map_by_index(
            student_index, student_file_worksheet))
    # insert group columns
    group_file_worksheet.cell(1, 1).value = 'groupid'
    meta_file_worksheet.cell(1, 1).value = 'groupid'
    meta_file_worksheet.cell(1, 2).value = 'metaid'
    start_index = 2
    max_num_students_pergroup = 0
    for group in set_of_group:
        group_file_worksheet.cell(start_index, 1).value = group
        student_emails = [x['Email']
                        for x in student_map_list if x['group'] == group]
        if len(student_emails) > max_num_students_pergroup:
            max_num_students_pergroup = len(student_emails)
        meta_file_worksheet.cell(start_index, 1).value = group
        meta_group = [x['meta']
                    for x in student_map_list if x['group'] == group][0]
        meta_file_worksheet.cell(start_index, 2).value = meta_group
        for insert_index in range(2, len(student_emails) + 2):
            group_file_worksheet.cell(
                start_index, insert_index).value = student_emails[insert_index - 2]
        start_index += 1
    for index in range(1, max_num_students_pergroup+1):
        group_file_worksheet.cell(1, 1+index).value = ("student" + str(index) )
    group_workbook.save(path_to_group_file)

    path_to_evaluation = "{}/evaluation.xlsx".format(
        path_to_current_user_project)
    evaluation_workbook = openpyxl.Workbook()
    evaluation_group = evaluation_workbook.create_sheet('group')
    evaluation_meta = evaluation_workbook.create_sheet('meta')
    evaluation_student = evaluation_workbook.create_sheet('students')
    copy_all_worksheet(evaluation_group, group_file_worksheet)
    copy_all_worksheet(evaluation_meta, meta_file_worksheet)
    copy_all_worksheet(evaluation_student, student_file_worksheet)
    # create EVA depending on the json file
    evaluation_eva = evaluation_workbook.create_sheet('eva')
    # open json file and load json
    myLock = FileLock(path_to_json_file_stored+'.lock', timeout=5)
    with myLock:
        with open(path_to_json_file_stored, 'r')as f:
            json_data = json.loads(f.read(), strict=False)
    # The group id, eva_name, date are defults
    tags_to_append = ['group_id', 'eva_name',
                    'owner', 'date', 'students']
    for category in json_data['category']:
        category_name = (category['name'])
        for section in category['section']:
            # instructors don't care about the text value, the text values will only be send to students.
            if section['type'] != 'text':
                value_to_append = "{}|{}".format(
                    category_name, section['name'])
                tags_to_append.append(value_to_append)
    tags_to_append.append("comment")
    tags_to_append.append("last_updates")
    evaluation_eva.append(tags_to_append)

    evaluation_workbook.save(path_to_evaluation)

    # create permission to owener himself
    project_id = "{}{}{}{}".format(
        email, email, projectName, 'full')
    self_permission = Permission(project_id=project_id, owner=email, shareTo=email,
                                project=projectName, status='full')
    db.session.add(self_permission)
    db.session.commit()

    # create the project in database
    project_to_add = Project(project_name=projectName, project_status='public',
                            owner=email, description=projectDescription)
    db.session.add(project_to_add)
    db.session.commit()


def copy_all_worksheet(copy_to, copy_from):
    for row in range(0, len(list(copy_from.iter_rows()))):
        for col in range(0, len(list(copy_from.iter_cols()))):
            copy_to.cell(row=row + 1, column=col +
                        1).value = copy_from.cell(row=row + 1, column=col + 1).value


def delete_project(email, projectName):


    project = Permission.query.filter_by(project_id=email+email+projectName+'full').first()
    permission_to_delete = Permission.query.filter_by(project=project.project).all()
    path_to_current_project = "{}/{}/{}".format(base_directory, email, project.project)

    if os.path.exists(path_to_current_project):
        shutil.rmtree(path_to_current_project)

        # after delete the folder, delete all the permissions that were send from the project
        for permission in permission_to_delete:
            db.session.delete(permission)
            db.session.commit()

        # delete the project in project table
        project_in_database = Project.query.filter_by(project_name=project.project, owner=project.owner).first()
        db.session.delete(project_in_database)
        db.session.commit()