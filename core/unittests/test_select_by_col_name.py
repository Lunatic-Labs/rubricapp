import sys
sys.path.append('..')
from functions import select_by_col_name
from core import app
import os
import openpyxl
import shutil
from flask import Flask
import random
from createTestProject import *
import unittest
import collections

def makeEvalNames(evalnameList):
    for i in range(40):
        evalnameList.insert(i,"EvalName" + str(i+2))

evalnameList = []
makeEvalNames(evalnameList)

class select_by_col_name_Test(unittest.TestCase):
   
    @classmethod
    def setUpClass(cls):
        
        basedirectoryAndmkuser()

        flask_app = app
        with flask_app.app_context():
            cls.projectName = "Test pName" + str(random.getrandbits(12)) + str(random.getrandbits(12)) + str(random.getrandbits(12))
            create_test_project("test@gmail.com", cls.projectName)
            createEvaluation("test@gmail.com", cls.projectName,evalnameList)

    @classmethod
    def tearDownClass(cls):    
        delete_project("test@gmail.com", cls.projectName)
    
    @classmethod
    def studentxlsx_sheet1_setup(cls):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, "test@gmail.com", cls.projectName)
        path_to_student_file_stored = "{}/student.xlsx".format(path_to_current_user_project)
        student_file_workbook = openpyxl.load_workbook(path_to_student_file_stored)
        return student_file_workbook['Sheet1']

    @classmethod
    def groupxlsx_sheet1_setup(cls):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, "test@gmail.com", cls.projectName)
        path_to_group_file_stored = "{}/group.xlsx".format(path_to_current_user_project)
        group_file_workbook = openpyxl.load_workbook(path_to_group_file_stored)
        return group_file_workbook['Sheet1']
    
    @classmethod
    def groupxlsx_sheet2_setup(cls):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, "test@gmail.com", cls.projectName)
        path_to_group_file_stored = "{}/group.xlsx".format(path_to_current_user_project)
        group_file_workbook = openpyxl.load_workbook(path_to_group_file_stored)
        return group_file_workbook['Sheet2']

    @classmethod
    def evaluationxlsx_eva_setup(cls):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, "test@gmail.com", cls.projectName)
        path_to_evaluation_file_stored = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_file_workbook = openpyxl.load_workbook(path_to_evaluation_file_stored)
        return evaluation_file_workbook['eva']
    
    @classmethod
    def evaluationxlsx_meta_setup(cls):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, "test@gmail.com", cls.projectName)
        path_to_evaluation_file_stored = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_file_workbook = openpyxl.load_workbook(path_to_evaluation_file_stored)
        return evaluation_file_workbook['meta']

    @classmethod
    def evaluationxlsx_group_setup(cls):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, "test@gmail.com", cls.projectName)
        path_to_evaluation_file_stored = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_file_workbook = openpyxl.load_workbook(path_to_evaluation_file_stored)
        return evaluation_file_workbook['group']

    @classmethod
    def evaluationxlsx_students_setup(cls):
        path_to_current_user_project = "{}/{}/{}".format(base_directory, "test@gmail.com", cls.projectName)
        path_to_evaluation_file_stored = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_file_workbook = openpyxl.load_workbook(path_to_evaluation_file_stored)
        return evaluation_file_workbook['students']
    
    def test_check_studentxlsx_student_col(self):
        student_file_worksheet = select_by_col_name_Test.studentxlsx_sheet1_setup()
        
        self.assertEqual(select_by_col_name("Student",student_file_worksheet),['Mccray, Maja', 'Guerrero, Fateh', 'Mcphee, Pearce', 'Michael, Olivia', 'Austin, Saarah', 'Velasquez, Rex', 'Coles, Rivka', 'Moody, Lyla', 'Morales, Conah', 'Cotton, Libbie', 'Ahmad, Rocky', 'Gardner, Shakira', 'Lopez, Cecilia', 'Aguilar, Sumayyah', 'Almond, Shae', 'Patton, Naima', "O'Moore, Kitty", 'Bowes, Aryan', 'Kirkpatrick, Arwen', 'Kendall, Jakob', 'Dowling, Safiyyah', 'Hull, Ciaran', 'Goddard, Rafael', 'Whittington, Aden', 'Mosley, Korban', 'Meyer, Sienna', 'Boyce, Bradlee', 'Patterson, Lilly', 'Wickens, Hebe', 'England, Kenzo', 'Parry, Sulayman', 'Baldwin, Ismail', 'Werner, Lillia', 'Emerson, Hiba', 'Casey, Gabriela', 'Johnston, Moshe', 'Thorpe, Shana', 'Lowry, Om', 'Lawson, Nikhil', 'Haney, Mara'])
        self.assertNotEqual(select_by_col_name("Student",student_file_worksheet),[])

    def test_check_studentxlsx_email_col(self):
        student_file_worksheet = select_by_col_name_Test.studentxlsx_sheet1_setup()
        
        self.assertEqual(select_by_col_name("Email",student_file_worksheet),['rubricapp-c0@mailinator.com', 'rubricapp-c1@mailinator.com', 'rubricapp-c2@mailinator.com', 'rubricapp-c3@mailinator.com', 'rubricapp-c4@mailinator.com', 'rubricapp-c5@mailinator.com', 'rubricapp-c6@mailinator.com', 'rubricapp-c7@mailinator.com', 'rubricapp-c8@mailinator.com', 'rubricapp-c9@mailinator.com', 'rubricapp-c10@mailinator.com', 'rubricapp-c11@mailinator.com', 'rubricapp-c12@mailinator.com', 'rubricapp-c13@mailinator.com', 'rubricapp-c14@mailinator.com', 'rubricapp-c15@mailinator.com', 'rubricapp-c16@mailinator.com', 'rubricapp-c17@mailinator.com', 'rubricapp-c18@mailinator.com', 'rubricapp-c19@mailinator.com', 'rubricapp-c20@mailinator.com', 'rubricapp-c21@mailinator.com', 'rubricapp-c22@mailinator.com', 'rubricapp-c23@mailinator.com', 'rubricapp-c24@mailinator.com', 'rubricapp-c25@mailinator.com', 'rubricapp-c26@mailinator.com', 'rubricapp-c27@mailinator.com', 'rubricapp-c28@mailinator.com', 'rubricapp-c29@mailinator.com', 'rubricapp-c30@mailinator.com', 'rubricapp-c31@mailinator.com', 'rubricapp-c32@mailinator.com', 'rubricapp-c33@mailinator.com', 'rubricapp-c34@mailinator.com', 'rubricapp-c35@mailinator.com', 'rubricapp-c36@mailinator.com', 'rubricapp-c37@mailinator.com', 'rubricapp-c38@mailinator.com', 'rubricapp-c39@mailinator.com'])
        self.assertNotEqual(select_by_col_name("Email",student_file_worksheet),[])
    
    def test_check_studentxlsx_group_col(self):
        student_file_worksheet = select_by_col_name_Test.studentxlsx_sheet1_setup()
        
        self.assertEqual(select_by_col_name("group",student_file_worksheet),['H', 'H', 'H', 'H', 'He', 'He', 'He', 'He', 'Li', 'Li', 'Li', 'Li', 'Be', 'Be', 'Be', 'Be', 'B', 'B', 'B', 'B', 'C', 'C', 'C', 'C', 'N', 'N', 'N', 'N', 'O', 'O', 'O', 'O', 'F', 'F', 'F', 'F', 'Ne', 'Ne', 'Ne', 'Ne'])
        self.assertNotEqual(select_by_col_name("group",student_file_worksheet),[])
    
    def test_check_studentxlsx_meta_col(self):
        student_file_worksheet = select_by_col_name_Test.studentxlsx_sheet1_setup()
        
        self.assertEqual(select_by_col_name("meta",student_file_worksheet),['a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'c', 'c', 'c', 'c'])
        self.assertNotEqual(select_by_col_name("meta",student_file_worksheet),[])
    
    def test_check_groupxlsx_groupid_col(self):
        group_file_worksheet1 = select_by_col_name_Test.groupxlsx_sheet1_setup()
        group_file_worksheet2 = select_by_col_name_Test.groupxlsx_sheet2_setup()
        
        self.assertNotEqual(select_by_col_name("groupid",group_file_worksheet1),[])
        self.assertNotEqual(select_by_col_name("groupid",group_file_worksheet2),[])

        self.assertEqual(select_by_col_name("groupid",group_file_worksheet1),['H', 'He', 'Li', 'Be', 'B', 'C', 'N', 'O', 'F', 'Ne'])
        self.assertEqual(select_by_col_name("groupid",group_file_worksheet1),['H', 'He', 'Li', 'Be', 'B', 'C', 'N', 'O', 'F', 'Ne'])
    
    def test_check_groupxlsx_metaid_col(self):
        group_file_worksheet2 = select_by_col_name_Test.groupxlsx_sheet2_setup()

        self.assertNotEqual(select_by_col_name("metaid",group_file_worksheet2),[])
        self.assertEqual(select_by_col_name("metaid",group_file_worksheet2),['a', 'a', 'a', 'a', 'a', 'b', 'b', 'b', 'b', 'c'])

    def test_check_groupxlsx_student1_col(self):
        group_file_worksheet1 = select_by_col_name_Test.groupxlsx_sheet1_setup()

        self.assertNotEqual(select_by_col_name("student1",group_file_worksheet1),[])
        self.assertEqual(select_by_col_name("student1",group_file_worksheet1),['rubricapp-c0@mailinator.com', 'rubricapp-c4@mailinator.com', 'rubricapp-c8@mailinator.com', 'rubricapp-c12@mailinator.com', 'rubricapp-c16@mailinator.com', 'rubricapp-c20@mailinator.com', 'rubricapp-c24@mailinator.com', 'rubricapp-c28@mailinator.com', 'rubricapp-c32@mailinator.com', 'rubricapp-c36@mailinator.com'])
    
    def test_check_groupxlsx_student2_col(self):
        group_file_worksheet1 = select_by_col_name_Test.groupxlsx_sheet1_setup()

        self.assertNotEqual(select_by_col_name("student2",group_file_worksheet1),[])
        self.assertEqual(select_by_col_name("student2",group_file_worksheet1),['rubricapp-c1@mailinator.com', 'rubricapp-c5@mailinator.com', 'rubricapp-c9@mailinator.com', 'rubricapp-c13@mailinator.com', 'rubricapp-c17@mailinator.com', 'rubricapp-c21@mailinator.com', 'rubricapp-c25@mailinator.com', 'rubricapp-c29@mailinator.com', 'rubricapp-c33@mailinator.com', 'rubricapp-c37@mailinator.com'])

    def test_check_groupxlsx_student3_col(self):
        group_file_worksheet1 = select_by_col_name_Test.groupxlsx_sheet1_setup()

        self.assertNotEqual(select_by_col_name("student3",group_file_worksheet1),[])
        self.assertEqual(select_by_col_name("student3",group_file_worksheet1),['rubricapp-c2@mailinator.com', 'rubricapp-c6@mailinator.com', 'rubricapp-c10@mailinator.com', 'rubricapp-c14@mailinator.com', 'rubricapp-c18@mailinator.com', 'rubricapp-c22@mailinator.com', 'rubricapp-c26@mailinator.com', 'rubricapp-c30@mailinator.com', 'rubricapp-c34@mailinator.com', 'rubricapp-c38@mailinator.com'])
    
    def test_check_groupxlsx_student4_col(self):
        group_file_worksheet1 = select_by_col_name_Test.groupxlsx_sheet1_setup()

        self.assertNotEqual(select_by_col_name("student4",group_file_worksheet1),[])
        self.assertEqual(select_by_col_name("student4",group_file_worksheet1),['rubricapp-c3@mailinator.com', 'rubricapp-c7@mailinator.com', 'rubricapp-c11@mailinator.com', 'rubricapp-c15@mailinator.com', 'rubricapp-c19@mailinator.com', 'rubricapp-c23@mailinator.com', 'rubricapp-c27@mailinator.com', 'rubricapp-c31@mailinator.com', 'rubricapp-c35@mailinator.com', 'rubricapp-c39@mailinator.com'])
    
    def test_check_evaluationxlsx_groupid_col(self):
        evaluation_file_worksheet4 = select_by_col_name_Test.evaluationxlsx_eva_setup()

        self.assertNotEqual(select_by_col_name("group_id",evaluation_file_worksheet4),[])
        self.assertEqual(select_by_col_name("group_id",evaluation_file_worksheet4),['H', 'H', 'H', 'H', 'He', 'He', 'He', 'He', 'Li', 'Li', 'Li', 'Li', 'Be', 'Be', 'Be', 'Be', 'B', 'B', 'B', 'B', 'C', 'C', 'C', 'C', 'N', 'N', 'N', 'N', 'O', 'O', 'O', 'O', 'F', 'F', 'F', 'F', 'Ne', 'Ne', 'Ne', 'Ne'])

    def test_check_evaluationxlsx_evaname_col(self):
        evaluation_file_worksheet4 = select_by_col_name_Test.evaluationxlsx_eva_setup()

        self.assertNotEqual(select_by_col_name("eva_name",evaluation_file_worksheet4),[])
        self.assertEqual(select_by_col_name("eva_name",evaluation_file_worksheet4),['EvalName2', 'EvalName3', 'EvalName4', 'EvalName5', 'EvalName6', 'EvalName7', 'EvalName8', 'EvalName9', 'EvalName10', 'EvalName11', 'EvalName12', 'EvalName13', 'EvalName14', 'EvalName15', 'EvalName16', 'EvalName17', 'EvalName18', 'EvalName19', 'EvalName20', 'EvalName21', 'EvalName22', 'EvalName23', 'EvalName24', 'EvalName25', 'EvalName26', 'EvalName27', 'EvalName28', 'EvalName29', 'EvalName30', 'EvalName31', 'EvalName32', 'EvalName33', 'EvalName34', 'EvalName35', 'EvalName36', 'EvalName37', 'EvalName38', 'EvalName39', 'EvalName40', 'EvalName41'])
    
    def test_check_evaluationxlsx_owner_col(self):
        evaluation_file_worksheet4 = select_by_col_name_Test.evaluationxlsx_eva_setup()

        self.assertNotEqual(select_by_col_name("owner",evaluation_file_worksheet4),[])
        self.assertEqual(select_by_col_name("owner",evaluation_file_worksheet4),['Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest'])

    def test_check_evaluationxlsx_date_col(self): 
        evaluation_file_worksheet4 = select_by_col_name_Test.evaluationxlsx_eva_setup()
        for i in range(0,40):
            self.assertIn(date[0],select_by_col_name("date",evaluation_file_worksheet4)[i])

    def test_check_evaluationxlsx_students_col(self):
        evaluation_file_worksheet4 = select_by_col_name_Test.evaluationxlsx_eva_setup()

        self.assertNotEqual(select_by_col_name("students",evaluation_file_worksheet4),[])
        self.assertEqual(select_by_col_name("students",evaluation_file_worksheet4),['Mccray, Maja', 'Guerrero, Fateh', 'Mcphee, Pearce', 'Michael, Olivia', 'Austin, Saarah', 'Velasquez, Rex', 'Coles, Rivka', 'Moody, Lyla', 'Morales, Conah', 'Cotton, Libbie', 'Ahmad, Rocky', 'Gardner, Shakira', 'Lopez, Cecilia', 'Aguilar, Sumayyah', 'Almond, Shae', 'Patton, Naima', "O'Moore, Kitty", 'Bowes, Aryan', 'Kirkpatrick, Arwen', 'Kendall, Jakob', 'Dowling, Safiyyah', 'Hull, Ciaran', 'Goddard, Rafael', 'Whittington, Aden', 'Mosley, Korban', 'Meyer, Sienna', 'Boyce, Bradlee', 'Patterson, Lilly', 'Wickens, Hebe', 'England, Kenzo', 'Parry, Sulayman', 'Baldwin, Ismail', 'Werner, Lillia', 'Emerson, Hiba', 'Casey, Gabriela', 'Johnston, Moshe', 'Thorpe, Shana', 'Lowry, Om', 'Lawson, Nikhil', 'Haney, Mara'])

    def test_check_evaluationxlsx_F_col(self):
        evaluation_file_worksheet4 = select_by_col_name_Test.evaluationxlsx_eva_setup()

        self.assertNotEqual(select_by_col_name("Evaluating|Determined the significance or relevance of information/data needed for the task.",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("Evaluating|Determined the significance or relevance of information/data needed for the task.",evaluation_file_worksheet4),[' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' '])

    def test_check_evaluationxlsx_G_col(self):
        evaluation_file_worksheet4 = select_by_col_name_Test.evaluationxlsx_eva_setup()

        self.assertNotEqual(select_by_col_name("Evaluating|Observed Characteristics",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("Evaluating|Observed Characteristics",evaluation_file_worksheet4),[' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' '])

    def test_check_evaluationxlsx_H_col(self):
        evaluation_file_worksheet4 = select_by_col_name_Test.evaluationxlsx_eva_setup()

        self.assertNotEqual(select_by_col_name("Interpreting|Provided meaning to data, made inferences and predictions from data, or extracted patterns from data.",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("Interpreting|Provided meaning to data, made inferences and predictions from data, or extracted patterns from data.",evaluation_file_worksheet4),[' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' '])

    def test_check_evaluationxlsx_I_col(self):
        evaluation_file_worksheet4 = select_by_col_name_Test.evaluationxlsx_eva_setup()

        self.assertNotEqual(select_by_col_name("Interpreting|Observed Characteristics",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("Interpreting|Observed Characteristics",evaluation_file_worksheet4),[' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' '])

    def test_check_evaluationxlsx_J_col(self):
        evaluation_file_worksheet4 = select_by_col_name_Test.evaluationxlsx_eva_setup()

        self.assertNotEqual(select_by_col_name("Manipulating or Transforming (Extent)|Observed Characteristics.",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("Manipulating or Transforming (Extent)|Observed Characteristics",evaluation_file_worksheet4),[' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' '])

    def test_check_evaluationxlsx_K_col(self):
        evaluation_file_worksheet4 = select_by_col_name_Test.evaluationxlsx_eva_setup()

        self.assertNotEqual(select_by_col_name("Manipulating or Transforming (Extent)|Observed Characteristics",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("Manipulating or Transforming (Extent)|Observed Characteristics",evaluation_file_worksheet4),[' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' '])

    def test_check_evaluationxlsx_L_col(self):
        evaluation_file_worksheet4 = select_by_col_name_Test.evaluationxlsx_eva_setup()

        self.assertNotEqual(select_by_col_name("Manipulating or Transforming (Accuracy)|Converted information/data from one form to another.",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("Manipulating or Transforming (Accuracy)|Converted information/data from one form to another.",evaluation_file_worksheet4),[' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' '])
    
    def test_check_evaluationxlsx_M_col(self):
        evaluation_file_worksheet4 = select_by_col_name_Test.evaluationxlsx_eva_setup()

        self.assertNotEqual(select_by_col_name("Manipulating or Transforming (Accuracy)|Observed Characteristics",evaluation_file_worksheet4),["Test"])
        self.assertEqual(select_by_col_name("Manipulating or Transforming (Accuracy)|Observed Characteristics",evaluation_file_worksheet4),[' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' '])

    def test_check_evaluationxlsx_O_col(self):
        evaluation_file_worksheet4 = select_by_col_name_Test.evaluationxlsx_eva_setup()

        self.assertNotEqual(select_by_col_name("last_updates",evaluation_file_worksheet4),[])
        self.assertEqual(select_by_col_name("last_updates",evaluation_file_worksheet4),['Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest', 'Guest'])

    def test_check_evaluationsxlsx_meta_groupid_col(self):
        evaluation_file_worksheet = select_by_col_name_Test.evaluationxlsx_meta_setup()
        self.assertNotEqual(select_by_col_name("groupid",evaluation_file_worksheet),[])
        self.assertEqual(select_by_col_name("groupid",evaluation_file_worksheet),['H', 'He', 'Li', 'Be', 'B', 'C', 'N', 'O', 'F', 'Ne'])

    def test_check_evaluationsxlsx_meta_metaid_col(self):
        evaluation_file_worksheet = select_by_col_name_Test.evaluationxlsx_meta_setup()

        self.assertNotEqual(select_by_col_name("metaid",evaluation_file_worksheet),[])
        self.assertEqual(select_by_col_name("metaid",evaluation_file_worksheet),['a', 'a', 'a', 'a', 'a', 'b', 'b', 'b', 'b', 'c'])

    def test_check_evaluationsxlsx_group_groupid_col(self):
        evaluation_file_worksheet = select_by_col_name_Test.evaluationxlsx_group_setup()
        self.assertNotEqual(select_by_col_name("groupid",evaluation_file_worksheet),[])
        self.assertEqual(select_by_col_name("groupid",evaluation_file_worksheet),['H', 'He', 'Li', 'Be', 'B', 'C', 'N', 'O', 'F', 'Ne'])

    def test_check_evaluationsxlsx_group_student1_col(self):
        evaluation_file_worksheet = select_by_col_name_Test.evaluationxlsx_group_setup()

        self.assertNotEqual(select_by_col_name("student1",evaluation_file_worksheet),[])
        self.assertEqual(select_by_col_name("student1",evaluation_file_worksheet),['rubricapp-c0@mailinator.com', 'rubricapp-c4@mailinator.com', 'rubricapp-c8@mailinator.com', 'rubricapp-c12@mailinator.com', 'rubricapp-c16@mailinator.com', 'rubricapp-c20@mailinator.com', 'rubricapp-c24@mailinator.com', 'rubricapp-c28@mailinator.com', 'rubricapp-c32@mailinator.com', 'rubricapp-c36@mailinator.com'])

    def test_check_evaluationsxlsx_group_student2_col(self):
        evaluation_file_worksheet = select_by_col_name_Test.evaluationxlsx_group_setup()
        self.assertNotEqual(select_by_col_name("student2",evaluation_file_worksheet),[])
        self.assertEqual(select_by_col_name("student2",evaluation_file_worksheet),['rubricapp-c1@mailinator.com', 'rubricapp-c5@mailinator.com', 'rubricapp-c9@mailinator.com', 'rubricapp-c13@mailinator.com', 'rubricapp-c17@mailinator.com', 'rubricapp-c21@mailinator.com', 'rubricapp-c25@mailinator.com', 'rubricapp-c29@mailinator.com', 'rubricapp-c33@mailinator.com', 'rubricapp-c37@mailinator.com'])

    def test_check_evaluationsxlsx_group_student3_col(self):
        evaluation_file_worksheet = select_by_col_name_Test.evaluationxlsx_group_setup()

        self.assertNotEqual(select_by_col_name("student3",evaluation_file_worksheet),[])
        self.assertEqual(select_by_col_name("student3",evaluation_file_worksheet),['rubricapp-c2@mailinator.com', 'rubricapp-c6@mailinator.com', 'rubricapp-c10@mailinator.com', 'rubricapp-c14@mailinator.com', 'rubricapp-c18@mailinator.com', 'rubricapp-c22@mailinator.com', 'rubricapp-c26@mailinator.com', 'rubricapp-c30@mailinator.com', 'rubricapp-c34@mailinator.com', 'rubricapp-c38@mailinator.com'])

    def test_check_evaluationsxlsx_group_student4_col(self):
        evaluation_file_worksheet = select_by_col_name_Test.evaluationxlsx_group_setup()
        self.assertNotEqual(select_by_col_name("student4",evaluation_file_worksheet),[])
        self.assertEqual(select_by_col_name("student4",evaluation_file_worksheet),['rubricapp-c3@mailinator.com', 'rubricapp-c7@mailinator.com', 'rubricapp-c11@mailinator.com', 'rubricapp-c15@mailinator.com', 'rubricapp-c19@mailinator.com', 'rubricapp-c23@mailinator.com', 'rubricapp-c27@mailinator.com', 'rubricapp-c31@mailinator.com', 'rubricapp-c35@mailinator.com', 'rubricapp-c39@mailinator.com'])

    def test_check_evaluationsxlsx_students_student_col(self):
        evaluation_file_worksheet = select_by_col_name_Test.evaluationxlsx_students_setup()
        self.assertNotEqual(select_by_col_name("Student",evaluation_file_worksheet),[])
        self.assertEqual(select_by_col_name("Student",evaluation_file_worksheet),['Mccray, Maja', 'Guerrero, Fateh', 'Mcphee, Pearce', 'Michael, Olivia', 'Austin, Saarah', 'Velasquez, Rex', 'Coles, Rivka', 'Moody, Lyla', 'Morales, Conah', 'Cotton, Libbie', 'Ahmad, Rocky', 'Gardner, Shakira', 'Lopez, Cecilia', 'Aguilar, Sumayyah', 'Almond, Shae', 'Patton, Naima', "O'Moore, Kitty", 'Bowes, Aryan', 'Kirkpatrick, Arwen', 'Kendall, Jakob', 'Dowling, Safiyyah', 'Hull, Ciaran', 'Goddard, Rafael', 'Whittington, Aden', 'Mosley, Korban', 'Meyer, Sienna', 'Boyce, Bradlee', 'Patterson, Lilly', 'Wickens, Hebe', 'England, Kenzo', 'Parry, Sulayman', 'Baldwin, Ismail', 'Werner, Lillia', 'Emerson, Hiba', 'Casey, Gabriela', 'Johnston, Moshe', 'Thorpe, Shana', 'Lowry, Om', 'Lawson, Nikhil', 'Haney, Mara'])

    def test_check_evaluationsxlsx_students_group_col(self):
        evaluation_file_worksheet = select_by_col_name_Test.evaluationxlsx_students_setup()
        self.assertNotEqual(select_by_col_name("group",evaluation_file_worksheet),[])
        self.assertEqual(select_by_col_name("group",evaluation_file_worksheet),['H', 'H', 'H', 'H', 'He', 'He', 'He', 'He', 'Li', 'Li', 'Li', 'Li', 'Be', 'Be', 'Be', 'Be', 'B', 'B', 'B', 'B', 'C', 'C', 'C', 'C', 'N', 'N', 'N', 'N', 'O', 'O', 'O', 'O', 'F', 'F', 'F', 'F', 'Ne', 'Ne', 'Ne', 'Ne'])

    def test_check_evaluationsxlsx_students_email_col(self):
        evaluation_file_worksheet = select_by_col_name_Test.evaluationxlsx_students_setup()
        self.assertNotEqual(select_by_col_name("Email",evaluation_file_worksheet),[])
        self.assertEqual(select_by_col_name("Email",evaluation_file_worksheet),['rubricapp-c0@mailinator.com', 'rubricapp-c1@mailinator.com', 'rubricapp-c2@mailinator.com', 'rubricapp-c3@mailinator.com', 'rubricapp-c4@mailinator.com', 'rubricapp-c5@mailinator.com', 'rubricapp-c6@mailinator.com', 'rubricapp-c7@mailinator.com', 'rubricapp-c8@mailinator.com', 'rubricapp-c9@mailinator.com', 'rubricapp-c10@mailinator.com', 'rubricapp-c11@mailinator.com', 'rubricapp-c12@mailinator.com', 'rubricapp-c13@mailinator.com', 'rubricapp-c14@mailinator.com', 'rubricapp-c15@mailinator.com', 'rubricapp-c16@mailinator.com', 'rubricapp-c17@mailinator.com', 'rubricapp-c18@mailinator.com', 'rubricapp-c19@mailinator.com', 'rubricapp-c20@mailinator.com', 'rubricapp-c21@mailinator.com', 'rubricapp-c22@mailinator.com', 'rubricapp-c23@mailinator.com', 'rubricapp-c24@mailinator.com', 'rubricapp-c25@mailinator.com', 'rubricapp-c26@mailinator.com', 'rubricapp-c27@mailinator.com', 'rubricapp-c28@mailinator.com', 'rubricapp-c29@mailinator.com', 'rubricapp-c30@mailinator.com', 'rubricapp-c31@mailinator.com', 'rubricapp-c32@mailinator.com', 'rubricapp-c33@mailinator.com', 'rubricapp-c34@mailinator.com', 'rubricapp-c35@mailinator.com', 'rubricapp-c36@mailinator.com', 'rubricapp-c37@mailinator.com', 'rubricapp-c38@mailinator.com', 'rubricapp-c39@mailinator.com'])

    def test_check_evaluationsxlsx_students_meta_col(self):
        evaluation_file_worksheet = select_by_col_name_Test.evaluationxlsx_students_setup()
        self.assertNotEqual(select_by_col_name("meta",evaluation_file_worksheet),[])
        self.assertEqual(select_by_col_name("meta",evaluation_file_worksheet),['a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'b', 'c', 'c', 'c', 'c'])
